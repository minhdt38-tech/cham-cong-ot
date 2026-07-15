# -*- coding: utf-8 -*-
"""
Cham Cong OT - Web Server
Chay: python server.py
Truy cap: http://localhost:3000
Yeu cau: Python 3.8+ (khong can cai them thu vien)
"""

import http.server
import socketserver
import json
import sqlite3
import hashlib
import os
import mimetypes
import uuid
import io
import re
import threading
import time
import base64
import zipfile
import unicodedata
import xml.etree.ElementTree as ET
import urllib.request as _urlreq
import urllib.parse
from urllib.parse import urlparse, parse_qs
from datetime import datetime, timedelta, timezone

VN_TZ = timezone(timedelta(hours=7))

try:
    from PIL import Image as PilImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import boto3
    from botocore.client import Config
    BOTO3_AVAILABLE = True
except ImportError:
    BOTO3_AVAILABLE = False

PORT = int(os.environ.get('PORT', 3000))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# DATA_DIR: thu muc luu DB + uploads
# - Local: cung thu muc voi server.py
# - Railway/cloud: dat bien moi truong DATA_DIR=/data (persistent volume)
DATA_DIR    = os.environ.get('DATA_DIR', BASE_DIR)
UPLOADS_DIR = os.path.join(DATA_DIR, 'uploads')
PUBLIC_DIR  = os.path.join(BASE_DIR, 'public')
DB_PATH     = os.path.join(DATA_DIR, 'cham_cong.db')

os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(PUBLIC_DIR,  exist_ok=True)

# R2 / S3-compatible storage
R2_ENDPOINT    = os.environ.get('R2_ENDPOINT_URL', '')
R2_ACCESS_KEY  = os.environ.get('R2_ACCESS_KEY_ID', '')
R2_SECRET_KEY  = os.environ.get('R2_SECRET_ACCESS_KEY', '')
R2_BUCKET      = os.environ.get('R2_BUCKET_NAME', '')
R2_ENABLED     = bool(BOTO3_AVAILABLE and R2_ENDPOINT and R2_ACCESS_KEY and R2_SECRET_KEY and R2_BUCKET)

def get_r2_client():
    return boto3.client(
        's3',
        endpoint_url=R2_ENDPOINT,
        aws_access_key_id=R2_ACCESS_KEY,
        aws_secret_access_key=R2_SECRET_KEY,
        config=Config(signature_version='s3v4'),
        region_name='auto'
    )

def r2_upload(file_bytes, key, content_type='application/octet-stream'):
    """Upload bytes len R2, tra ve True neu thanh cong"""
    try:
        client = get_r2_client()
        client.put_object(
            Bucket=R2_BUCKET,
            Key=key,
            Body=file_bytes,
            ContentType=content_type
        )
        return True
    except Exception as e:
        print(f'[R2 upload error] {e}')
        return False

def r2_delete(key):
    """Xoa object tren R2"""
    try:
        client = get_r2_client()
        client.delete_object(Bucket=R2_BUCKET, Key=key)
    except Exception as e:
        print(f'[R2 delete error] {e}')

def r2_presigned_url(key, expires=3600):
    """Tao presigned URL de download/preview file, het han sau `expires` giay"""
    try:
        client = get_r2_client()
        return client.generate_presigned_url(
            'get_object',
            Params={'Bucket': R2_BUCKET, 'Key': key},
            ExpiresIn=expires
        )
    except Exception as e:
        print(f'[R2 presign error] {e}')
        return None


def backup_db_to_r2():
    """Chup nhanh toan bo cham_cong.db (an toan voi WAL) roi day len R2 duoi thu muc backups/."""
    if not R2_ENABLED:
        return {'ok': False, 'error': 'R2 chua duoc bat (thieu bien moi truong R2_*)'}
    tmp_path = os.path.join(DATA_DIR, f'_backup_tmp_{uuid.uuid4().hex}.db')
    try:
        src = sqlite3.connect(DB_PATH)
        dst = sqlite3.connect(tmp_path)
        with dst:
            src.backup(dst)
        src.close()
        dst.close()
        with open(tmp_path, 'rb') as f:
            data = f.read()
        ts  = datetime.now(VN_TZ).strftime('%Y-%m-%d_%H%M%S')
        key = f'backups/cham_cong_{ts}.db'
        ok  = r2_upload(data, key, 'application/octet-stream')
        if not ok:
            return {'ok': False, 'error': 'Upload len R2 that bai'}
        print(f'[Backup] OK -> {key} ({len(data)} bytes)')
        cleanup_old_backups()
        return {'ok': True, 'key': key, 'size': len(data), 'time': ts}
    except Exception as e:
        print(f'[Backup] error: {e}')
        return {'ok': False, 'error': str(e)}
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def list_r2_backups():
    if not R2_ENABLED:
        return []
    try:
        client = get_r2_client()
        resp = client.list_objects_v2(Bucket=R2_BUCKET, Prefix='backups/')
        items = [
            {'key': o['Key'], 'size': o['Size'], 'modified': o['LastModified'].isoformat()}
            for o in resp.get('Contents', [])
        ]
        items.sort(key=lambda x: x['key'], reverse=True)
        return items
    except Exception as e:
        print(f'[Backup list] error: {e}')
        return []


def cleanup_old_backups(keep=30):
    items = list_r2_backups()
    for item in items[keep:]:
        r2_delete(item['key'])


def backup_scheduler_loop():
    time.sleep(15)
    while True:
        backup_db_to_r2()
        time.sleep(24 * 60 * 60)


# In-memory sessions: { token: {userId, role, fullName} }
SESSIONS = {}
SESSION_LOCK = threading.Lock()


# --- DATABASE ---

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            username    TEXT UNIQUE NOT NULL,
            password    TEXT NOT NULL,
            full_name   TEXT NOT NULL,
            role        TEXT DEFAULT 'employee',
            department  TEXT DEFAULT '',
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS overtime_requests (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id      INTEGER NOT NULL,
            request_date TEXT NOT NULL,
            ot_type      TEXT NOT NULL,
            start_time   TEXT NOT NULL,
            end_time     TEXT NOT NULL,
            hours        REAL,
            reason       TEXT NOT NULL,
            image_path   TEXT,
            status       TEXT DEFAULT 'pending',
            manager_note TEXT,
            reviewed_by  INTEGER,
            reviewed_at  TEXT,
            created_at   TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS notifications (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            ot_id       INTEGER NOT NULL,
            title       TEXT NOT NULL,
            message     TEXT NOT NULL,
            is_read     INTEGER DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(ot_id)   REFERENCES overtime_requests(id)
        );
        CREATE TABLE IF NOT EXISTS doc_categories (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL,
            description TEXT DEFAULT '',
            color       TEXT DEFAULT '#1B2A4A',
            created_by  INTEGER,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS doc_types (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL,
            description TEXT DEFAULT '',
            created_by  INTEGER,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS documents (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            title       TEXT NOT NULL,
            description TEXT DEFAULT '',
            file_path   TEXT NOT NULL,
            file_name   TEXT NOT NULL,
            file_type   TEXT NOT NULL,
            file_size   INTEGER DEFAULT 0,
            category_id INTEGER,
            tags        TEXT DEFAULT '',
            uploaded_by INTEGER NOT NULL,
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(category_id) REFERENCES doc_categories(id),
            FOREIGN KEY(uploaded_by) REFERENCES users(id)
        );

        -- ── BỒI THƯỜNG, HỖ TRỢ ──────────────────────────────────────────
        CREATE TABLE IF NOT EXISTS bt_projects (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL,
            mo_ta       TEXT DEFAULT '',
            dia_diem    TEXT DEFAULT '',
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS bt_owners (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            ho_ten              TEXT NOT NULL,
            dia_chi_thuong_tru  TEXT DEFAULT '',
            so_cccd             TEXT DEFAULT '',
            ngay_cap_cccd       TEXT DEFAULT NULL,
            so_dien_thoai       TEXT DEFAULT '',
            ghi_chu             TEXT DEFAULT '',
            created_at          TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS bt_parcels (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id            INTEGER NOT NULL,
            owner_id              INTEGER,
            so_to                 TEXT DEFAULT '',
            so_thua               TEXT DEFAULT '',
            loai_dat              TEXT DEFAULT '',
            tong_dien_tich        REAL DEFAULT 0,
            dien_tich_thu_hoi     REAL DEFAULT 0,
            dien_tich_con_lai     REAL DEFAULT 0,
            dia_diem_thu_hoi      TEXT DEFAULT '',
            so_gcn                TEXT DEFAULT '',
            ngay_cap_gcn          TEXT DEFAULT NULL,
            ghi_chu               TEXT DEFAULT '',
            created_at            TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(project_id) REFERENCES bt_projects(id) ON DELETE CASCADE,
            FOREIGN KEY(owner_id)   REFERENCES bt_owners(id)
        );

        CREATE TABLE IF NOT EXISTS bt_records (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            parcel_id             INTEGER NOT NULL,
            trang_thai            TEXT DEFAULT 'Chưa thực hiện',
            ngay_hop_dan          TEXT DEFAULT NULL,
            dot_niem_yet_ck       TEXT DEFAULT '',
            so_quyet_dinh_pd      TEXT DEFAULT '',
            ngay_quyet_dinh_pd    TEXT DEFAULT NULL,
            so_tien_pd            REAL DEFAULT 0,
            so_quyet_dinh_thd     TEXT DEFAULT '',
            ngay_quyet_dinh_thd   TEXT DEFAULT NULL,
            tien_bt_dat           REAL DEFAULT 0,
            tien_bt_cay_trong     REAL DEFAULT 0,
            tien_csht             REAL DEFAULT 0,
            tien_ho_tro           REAL DEFAULT 0,
            tong_tien_nyck        REAL DEFAULT 0,
            so_tien_tam_ung       REAL DEFAULT 0,
            tong_tien_btht        REAL DEFAULT 0,
            so_tien_da_chi_tra    REAL DEFAULT 0,
            ngay_chi_tra          TEXT DEFAULT NULL,
            co_don_y_kien         INTEGER DEFAULT 0,
            noi_dung_y_kien       TEXT DEFAULT '',
            ghi_chu               TEXT DEFAULT '',
            created_at            TEXT DEFAULT (datetime('now','localtime')),
            updated_at            TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(parcel_id) REFERENCES bt_parcels(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS bt_status_config (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            ten     TEXT NOT NULL,
            mau     TEXT DEFAULT '#95a5a6',
            thu_tu  INTEGER DEFAULT 0
        );

        -- ── BỒI THƯỜNG v2: Chủ thể & Nhân khẩu ──────────────────────────
        CREATE TABLE IF NOT EXISTS bt_parties (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            loai_chu_the        TEXT DEFAULT 'Cá nhân',
            ho_ten              TEXT NOT NULL,
            gioi_tinh           TEXT DEFAULT '',
            ngay_sinh           TEXT DEFAULT NULL,
            so_cccd             TEXT DEFAULT '',
            ngay_cap_cccd       TEXT DEFAULT NULL,
            noi_cap_cccd        TEXT DEFAULT '',
            dia_chi_thuong_tru  TEXT DEFAULT '',
            so_dien_thoai       TEXT DEFAULT '',
            ghi_chu             TEXT DEFAULT '',
            custom_fields       TEXT DEFAULT '{}',
            created_at          TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS bt_household_members (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            chu_the_id          INTEGER NOT NULL,
            ho_ten              TEXT NOT NULL,
            quan_he_voi_chu_ho  TEXT DEFAULT '',
            ngay_sinh           TEXT DEFAULT NULL,
            so_cccd             TEXT DEFAULT '',
            ghi_chu             TEXT DEFAULT '',
            custom_fields       TEXT DEFAULT '{}',
            created_at          TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(chu_the_id) REFERENCES bt_parties(id) ON DELETE CASCADE
        );

        -- ── BỒI THƯỜNG v2: Thửa đất gốc & Bản đồ ────────────────────────
        CREATE TABLE IF NOT EXISTS bt_parcel_master (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            so_to_hien_hanh     TEXT DEFAULT '',
            so_thua_hien_hanh   TEXT DEFAULT '',
            dia_chi_vi_tri      TEXT DEFAULT '',
            toa_do_ranh_gioi    TEXT DEFAULT NULL,
            ghi_chu             TEXT DEFAULT '',
            custom_fields       TEXT DEFAULT '{}',
            created_at          TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS bt_parcel_owners (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            parcel_id       INTEGER NOT NULL,
            chu_the_id      INTEGER NOT NULL,
            vai_tro         TEXT DEFAULT 'Đại diện đứng tên',
            ty_le_so_huu    REAL DEFAULT NULL,
            FOREIGN KEY(parcel_id)  REFERENCES bt_parcels(id) ON DELETE CASCADE,
            FOREIGN KEY(chu_the_id) REFERENCES bt_parties(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS bt_maps (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id      INTEGER DEFAULT NULL,
            loai_ban_do     TEXT DEFAULT '',
            ten_ban_do      TEXT DEFAULT '',
            ngay_lap        TEXT DEFAULT NULL,
            don_vi_lap      TEXT DEFAULT '',
            ghi_chu         TEXT DEFAULT '',
            custom_fields   TEXT DEFAULT '{}',
            created_at      TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(project_id) REFERENCES bt_projects(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS bt_map_parcels (
            id                      INTEGER PRIMARY KEY AUTOINCREMENT,
            map_id                  INTEGER NOT NULL,
            parcel_master_id        INTEGER DEFAULT NULL,
            so_to_tren_ban_do       TEXT DEFAULT '',
            so_thua_tren_ban_do     TEXT DEFAULT '',
            dien_tich_tren_ban_do   REAL DEFAULT 0,
            FOREIGN KEY(map_id) REFERENCES bt_maps(id) ON DELETE CASCADE,
            FOREIGN KEY(parcel_master_id) REFERENCES bt_parcel_master(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS bt_map_files (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            map_id          INTEGER NOT NULL,
            file_path       TEXT NOT NULL,
            file_name       TEXT DEFAULT '',
            uploaded_at     TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(map_id) REFERENCES bt_maps(id) ON DELETE CASCADE
        );

        -- ── BỒI THƯỜNG v2: Tài sản trên đất ──────────────────────────────
        CREATE TABLE IF NOT EXISTS bt_assets (
            id                      INTEGER PRIMARY KEY AUTOINCREMENT,
            loai_tai_san_nhom       TEXT DEFAULT '',
            loai_tai_san_cu_the     TEXT DEFAULT '',
            chu_tai_san_id          INTEGER DEFAULT NULL,
            don_vi_tinh             TEXT DEFAULT '',
            so_luong_khoi_luong     REAL DEFAULT 0,
            thoi_diem_hinh_thanh    TEXT DEFAULT NULL,
            tinh_trang_phap_ly      TEXT DEFAULT 'Đúng quy định',
            ngay_kiem_dem           TEXT DEFAULT NULL,
            nguoi_kiem_dem          TEXT DEFAULT '',
            ghi_chu                 TEXT DEFAULT '',
            custom_fields           TEXT DEFAULT '{}',
            created_at              TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(chu_tai_san_id) REFERENCES bt_parties(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS bt_asset_parcels (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id    INTEGER NOT NULL,
            parcel_id   INTEGER NOT NULL,
            FOREIGN KEY(asset_id)  REFERENCES bt_assets(id) ON DELETE CASCADE,
            FOREIGN KEY(parcel_id) REFERENCES bt_parcels(id) ON DELETE CASCADE
        );

        -- ── BỒI THƯỜNG v2: Hồ sơ Hộ & Quyết định theo Thửa ──────────────
        CREATE TABLE IF NOT EXISTS bt_dossiers (
            id                                  INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id                          INTEGER NOT NULL,
            chu_the_id                          INTEGER NOT NULL,
            tien_thuong_tien_do                 REAL DEFAULT 0,
            tien_ho_tro_on_dinh_doi_song         REAL DEFAULT 0,
            tien_ho_tro_dao_tao_chuyen_doi_nghe  REAL DEFAULT 0,
            tien_ho_tro_tai_dinh_cu              REAL DEFAULT 0,
            tien_ho_tro_khac                     REAL DEFAULT 0,
            so_tien_tam_ung                      REAL DEFAULT 0,
            so_tien_da_chi_tra                   REAL DEFAULT 0,
            ngay_chi_tra                         TEXT DEFAULT NULL,
            co_don_y_kien                        INTEGER DEFAULT 0,
            noi_dung_y_kien                      TEXT DEFAULT '',
            ghi_chu                              TEXT DEFAULT '',
            custom_fields                        TEXT DEFAULT '{}',
            created_at                           TEXT DEFAULT (datetime('now','localtime')),
            updated_at                           TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(project_id) REFERENCES bt_projects(id) ON DELETE CASCADE,
            FOREIGN KEY(chu_the_id) REFERENCES bt_parties(id)
        );

        CREATE TABLE IF NOT EXISTS bt_parcel_decisions (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            ho_so_ho_id         INTEGER NOT NULL,
            parcel_id           INTEGER NOT NULL,
            so_quyet_dinh_pd    TEXT DEFAULT '',
            ngay_quyet_dinh_pd  TEXT DEFAULT NULL,
            tien_bt_dat         REAL DEFAULT 0,
            tien_bt_tai_san     REAL DEFAULT 0,
            trang_thai          TEXT DEFAULT 'Chưa thực hiện',
            ghi_chu             TEXT DEFAULT '',
            created_at          TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(ho_so_ho_id) REFERENCES bt_dossiers(id) ON DELETE CASCADE,
            FOREIGN KEY(parcel_id)   REFERENCES bt_parcels(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS bt_dossier_persons (
            id                            INTEGER PRIMARY KEY AUTOINCREMENT,
            ho_so_ho_id                   INTEGER NOT NULL,
            chu_the_id                    INTEGER DEFAULT NULL,
            nhan_khau_id                  INTEGER DEFAULT NULL,
            la_doi_tuong_sxnn_truc_tiep   INTEGER DEFAULT 0,
            ghi_chu                       TEXT DEFAULT '',
            FOREIGN KEY(ho_so_ho_id)  REFERENCES bt_dossiers(id) ON DELETE CASCADE,
            FOREIGN KEY(chu_the_id)   REFERENCES bt_parties(id) ON DELETE CASCADE,
            FOREIGN KEY(nhan_khau_id) REFERENCES bt_household_members(id) ON DELETE CASCADE
        );
        """)
        # Migrations
        cols = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
        if 'can_upload_docs' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN can_upload_docs INTEGER DEFAULT 0")
        if 'can_view_docs' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN can_view_docs INTEGER DEFAULT 1")
        if 'can_download_docs' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN can_download_docs INTEGER DEFAULT 1")
        if 'can_edit_docs' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN can_edit_docs INTEGER DEFAULT 0")
        if 'can_delete_docs' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN can_delete_docs INTEGER DEFAULT 0")
        doc_cols = [r[1] for r in db.execute("PRAGMA table_info(documents)").fetchall()]
        if 'notes' not in doc_cols:
            db.execute("ALTER TABLE documents ADD COLUMN notes TEXT DEFAULT ''")
        if 'doc_number' not in doc_cols:
            db.execute("ALTER TABLE documents ADD COLUMN doc_number TEXT DEFAULT ''")
        if 'issued_date' not in doc_cols:
            db.execute("ALTER TABLE documents ADD COLUMN issued_date TEXT DEFAULT NULL")
        if 'issuer' not in doc_cols:
            db.execute("ALTER TABLE documents ADD COLUMN issuer TEXT DEFAULT ''")
        if 'doc_type_id' not in doc_cols:
            db.execute("ALTER TABLE documents ADD COLUMN doc_type_id INTEGER DEFAULT NULL")
        cat_cols = [r[1] for r in db.execute("PRAGMA table_info(doc_categories)").fetchall()]
        if 'parent_id' not in cat_cols:
            db.execute("ALTER TABLE doc_categories ADD COLUMN parent_id INTEGER DEFAULT NULL")

        # Bồi thường v2: mở rộng bt_projects (Dự án có ranh giới không gian)
        proj_cols = [r[1] for r in db.execute("PRAGMA table_info(bt_projects)").fetchall()]
        if 'chu_dau_tu' not in proj_cols:
            db.execute("ALTER TABLE bt_projects ADD COLUMN chu_dau_tu TEXT DEFAULT ''")
        if 'dien_tich_du_an' not in proj_cols:
            db.execute("ALTER TABLE bt_projects ADD COLUMN dien_tich_du_an REAL DEFAULT 0")
        if 'ranh_gioi_du_an' not in proj_cols:
            db.execute("ALTER TABLE bt_projects ADD COLUMN ranh_gioi_du_an TEXT DEFAULT NULL")
        if 'ngay_thong_bao_thu_hoi' not in proj_cols:
            db.execute("ALTER TABLE bt_projects ADD COLUMN ngay_thong_bao_thu_hoi TEXT DEFAULT NULL")
        if 'ngay_bat_dau' not in proj_cols:
            db.execute("ALTER TABLE bt_projects ADD COLUMN ngay_bat_dau TEXT DEFAULT NULL")
        if 'ngay_ket_thuc_du_kien' not in proj_cols:
            db.execute("ALTER TABLE bt_projects ADD COLUMN ngay_ket_thuc_du_kien TEXT DEFAULT NULL")
        if 'custom_fields' not in proj_cols:
            db.execute("ALTER TABLE bt_projects ADD COLUMN custom_fields TEXT DEFAULT '{}'")
        if 'kinh_tuyen_truc' not in proj_cols:
            # Kinh tuyến trục VN-2000 (độ) của tỉnh nơi dự án tọa lạc — dùng để quy đổi tọa độ VN-2000
            # (X-Bắc, Y-Đông) sang kinh độ/vĩ độ WGS84 khi hiển thị trên bản đồ nền thực (OpenStreetMap).
            db.execute("ALTER TABLE bt_projects ADD COLUMN kinh_tuyen_truc REAL DEFAULT NULL")

        # Bồi thường v2: mở rộng bt_parcels (Thửa đất theo dự án -> tham chiếu thửa gốc)
        parcel_cols = [r[1] for r in db.execute("PRAGMA table_info(bt_parcels)").fetchall()]
        if 'parcel_master_id' not in parcel_cols:
            db.execute("ALTER TABLE bt_parcels ADD COLUMN parcel_master_id INTEGER DEFAULT NULL")
        if 'nguon_goc_su_dung' not in parcel_cols:
            db.execute("ALTER TABLE bt_parcels ADD COLUMN nguon_goc_su_dung TEXT DEFAULT ''")
        if 'custom_fields' not in parcel_cols:
            db.execute("ALTER TABLE bt_parcels ADD COLUMN custom_fields TEXT DEFAULT '{}'")

        # Bồi thường v2.1: Bản đồ GPMB là nguồn sinh Thửa đất
        map_parcel_cols = [r[1] for r in db.execute("PRAGMA table_info(bt_map_parcels)").fetchall()]
        if 'parcel_id' not in map_parcel_cols:
            db.execute("ALTER TABLE bt_map_parcels ADD COLUMN parcel_id INTEGER DEFAULT NULL")
        if 'dien_tich_thu_hoi_tren_ban_do' not in map_parcel_cols:
            db.execute("ALTER TABLE bt_map_parcels ADD COLUMN dien_tich_thu_hoi_tren_ban_do REAL DEFAULT 0")
        if 'toa_do' not in map_parcel_cols:
            db.execute("ALTER TABLE bt_map_parcels ADD COLUMN toa_do TEXT DEFAULT NULL")

        # Bồi thường v2.2: tab Bản đồ (chung) — nhóm các mảnh bản đồ lại thành đầu mục tùy chỉnh
        # (VD "Mảnh bản đồ xã Chương Dương" gồm nhiều tờ) — nhãn tự do, gom nhóm ở phía UI theo giá trị trùng nhau.
        map_cols = [r[1] for r in db.execute("PRAGMA table_info(bt_maps)").fetchall()]
        if 'nhom_ban_do' not in map_cols:
            db.execute("ALTER TABLE bt_maps ADD COLUMN nhom_ban_do TEXT DEFAULT ''")

        # Bồi thường v2.3: Bản đồ mốc giới GPMB — khai báo mảnh trích đo GPMB liên quan (JSON mảng id)
        # và hướng tính diện tích thu hồi (phần nằm TRONG polygon là 'thu_hoi' hay 'con_lai') để tự
        # tính giao hình học (shapely) giữa mốc giới và từng thửa trên các mảnh GPMB đã liên kết.
        if 'mocgioi_manh_ids' not in map_cols:
            db.execute("ALTER TABLE bt_maps ADD COLUMN mocgioi_manh_ids TEXT DEFAULT ''")
        if 'mocgioi_trong_la' not in map_cols:
            db.execute("ALTER TABLE bt_maps ADD COLUMN mocgioi_trong_la TEXT DEFAULT ''")

        row = db.execute("SELECT id FROM users WHERE username='admin'").fetchone()
        if not row:
            pwd = hash_password('admin123')
            db.execute(
                "INSERT INTO users (username,password,full_name,role) VALUES (?,?,?,?)",
                ('admin', pwd, 'Quan ly', 'manager')
            )
        # Seed trạng thái mặc định nếu chưa có
        cnt = db.execute("SELECT COUNT(*) FROM bt_status_config").fetchone()[0]
        if cnt == 0:
            defaults = [
                ('Chưa thực hiện',                      '#95a5a6', 0),
                ('Đã họp dân',                           '#3498db', 1),
                ('Đã ban hành TBTH',                     '#9b59b6', 2),
                ('Đã kiểm kê / Niêm yết công khai',     '#f39c12', 3),
                ('Đã lập PA bồi thường, hỗ trợ',        '#e67e22', 4),
                ('Đã nhận tiền bồi thường, hỗ trợ',     '#27ae60', 5),
                ('Đã ban hành QĐ thu hồi đất',           '#1abc9c', 6),
                ('Vướng mắc / Khiếu nại',                '#e74c3c', 99),
            ]
            db.executemany(
                "INSERT INTO bt_status_config (ten, mau, thu_tu) VALUES (?,?,?)",
                defaults
            )
        db.commit()


def compress_image(file_bytes, ext):
    if not PIL_AVAILABLE:
        return file_bytes, ext
    try:
        img = PilImage.open(io.BytesIO(file_bytes))
        if img.mode in ('RGBA', 'P', 'LA'):
            bg = PilImage.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            bg.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
            img = bg
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        max_px = 1920
        w, h = img.size
        if w > max_px or h > max_px:
            ratio = min(max_px / w, max_px / h)
            img = img.resize((int(w * ratio), int(h * ratio)), PilImage.LANCZOS)
        out = io.BytesIO()
        img.save(out, format='JPEG', quality=75, optimize=True)
        return out.getvalue(), '.jpg'
    except Exception:
        return file_bytes, ext


def hash_password(pw):
    salt = 'chamcong_salt_2024'
    return hashlib.sha256((salt + pw).encode('utf-8')).hexdigest()


def check_password(pw, hashed):
    return hash_password(pw) == hashed


def rows_to_list(rows):
    return [dict(r) for r in rows]


def generate_temp_password():
    import random, string
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=8))


# --- SESSION ---

def create_session(user_id, role, full_name):
    token = str(uuid.uuid4())
    with SESSION_LOCK:
        SESSIONS[token] = {'userId': user_id, 'role': role, 'fullName': full_name}
    return token


def get_session(token):
    if not token:
        return None
    with SESSION_LOCK:
        return SESSIONS.get(token)


def delete_session(token):
    with SESSION_LOCK:
        SESSIONS.pop(token, None)


def parse_cookie(cookie_str):
    cookies = {}
    if not cookie_str:
        return cookies
    for part in cookie_str.split(';'):
        if '=' in part:
            k, v = part.strip().split('=', 1)
            cookies[k.strip()] = v.strip()
    return cookies


# --- MULTIPART PARSER ---

def parse_multipart(content_type, body):
    boundary = None
    for part in content_type.split(';'):
        part = part.strip()
        if part.startswith('boundary='):
            boundary = part[9:].strip().encode('latin-1')
            break
    if not boundary:
        return {}, {}

    fields = {}
    files  = {}
    delimiter = b'--' + boundary
    raw_parts = body.split(delimiter)

    for raw in raw_parts:
        if raw in (b'', b'--\r\n', b'--', b'\r\n'):
            continue
        if raw.startswith(b'--'):
            continue
        if raw.startswith(b'\r\n'):
            raw = raw[2:]
        if b'\r\n\r\n' not in raw:
            continue

        header_block, _, body_part = raw.partition(b'\r\n\r\n')
        if body_part.endswith(b'\r\n'):
            body_part = body_part[:-2]

        headers = {}
        for line in header_block.decode('utf-8', errors='replace').split('\r\n'):
            if ':' in line:
                k, _, v = line.partition(':')
                headers[k.strip().lower()] = v.strip()

        disposition = headers.get('content-disposition', '')
        name     = None
        filename = None
        for seg in disposition.split(';'):
            seg = seg.strip()
            if seg.startswith('name='):
                name = seg[5:].strip('"')
            elif seg.startswith('filename='):
                filename = seg[9:].strip('"')

        if name is None:
            continue

        if filename:
            class FileItem:
                pass
            fi = FileItem()
            fi.filename = filename
            fi.file = io.BytesIO(body_part)
            fi.content_type = headers.get('content-type', 'application/octet-stream')
            files[name] = fi
        else:
            fields[name] = body_part.decode('utf-8', errors='replace')

    return fields, files


# --- DOCUMENT META EXTRACTION (Claude AI) ---

# Prompt chuẩn cho văn bản hành chính Việt Nam
_PROMPT_VN_DOC = (
    'Phân tích văn bản hành chính Việt Nam và trích xuất CHÍNH XÁC 3 trường:\n\n'
    '1. doc_number — Số hiệu văn bản:\n'
    '   Tìm dòng bắt đầu bằng "Số:" ở góc trên bên trái.\n'
    '   Lấy TOÀN BỘ ký hiệu sau "Số:" (ví dụ: "11369/VP-ĐT", "123/2024/QĐ-UBND").\n'
    '   Nếu số và ký hiệu nằm tách nhau (ví dụ "11369" và "/VP-ĐT"), ghép lại thành "11369/VP-ĐT".\n\n'
    '2. issued_date — Ngày ban hành (định dạng YYYY-MM-DD):\n'
    '   Tìm cụm "ngày DD tháng MM năm YYYY" thường ở góc phải hoặc dòng địa danh.\n'
    '   Ví dụ: "Hà Nội, ngày 12 tháng 6 năm 2026" → "2026-06-12".\n\n'
    '3. issuer — Cơ quan ban hành:\n'
    '   Lấy từ khối tiêu đề góc TRÊN BÊN TRÁI (thường 2-3 dòng).\n'
    '   Ghép tên cơ quan cấp trên và tên đơn vị thành tên đầy đủ.\n'
    '   Ví dụ: "ỦY BAN NHÂN DÂN / THÀNH PHỐ HÀ NỘI / VĂN PHÒNG"\n'
    '   → "Văn phòng UBND Thành phố Hà Nội".\n\n'
    'Chỉ trả về JSON thuần, KHÔNG markdown, KHÔNG giải thích:\n'
    '{"doc_number":"...","issued_date":"YYYY-MM-DD","issuer":"..."}\n'
    'Dùng chuỗi rỗng "" nếu không tìm thấy.'
)

def extract_docx_text(file_bytes):
    """Extract plain text from a DOCX file using stdlib only."""
    from xml.etree import ElementTree as ET
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            with z.open('word/document.xml') as f:
                tree = ET.parse(f)
        ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
        parts = [el.text for el in tree.findall('.//w:t', ns) if el.text]
        return ' '.join(parts)[:8000]
    except Exception:
        return ''

def extract_pdf_text(file_bytes):
    """Extract text from PDF using pypdf (fallback when vision fails)."""
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        texts = []
        for i, page in enumerate(reader.pages):
            if i >= 3: break
            t = page.extract_text()
            if t:
                texts.append(t)
        return '\n'.join(texts)[:8000]
    except Exception as e:
        print(f'[pypdf] {e}')
        return ''

def _call_claude_api(content, extra_headers=None):
    """Call Claude Haiku API. Returns parsed dict or None."""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return None
    payload = json.dumps({
        'model': 'claude-haiku-4-5-20251001',
        'max_tokens': 300,
        'messages': [{'role': 'user', 'content': content}]
    }).encode('utf-8')
    headers = {
        'x-api-key': api_key,
        'anthropic-version': '2023-06-01',
        'content-type': 'application/json',
        **(extra_headers or {})
    }
    try:
        req = _urlreq.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload, headers=headers, method='POST'
        )
        with _urlreq.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        resp_text = data.get('content', [{}])[0].get('text', '').strip()
        print(f'[Claude meta] OK: {resp_text[:200]}')
        # Extract first complete JSON object (handles markdown code blocks too)
        m = re.search(r'\{[^{}]*\}', resp_text)
        if m:
            parsed = json.loads(m.group())
            # Normalize: replace null/"null" with empty string
            return {k: ('' if v in (None, 'null') else str(v)) for k, v in parsed.items()}
    except Exception as e:
        err_body = ''
        if hasattr(e, 'read'):
            try:
                err_body = e.read().decode('utf-8', errors='ignore')[:200]
            except Exception:
                pass
        print(f'[Claude meta] error: {e} {err_body}')
    return None

def call_claude_for_metadata(file_bytes, file_ext):
    """Call Claude Haiku to extract doc_number, issued_date, issuer from a document."""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        print('[Claude meta] ANTHROPIC_API_KEY not set')
        return None

    if file_ext == '.pdf':
        pdf_content = [
            {'type': 'document', 'source': {
                'type': 'base64', 'media_type': 'application/pdf',
                'data': base64.b64encode(file_bytes).decode()
            }},
            {'type': 'text', 'text': _PROMPT_VN_DOC}
        ]
        # Attempt 1: without beta header (PDF GA for claude 4.x models)
        result = _call_claude_api(pdf_content)
        if result:
            return result
        # Attempt 2: with beta header (for claude 3.x compatibility)
        result = _call_claude_api(pdf_content, {'anthropic-beta': 'pdfs-2024-09-25'})
        if result:
            return result
        # Fallback: pypdf text extraction (may miss visual-only text)
        text = extract_pdf_text(file_bytes)
        if text.strip():
            print('[Claude meta] Falling back to PDF text extraction')
            text_content = [{'type': 'text', 'text': f'Nội dung văn bản:\n{text}\n\n{_PROMPT_VN_DOC}'}]
            return _call_claude_api(text_content)
        return None

    elif file_ext == '.docx':
        text = extract_docx_text(file_bytes)
        if not text.strip():
            return None
        text_content = [{'type': 'text', 'text': f'Nội dung văn bản:\n{text}\n\n{_PROMPT_VN_DOC}'}]
        return _call_claude_api(text_content)

    return None


# --- HTTP HANDLER ---

class Handler(http.server.BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        pass

    def get_token(self):
        cookies = parse_cookie(self.headers.get('Cookie', ''))
        return cookies.get('session_token')

    def get_session(self):
        return get_session(self.get_token())

    def require_auth(self):
        s = self.get_session()
        if not s:
            self.send_json({'error': 'Chua dang nhap'}, 401)
        return s

    def require_manager(self):
        s = self.get_session()
        if not s or s['role'] != 'manager':
            self.send_json({'error': 'Khong co quyen'}, 403)
            return None
        return s

    def send_json(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, path):
        try:
            with open(path, 'rb') as f:
                data = f.read()
            ct = mimetypes.guess_type(path)[0] or 'application/octet-stream'
            self.send_response(200)
            self.send_header('Content-Type', ct)
            self.send_header('Content-Length', len(data))
            self.end_headers()
            self.wfile.write(data)
        except FileNotFoundError:
            self.send_response(404)
            self.end_headers()

    def read_json(self):
        length = int(self.headers.get('Content-Length', 0))
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length))

    def read_form(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length).decode('utf-8', errors='replace')
        raw = parse_qs(body, keep_blank_values=True)
        return {k: v[0] for k, v in raw.items()}

    def read_multipart(self):
        ct     = self.headers.get('Content-Type', '')
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length)
        return parse_multipart(ct, body)

    def get_path_and_query(self):
        parsed = urlparse(self.path)
        return parsed.path, parse_qs(parsed.query)

    # Routing

    def do_GET(self):
        path, qs = self.get_path_and_query()

        if path.startswith('/uploads/'):
            fname = os.path.basename(path)
            fpath = os.path.join(UPLOADS_DIR, fname)
            self.send_file(fpath)
            return

        routes = {
            '/api/me':                   self.api_me,
            '/api/manager/users':        self.api_manager_users_list,
            '/api/notifications':        self.api_notifications_list,
            '/api/notifications/unread': self.api_notifications_unread_count,
        }
        if path in routes:
            routes[path]()
            return

        if path == '/api/overtime/my':
            self.api_overtime_my(qs); return
        if path == '/api/manager/overtime':
            self.api_manager_overtime_list(qs); return
        if path == '/api/manager/stats':
            self.api_manager_stats(qs); return
        if path == '/api/manager/export':
            self.api_manager_export(qs); return
        if path == '/api/manager/storage':
            self.api_manager_storage(); return
        if path == '/api/manager/backups':
            self.api_manager_backups_list(); return
        if path == '/api/manager/backups/download':
            self.api_manager_backup_download(qs); return
        if path == '/api/docs':
            self.api_docs_list(qs); return
        if path == '/api/docs/my-permissions':
            self.api_docs_my_permissions(); return
        if path == '/api/docs/categories':
            self.api_doc_categories_list(); return
        if path == '/api/docs/types':
            self.api_doc_types_list(); return
        if path == '/api/docs/storage':
            self.api_docs_storage(); return
        m = re.match(r'^/api/docs/url/(\d+)$', path)
        if m:
            self.api_docs_presigned_url(int(m.group(1))); return
        m = re.match(r'^/api/docs/download/(\d+)$', path)
        if m:
            self.api_docs_download(int(m.group(1))); return
        if path == '/api/manager/doc-permissions':
            self.api_manager_doc_permissions(); return

        if path == '/docs':
            self.send_file(os.path.join(PUBLIC_DIR, 'docs.html'))
            return

        if path == '/cay':
            self.send_file(os.path.join(PUBLIC_DIR, 'cay.html'))
            return

        if path == '/boi-thuong':
            self.send_file(os.path.join(PUBLIC_DIR, 'boi-thuong.html'))
            return

        # BT API — GET
        if path == '/api/bt/projects':
            self.api_bt_projects_list(); return
        if path == '/api/bt/status-config':
            self.api_bt_status_config_list(); return
        m = re.match(r'^/api/bt/projects/(\d+)/stats$', path)
        if m:
            self.api_bt_project_stats(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/gpmb-status$', path)
        if m:
            self.api_bt_project_gpmb_status(int(m.group(1))); return
        m = re.match(r'^/api/bt/maps/(\d+)/mocgioi/recompute-check$', path)
        if m:
            self.api_bt_mocgioi_recompute_check(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/records$', path)
        if m:
            self.api_bt_records_list(int(m.group(1)), qs); return
        m = re.match(r'^/api/bt/projects/(\d+)/export$', path)
        if m:
            self.api_bt_export(int(m.group(1))); return
        if path == '/api/bt/parties':
            self.api_bt_parties_list(qs); return
        m = re.match(r'^/api/bt/parties/(\d+)/members$', path)
        if m:
            self.api_bt_members_list(int(m.group(1))); return
        if path == '/api/bt/parcel-master/search':
            self.api_bt_parcel_master_search(qs); return
        if path == '/api/bt/gpmb-template':
            self.api_bt_gpmb_template(); return
        m = re.match(r'^/api/bt/projects/(\d+)/parcels$', path)
        if m:
            self.api_bt_parcels_list(int(m.group(1)), qs); return
        m = re.match(r'^/api/bt/parcels/(\d+)$', path)
        if m:
            self.api_bt_parcel_detail(int(m.group(1))); return
        if path == '/api/bt/maps':
            self.api_bt_maps_list(qs); return
        m = re.match(r'^/api/bt/maps/(\d+)$', path)
        if m:
            self.api_bt_map_detail(int(m.group(1))); return
        m = re.match(r'^/api/bt/map-files/(\d+)/download$', path)
        if m:
            self.api_bt_map_files_download(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/assets$', path)
        if m:
            self.api_bt_assets_list(int(m.group(1)), qs); return
        m = re.match(r'^/api/bt/assets/(\d+)$', path)
        if m:
            self.api_bt_asset_detail(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/dossiers$', path)
        if m:
            self.api_bt_dossiers_list(int(m.group(1)), qs); return
        m = re.match(r'^/api/bt/dossiers/(\d+)$', path)
        if m:
            self.api_bt_dossier_detail(int(m.group(1))); return

        self.send_file(os.path.join(PUBLIC_DIR, 'index.html'))

    def do_POST(self):
        path, _ = self.get_path_and_query()
        routes = {
            '/api/login':                    self.api_login,
            '/api/logout':                   self.api_logout,
            '/api/change-password':          self.api_change_password,
            '/api/verify-password':          self.api_verify_password,
            '/api/overtime':                 self.api_submit_overtime,
            '/api/manager/users':            self.api_manager_create_user,
            '/api/notifications/read-all':   self.api_notifications_read_all,
            '/api/docs/upload':              self.api_docs_upload,
            '/api/docs/extract-meta':        self.api_docs_extract_meta,
            '/api/docs/categories':          self.api_doc_categories_create,
            '/api/docs/types':               self.api_doc_types_create,
            '/api/bt/projects':              self.api_bt_projects_create,
            '/api/bt/status-config':         self.api_bt_status_config_create,
            '/api/bt/parties':               self.api_bt_parties_create,
            '/api/bt/maps':                   self.api_bt_maps_create,
            '/api/manager/backups/run':      self.api_manager_backup_run,
        }
        if path in routes:
            routes[path]()
            return
        m = re.match(r'^/api/bt/projects/(\d+)/records$', path)
        if m:
            self.api_bt_records_create(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/import$', path)
        if m:
            self.api_bt_import_excel(int(m.group(1))); return
        m = re.match(r'^/api/bt/parties/(\d+)/members$', path)
        if m:
            self.api_bt_members_create(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/parcels$', path)
        if m:
            self.api_bt_parcels_create(int(m.group(1))); return
        m = re.match(r'^/api/bt/maps/(\d+)/parcels$', path)
        if m:
            self.api_bt_map_parcels_add(int(m.group(1))); return
        m = re.match(r'^/api/bt/maps/(\d+)/parcels/import$', path)
        if m:
            self.api_bt_map_parcels_import(int(m.group(1))); return
        m = re.match(r'^/api/bt/maps/(\d+)/mocgioi/apply-recompute$', path)
        if m:
            self.api_bt_mocgioi_apply_recompute(int(m.group(1))); return
        if path == '/api/bt/mocgioi/apply-recompute-for-manh':
            self.api_bt_mocgioi_apply_recompute_for_manh(); return
        m = re.match(r'^/api/bt/maps/(\d+)/files$', path)
        if m:
            self.api_bt_map_files_upload(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/assets$', path)
        if m:
            self.api_bt_assets_create(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)/dossiers$', path)
        if m:
            self.api_bt_dossiers_create(int(m.group(1))); return
        self.send_json({'error': 'Not found'}, 404)

    def do_PUT(self):
        path, _ = self.get_path_and_query()
        m = re.match(r'^/api/overtime/(\d+)$', path)
        if m:
            self.api_overtime_edit(int(m.group(1))); return
        m = re.match(r'^/api/manager/overtime/(\d+)$', path)
        if m:
            self.api_manager_review_overtime(int(m.group(1))); return
        m = re.match(r'^/api/manager/overtime/(\d+)/cancel$', path)
        if m:
            self.api_manager_cancel_overtime(int(m.group(1))); return
        m = re.match(r'^/api/manager/users/(\d+)$', path)
        if m:
            self.api_manager_update_user(int(m.group(1))); return
        m = re.match(r'^/api/manager/users/(\d+)/reset-password$', path)
        if m:
            self.api_manager_reset_password(int(m.group(1))); return
        m = re.match(r'^/api/notifications/(\d+)/read$', path)
        if m:
            self.api_notification_mark_read(int(m.group(1))); return
        m = re.match(r'^/api/docs/(\d+)$', path)
        if m:
            self.api_docs_update(int(m.group(1))); return
        m = re.match(r'^/api/docs/categories/(\d+)$', path)
        if m:
            self.api_doc_categories_update(int(m.group(1))); return
        m = re.match(r'^/api/docs/types/(\d+)$', path)
        if m:
            self.api_doc_types_update(int(m.group(1))); return
        m = re.match(r'^/api/manager/doc-permissions/(\d+)$', path)
        if m:
            self.api_manager_doc_permission_set(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)$', path)
        if m:
            self.api_bt_projects_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/records/(\d+)$', path)
        if m:
            self.api_bt_records_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/status-config/(\d+)$', path)
        if m:
            self.api_bt_status_config_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/parties/(\d+)$', path)
        if m:
            self.api_bt_parties_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/members/(\d+)$', path)
        if m:
            self.api_bt_members_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/parcels/(\d+)$', path)
        if m:
            self.api_bt_parcels_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/maps/(\d+)$', path)
        if m:
            self.api_bt_maps_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/map-parcels/(\d+)$', path)
        if m:
            self.api_bt_map_parcels_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/assets/(\d+)$', path)
        if m:
            self.api_bt_assets_update(int(m.group(1))); return
        m = re.match(r'^/api/bt/dossiers/(\d+)$', path)
        if m:
            self.api_bt_dossiers_update(int(m.group(1))); return
        self.send_json({'error': 'Not found'}, 404)

    def do_DELETE(self):
        path, qs = self.get_path_and_query()
        m = re.match(r'^/api/overtime/(\d+)$', path)
        if m:
            self.api_overtime_delete(int(m.group(1))); return
        m = re.match(r'^/api/manager/users/(\d+)$', path)
        if m:
            self.api_manager_delete_user(int(m.group(1))); return
        if path == '/api/manager/data/cleanup':
            self.api_manager_cleanup_data(); return
        if path == '/api/manager/data/reset-all':
            self.api_manager_reset_all(); return
        m = re.match(r'^/api/docs/(\d+)$', path)
        if m:
            self.api_docs_delete(int(m.group(1))); return
        m = re.match(r'^/api/docs/categories/(\d+)$', path)
        if m:
            self.api_doc_categories_delete(int(m.group(1)), qs); return
        m = re.match(r'^/api/docs/types/(\d+)$', path)
        if m:
            self.api_doc_types_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/projects/(\d+)$', path)
        if m:
            self.api_bt_projects_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/records/(\d+)$', path)
        if m:
            self.api_bt_records_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/status-config/(\d+)$', path)
        if m:
            self.api_bt_status_config_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/parties/(\d+)$', path)
        if m:
            self.api_bt_parties_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/members/(\d+)$', path)
        if m:
            self.api_bt_members_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/parcels/(\d+)$', path)
        if m:
            self.api_bt_parcels_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/maps/(\d+)$', path)
        if m:
            self.api_bt_maps_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/map-parcels/(\d+)$', path)
        if m:
            self.api_bt_map_parcels_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/map-files/(\d+)$', path)
        if m:
            self.api_bt_map_files_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/assets/(\d+)$', path)
        if m:
            self.api_bt_assets_delete(int(m.group(1))); return
        m = re.match(r'^/api/bt/dossiers/(\d+)$', path)
        if m:
            self.api_bt_dossiers_delete(int(m.group(1))); return
        self.send_json({'error': 'Not found'}, 404)

    def do_PATCH(self):
        path, _ = self.get_path_and_query()
        m = re.match(r'^/api/docs/categories/(\d+)/parent$', path)
        if m:
            self.api_doc_categories_reparent(int(m.group(1))); return
        self.send_json({'error': 'Not found'}, 404)

    def api_doc_categories_reparent(self, cat_id):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        raw = body.get('parent_id')
        parent_id = int(raw) if raw else None
        if parent_id == cat_id:
            self.send_json({'error': 'Khong the dat chinh no lam cha'}, 400); return
        with get_db() as db:
            if parent_id:
                # Prevent circular: recursively collect all descendants of cat_id
                def get_descendants(anc_id):
                    kids = [r[0] for r in db.execute(
                        'SELECT id FROM doc_categories WHERE parent_id=?', (anc_id,)).fetchall()]
                    result = list(kids)
                    for k in kids:
                        result.extend(get_descendants(k))
                    return result
                if parent_id in get_descendants(cat_id):
                    self.send_json({'error': 'Khong the tao vong lap danh muc'}, 400); return
            db.execute('UPDATE doc_categories SET parent_id=? WHERE id=?', (parent_id, cat_id))
            db.commit()
        self.send_json({'ok': True})

    # --- AUTH ---

    def api_login(self):
        ct = self.headers.get('Content-Type', '')
        body = self.read_json() if 'application/json' in ct else self.read_form()
        username = body.get('username', '').strip()
        password = body.get('password', '')
        with get_db() as db:
            row = db.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        if not row or not check_password(password, row['password']):
            self.send_json({'error': 'Sai tai khoan hoac mat khau'}, 401)
            return
        token = create_session(row['id'], row['role'], row['full_name'])
        data  = {
            'id': row['id'], 'username': row['username'],
            'full_name': row['full_name'], 'role': row['role']
        }
        body_bytes = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', len(body_bytes))
        self.send_header('Set-Cookie',
            f'session_token={token}; Path=/; HttpOnly; Max-Age=28800')
        self.end_headers()
        self.wfile.write(body_bytes)

    def api_logout(self):
        token = self.get_token()
        if token:
            delete_session(token)
        body_bytes = b'{"ok":true}'
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', len(body_bytes))
        self.send_header('Set-Cookie', 'session_token=; Path=/; Max-Age=0')
        self.end_headers()
        self.wfile.write(body_bytes)

    def api_me(self):
        sess = self.require_auth()
        if not sess:
            return
        with get_db() as db:
            row = db.execute(
                'SELECT id,username,full_name,role,department FROM users WHERE id=?',
                (sess['userId'],)
            ).fetchone()
        if not row:
            self.send_json({'error': 'Not found'}, 404)
            return
        self.send_json(dict(row))

    def api_verify_password(self):
        """Xác thực lại mật khẩu tài khoản hiện tại — dùng để yêu cầu xác nhận trước các hành động xóa dữ liệu."""
        sess = self.require_auth()
        if not sess:
            return
        body = self.read_json()
        pw = body.get('password', '')
        with get_db() as db:
            row = db.execute('SELECT password FROM users WHERE id=?', (sess['userId'],)).fetchone()
        if not row or not check_password(pw, row['password']):
            self.send_json({'ok': False, 'error': 'Mật khẩu không đúng'}, 401)
            return
        self.send_json({'ok': True})

    def api_change_password(self):
        sess = self.require_auth()
        if not sess:
            return
        body   = self.read_json()
        old_pw = body.get('old_password', '')
        new_pw = body.get('new_password', '')
        with get_db() as db:
            row = db.execute('SELECT password FROM users WHERE id=?',
                             (sess['userId'],)).fetchone()
            if not check_password(old_pw, row['password']):
                self.send_json({'error': 'Mat khau cu khong dung'}, 400)
                return
            if len(new_pw) < 6:
                self.send_json({'error': 'Mat khau moi it nhat 6 ky tu'}, 400)
                return
            db.execute('UPDATE users SET password=? WHERE id=?',
                       (hash_password(new_pw), sess['userId']))
            db.commit()
        self.send_json({'ok': True})

    # --- EMPLOYEE OVERTIME ---

    def api_submit_overtime(self):
        sess = self.require_auth()
        if not sess:
            return
        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' in ct:
            fields, files = self.read_multipart()
        else:
            fields = self.read_form()
            files  = {}

        req_date   = fields.get('request_date', '').strip()
        start_time = fields.get('start_time', '').strip()
        end_time   = fields.get('end_time', '').strip()
        reason     = fields.get('reason', '').strip()

        if not all([req_date, start_time, end_time, reason]):
            self.send_json({'error': 'Vui long dien day du thong tin'}, 400)
            return

        try:
            from datetime import date as dclass
            y, mo, d = map(int, req_date.split('-'))
            dow = dclass(y, mo, d).weekday()
        except Exception:
            self.send_json({'error': 'Ngay khong hop le'}, 400)
            return

        if dow <= 4:
            ot_type = 'weekday'; min_start = (17,30); max_end = (23,0)
            rule_msg = 'OT ngay thuong chi duoc tu 17:30 den 23:00'
        elif dow == 5:
            ot_type = 'weekend'; min_start = (13,30); max_end = (23,0)
            rule_msg = 'OT thu 7 chi duoc tu 13:30 den 23:00'
        else:
            ot_type = 'weekend'; min_start = (8,0); max_end = (23,0)
            rule_msg = 'OT chu nhat chi duoc tu 08:00 den 23:00'

        try:
            sh, sm = map(int, start_time.split(':'))
            eh, em = map(int, end_time.split(':'))
        except Exception:
            self.send_json({'error': 'Gio khong hop le'}, 400)
            return

        start_min = sh*60+sm
        end_min   = eh*60+em
        min_start_min = min_start[0]*60+min_start[1]
        max_end_min   = max_end[0]*60+max_end[1]

        if start_min < min_start_min or end_min > max_end_min:
            self.send_json({'error': rule_msg}, 400)
            return
        if end_min <= start_min:
            self.send_json({'error': 'Gio ket thuc phai sau gio bat dau'}, 400)
            return

        now = datetime.now(VN_TZ).replace(tzinfo=None)
        today_str = now.strftime('%Y-%m-%d')
        if req_date == today_str:
            now_min = now.hour*60+now.minute
            now_hm  = now.strftime('%H:%M')
            if end_min > now_min:
                self.send_json({'error': 'Gio ket thuc (' + end_time + ') vuot qua thoi diem hien tai (' + now_hm + ')'}, 400)
                return
            if start_min >= now_min:
                self.send_json({'error': 'Gio bat dau (' + start_time + ') chua den gio hien tai (' + now_hm + ')'}, 400)
                return

        raw_minutes = end_min - start_min
        deduct_min  = 0
        if dow == 6:
            lunch_s = 720; lunch_e = 810
            overlap = max(0, min(end_min, lunch_e) - max(start_min, lunch_s))
            deduct_min = overlap

        hours = (raw_minutes - deduct_min) / 60.0
        if hours <= 0:
            self.send_json({'error': 'Tong gio OT sau khi tru nghi trua phai lon hon 0'}, 400)
            return

        image_path = None
        if 'image' in files:
            fitem = files['image']
            ext   = os.path.splitext(fitem.filename)[1].lower()
            if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'):
                self.send_json({'error': 'Dinh dang file khong ho tro'}, 400)
                return
            file_bytes = fitem.file.read()
            if ext != '.pdf':
                file_bytes, ext = compress_image(file_bytes, ext)
            fname = f"{int(datetime.now().timestamp()*1000)}_{sess['userId']}{ext}"
            fpath = os.path.join(UPLOADS_DIR, fname)
            with open(fpath, 'wb') as f:
                f.write(file_bytes)
            image_path = f'/uploads/{fname}'

        with get_db() as db:
            db.execute(
                """INSERT INTO overtime_requests
                   (user_id, request_date, ot_type, start_time, end_time,
                    hours, reason, image_path)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (sess['userId'], req_date, ot_type, start_time,
                 end_time, hours, reason, image_path)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_overtime_delete(self, ot_id):
        sess = self.require_auth()
        if not sess:
            return
        with get_db() as db:
            row = db.execute(
                'SELECT * FROM overtime_requests WHERE id=? AND user_id=?',
                (ot_id, sess['userId'])
            ).fetchone()
            if not row:
                self.send_json({'error': 'Khong tim thay khai bao'}, 404)
                return
            if row['status'] != 'pending':
                self.send_json({'error': 'Yeu cau huy Khai bao cua ban khong the thuc hien do Admin da phe duyet Khai bao nay'}, 400)
                return
            if row['image_path']:
                try:
                    fpath = os.path.join(DATA_DIR, row['image_path'].lstrip('/'))
                    if os.path.exists(fpath):
                        os.remove(fpath)
                except Exception:
                    pass
            db.execute('DELETE FROM overtime_requests WHERE id=?', (ot_id,))
            db.execute('DELETE FROM notifications WHERE ot_id=?', (ot_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_overtime_edit(self, ot_id):
        sess = self.require_auth()
        if not sess:
            return
        with get_db() as db:
            row = db.execute(
                'SELECT * FROM overtime_requests WHERE id=? AND user_id=?',
                (ot_id, sess['userId'])
            ).fetchone()
            if not row:
                self.send_json({'error': 'Khong tim thay khai bao'}, 404)
                return
            if row['status'] != 'pending':
                self.send_json({'error': 'Yeu cau chinh sua Khai bao cua ban khong the thuc hien do Admin da phe duyet Khai bao nay'}, 400)
                return

        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' in ct:
            fields, files = self.read_multipart()
        else:
            fields = self.read_form()
            files  = {}

        req_date   = fields.get('request_date', '').strip()
        start_time = fields.get('start_time', '').strip()
        end_time   = fields.get('end_time', '').strip()
        reason     = fields.get('reason', '').strip()

        if not all([req_date, start_time, end_time, reason]):
            self.send_json({'error': 'Vui long dien day du thong tin'}, 400)
            return

        try:
            from datetime import date as dclass
            y, mo, d = map(int, req_date.split('-'))
            dow = dclass(y, mo, d).weekday()
        except Exception:
            self.send_json({'error': 'Ngay khong hop le'}, 400)
            return

        if dow <= 4:
            ot_type = 'weekday'; min_start = (17,30); max_end = (23,0)
            rule_msg = 'OT ngay thuong chi duoc tu 17:30 den 23:00'
        elif dow == 5:
            ot_type = 'weekend'; min_start = (13,30); max_end = (23,0)
            rule_msg = 'OT thu 7 chi duoc tu 13:30 den 23:00'
        else:
            ot_type = 'weekend'; min_start = (8,0); max_end = (23,0)
            rule_msg = 'OT chu nhat chi duoc tu 08:00 den 23:00'

        try:
            sh, sm = map(int, start_time.split(':'))
            eh, em = map(int, end_time.split(':'))
        except Exception:
            self.send_json({'error': 'Gio khong hop le'}, 400)
            return

        start_min = sh*60+sm; end_min = eh*60+em
        min_start_min = min_start[0]*60+min_start[1]
        max_end_min   = max_end[0]*60+max_end[1]

        if start_min < min_start_min or end_min > max_end_min:
            self.send_json({'error': rule_msg}, 400); return
        if end_min <= start_min:
            self.send_json({'error': 'Gio ket thuc phai sau gio bat dau'}, 400); return

        now = datetime.now(VN_TZ).replace(tzinfo=None)
        today_str = now.strftime('%Y-%m-%d')
        if req_date == today_str:
            now_min = now.hour*60+now.minute
            now_hm  = now.strftime('%H:%M')
            if end_min > now_min:
                self.send_json({'error': 'Gio ket thuc (' + end_time + ') vuot qua thoi diem hien tai (' + now_hm + ')'}, 400); return
            if start_min >= now_min:
                self.send_json({'error': 'Gio bat dau (' + start_time + ') chua den gio hien tai (' + now_hm + ')'}, 400); return

        raw_minutes = end_min - start_min
        deduct_min = 0
        if dow == 6:
            lunch_s = 720; lunch_e = 810
            deduct_min = max(0, min(end_min, lunch_e) - max(start_min, lunch_s))
        hours = (raw_minutes - deduct_min) / 60.0
        if hours <= 0:
            self.send_json({'error': 'Tong gio OT phai lon hon 0'}, 400); return

        image_path = row['image_path']
        if 'image' in files:
            fitem = files['image']
            ext = os.path.splitext(fitem.filename)[1].lower()
            if ext not in ('.jpg','.jpeg','.png','.gif','.webp','.pdf'):
                self.send_json({'error': 'Dinh dang file khong ho tro'}, 400); return
            file_bytes = fitem.file.read()
            if ext != '.pdf':
                file_bytes, ext = compress_image(file_bytes, ext)
            fname = str(int(datetime.now().timestamp()*1000)) + '_' + str(sess['userId']) + ext
            fpath = os.path.join(UPLOADS_DIR, fname)
            with open(fpath, 'wb') as f:
                f.write(file_bytes)
            if row['image_path']:
                try:
                    old = os.path.join(DATA_DIR, row['image_path'].lstrip('/'))
                    if os.path.exists(old):
                        os.remove(old)
                except Exception:
                    pass
            image_path = '/uploads/' + fname

        with get_db() as db:
            db.execute(
                """UPDATE overtime_requests
                   SET request_date=?, ot_type=?, start_time=?, end_time=?,
                       hours=?, reason=?, image_path=?
                   WHERE id=? AND user_id=? AND status='pending'""",
                (req_date, ot_type, start_time, end_time,
                 hours, reason, image_path, ot_id, sess['userId'])
            )
            db.commit()
        self.send_json({'ok': True})

    def api_overtime_my(self, qs):
        sess = self.require_auth()
        if not sess:
            return
        month  = qs.get('month', [None])[0]
        year   = qs.get('year',  [None])[0]
        sql    = """SELECT r.*, u.full_name, u.department
                    FROM overtime_requests r
                    JOIN users u ON r.user_id = u.id
                    WHERE r.user_id = ?"""
        params = [sess['userId']]
        if month and year:
            sql += " AND strftime('%Y-%m', r.request_date) = ?"
            params.append(f"{year}-{month.zfill(2)}")
        sql += ' ORDER BY r.request_date DESC, r.created_at DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json(rows_to_list(rows))

    # --- NOTIFICATIONS ---

    def api_notifications_list(self):
        sess = self.require_auth()
        if not sess:
            return
        with get_db() as db:
            rows = db.execute(
                """SELECT n.*, r.request_date, r.start_time, r.end_time, r.ot_type
                   FROM notifications n
                   JOIN overtime_requests r ON n.ot_id = r.id
                   WHERE n.user_id = ?
                   ORDER BY n.created_at DESC
                   LIMIT 50""",
                (sess['userId'],)
            ).fetchall()
        self.send_json(rows_to_list(rows))

    def api_notifications_unread_count(self):
        sess = self.require_auth()
        if not sess:
            return
        with get_db() as db:
            row = db.execute(
                'SELECT COUNT(*) as cnt FROM notifications WHERE user_id=? AND is_read=0',
                (sess['userId'],)
            ).fetchone()
        self.send_json({'count': row['cnt']})

    def api_notification_mark_read(self, notif_id):
        sess = self.require_auth()
        if not sess:
            return
        with get_db() as db:
            db.execute(
                'UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?',
                (notif_id, sess['userId'])
            )
            db.commit()
        self.send_json({'ok': True})

    def api_notifications_read_all(self):
        sess = self.require_auth()
        if not sess:
            return
        with get_db() as db:
            db.execute(
                'UPDATE notifications SET is_read=1 WHERE user_id=?',
                (sess['userId'],)
            )
            db.commit()
        self.send_json({'ok': True})

    # --- MANAGER ---

    def api_manager_overtime_list(self, qs):
        sess = self.require_manager()
        if not sess:
            return
        status  = qs.get('status',  [None])[0]
        month   = qs.get('month',   [None])[0]
        year    = qs.get('year',    [None])[0]
        user_id = qs.get('user_id', [None])[0]
        sql     = """SELECT r.*, u.full_name, u.department
                     FROM overtime_requests r
                     JOIN users u ON r.user_id = u.id
                     WHERE 1=1"""
        params  = []
        if status:
            sql += ' AND r.status = ?'
            params.append(status)
        if month and year:
            sql += " AND strftime('%Y-%m', r.request_date) = ?"
            params.append(f"{year}-{month.zfill(2)}")
        if user_id:
            sql += ' AND r.user_id = ?'
            params.append(int(user_id))
        sql += ' ORDER BY r.created_at DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json(rows_to_list(rows))

    def api_manager_review_overtime(self, ot_id):
        sess = self.require_manager()
        if not sess:
            return
        body   = self.read_json()
        status = body.get('status', '')
        note   = body.get('manager_note', '')
        if status not in ('approved', 'rejected'):
            self.send_json({'error': 'Trang thai khong hop le'}, 400)
            return
        now = datetime.now(VN_TZ).strftime('%Y-%m-%d %H:%M:%S')
        with get_db() as db:
            db.execute(
                """UPDATE overtime_requests
                   SET status=?, manager_note=?, reviewed_by=?, reviewed_at=?
                   WHERE id=?""",
                (status, note, sess['userId'], now, ot_id)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_manager_cancel_overtime(self, ot_id):
        sess = self.require_manager()
        if not sess:
            return
        body   = self.read_json()
        reason = body.get('cancel_reason', '').strip()
        now    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with get_db() as db:
            row = db.execute(
                'SELECT * FROM overtime_requests WHERE id=?', (ot_id,)
            ).fetchone()
            if not row:
                self.send_json({'error': 'Khong tim thay yeu cau'}, 404)
                return
            if row['status'] != 'approved':
                self.send_json({'error': 'Chi huy duoc yeu cau da duyet'}, 400)
                return
            db.execute(
                """UPDATE overtime_requests
                   SET status='cancelled', manager_note=?, reviewed_by=?, reviewed_at=?
                   WHERE id=?""",
                (reason, sess['userId'], now, ot_id)
            )
            ot_date  = row['request_date']
            st_time  = row['start_time']
            en_time  = row['end_time']
            title    = 'Yeu cau OT bi huy'
            msg_body = (
                f"Yeu cau OT ngay {ot_date} ({st_time}-{en_time}) "
                f"da duoc duyet truoc do vua bi HUY boi quan ly."
            )
            if reason:
                msg_body += f" Ly do: {reason}"
            db.execute(
                'INSERT INTO notifications (user_id, ot_id, title, message) VALUES (?,?,?,?)',
                (row['user_id'], ot_id, title, msg_body)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_manager_users_list(self):
        sess = self.require_manager()
        if not sess:
            return
        with get_db() as db:
            rows = db.execute(
                'SELECT id,username,full_name,role,department,created_at FROM users ORDER BY full_name'
            ).fetchall()
        self.send_json(rows_to_list(rows))

    def api_manager_create_user(self):
        sess = self.require_manager()
        if not sess:
            return
        body      = self.read_json()
        username  = body.get('username', '').strip()
        password  = body.get('password', '')
        full_name = body.get('full_name', '').strip()
        role      = body.get('role', 'employee')
        dept      = body.get('department', '')
        if not username or not password or not full_name:
            self.send_json({'error': 'Thieu thong tin bat buoc'}, 400)
            return
        try:
            with get_db() as db:
                db.execute(
                    'INSERT INTO users (username,password,full_name,role,department) VALUES (?,?,?,?,?)',
                    (username, hash_password(password), full_name, role, dept)
                )
                db.commit()
            self.send_json({'ok': True})
        except sqlite3.IntegrityError:
            self.send_json({'error': 'Ten dang nhap da ton tai'}, 400)

    def api_manager_update_user(self, uid):
        sess = self.require_manager()
        if not sess:
            return
        body      = self.read_json()
        full_name = body.get('full_name', '').strip()
        role      = body.get('role', 'employee')
        dept      = body.get('department', '')
        password  = body.get('password')
        with get_db() as db:
            if password:
                db.execute(
                    'UPDATE users SET full_name=?,role=?,department=?,password=? WHERE id=?',
                    (full_name, role, dept, hash_password(password), uid)
                )
            else:
                db.execute(
                    'UPDATE users SET full_name=?,role=?,department=? WHERE id=?',
                    (full_name, role, dept, uid)
                )
            db.commit()
        self.send_json({'ok': True})

    def api_manager_delete_user(self, uid):
        sess = self.require_manager()
        if not sess:
            return
        if uid == sess['userId']:
            self.send_json({'error': 'Khong the xoa chinh minh'}, 400)
            return
        with get_db() as db:
            db.execute('DELETE FROM users WHERE id=?', (uid,))
            db.commit()
        self.send_json({'ok': True})

    def api_manager_stats(self, qs):
        sess = self.require_manager()
        if not sess:
            return
        month  = qs.get('month', [None])[0]
        year   = qs.get('year',  [None])[0]
        date_filter = ''
        params = []
        if month and year:
            date_filter = "AND strftime('%Y-%m', r.request_date) = ?"
            params.append(f"{year}-{month.zfill(2)}")
        sql = f"""
            SELECT
                u.id, u.full_name, u.department,
                COUNT(CASE WHEN r.status='approved' THEN 1 END) as approved_count,
                COALESCE(SUM(CASE WHEN r.status='approved' THEN r.hours ELSE 0 END), 0) as total_hours,
                COALESCE(SUM(CASE WHEN r.status='approved' AND r.ot_type='weekday' THEN r.hours ELSE 0 END), 0) as weekday_hours,
                COALESCE(SUM(CASE WHEN r.status='approved' AND r.ot_type='weekend' THEN r.hours ELSE 0 END), 0) as weekend_hours,
                COUNT(CASE WHEN r.status='pending' THEN 1 END) as pending_count
            FROM users u
            LEFT JOIN overtime_requests r ON u.id = r.user_id {date_filter}
            WHERE u.role = 'employee'
            GROUP BY u.id
            ORDER BY u.full_name
        """
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json(rows_to_list(rows))

    def api_manager_export(self, qs):
        sess = self.require_manager()
        if not sess:
            return
        month = qs.get('month', [None])[0]
        year  = qs.get('year',  [None])[0]
        sql = """SELECT r.id, u.full_name, u.department,
                        r.request_date, r.ot_type,
                        r.start_time, r.end_time, r.hours,
                        r.reason, r.status, r.manager_note, r.created_at
                 FROM overtime_requests r
                 JOIN users u ON r.user_id = u.id
                 WHERE 1=1"""
        params = []
        if month and year:
            sql += " AND strftime('%Y-%m', r.request_date) = ?"
            params.append(f"{year}-{month.zfill(2)}")
        sql += ' ORDER BY r.request_date DESC, u.full_name'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()

        label_month = f'T{month}_{year}' if month and year else 'tat_ca'

        if OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Cham Cong OT'

            hdr_fill  = PatternFill('solid', fgColor='2563EB')
            hdr_font  = Font(bold=True, color='FFFFFF')
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            headers = ['STT','Ho ten','Bo phan','Ngay OT','Loai OT',
                       'Gio bat dau','Gio ket thuc','So gio',
                       'Ly do','Trang thai','Ghi chu QT','Ngay gui']
            col_widths = [5, 22, 16, 13, 14, 13, 13, 9, 40, 14, 30, 18]

            for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'A2'

            status_map = {
                'pending':   'Cho duyet',
                'approved':  'Da duyet',
                'rejected':  'Tu choi',
                'cancelled': 'Da huy',
            }
            status_colors = {
                'pending':   'FEF3C7',
                'approved':  'DCFCE7',
                'rejected':  'FEE2E2',
                'cancelled': 'F1F5F9',
            }

            for i, r in enumerate(rows, 1):
                ot_label = 'Ngay thuong' if r['ot_type'] == 'weekday' else 'Cuoi tuan/Le'
                status   = r['status']
                values   = [
                    i, r['full_name'], r['department'] or '',
                    r['request_date'], ot_label,
                    r['start_time'], r['end_time'],
                    round(r['hours'] or 0, 2),
                    r['reason'], status_map.get(status, status),
                    r['manager_note'] or '', r['created_at']
                ]
                fill = PatternFill('solid', fgColor=status_colors.get(status, 'FFFFFF')) if status in status_colors else None
                for ci, v in enumerate(values, 1):
                    cell = ws.cell(row=i+1, column=ci, value=v)
                    cell.alignment = Alignment(vertical='top', wrap_text=(ci == 9))
                    if fill and ci == 10:
                        cell.fill = fill

            total_row = len(rows) + 3
            ws.cell(row=total_row, column=1, value='Tong cong').font = Font(bold=True)
            approved_hours = sum(r['hours'] or 0 for r in rows if r['status'] == 'approved')
            ws.cell(row=total_row, column=8, value=round(approved_hours, 2)).font = Font(bold=True)
            ws.cell(row=total_row, column=9, value=f"Tong gio OT da duyet: {approved_hours:.1f}h").font = Font(bold=True)

            buf = io.BytesIO()
            wb.save(buf)
            data = buf.getvalue()
            fname = f'cham_cong_OT_{label_month}.xlsx'
            ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            import csv
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(['STT','Ho ten','Bo phan','Ngay OT','Loai OT',
                        'Gio bat dau','Gio ket thuc','So gio',
                        'Ly do','Trang thai','Ghi chu QT','Ngay gui'])
            for i, r in enumerate(rows, 1):
                w.writerow([i, r['full_name'], r['department'] or '',
                             r['request_date'],
                             'Ngay thuong' if r['ot_type']=='weekday' else 'Cuoi tuan',
                             r['start_time'], r['end_time'], r['hours'] or 0,
                             r['reason'], r['status'], r['manager_note'] or '', r['created_at']])
            data = buf.getvalue().encode('utf-8-sig')
            fname = f'cham_cong_OT_{label_month}.csv'
            ct = 'text/csv; charset=utf-8'

        self.send_response(200)
        self.send_header('Content-Type', ct)
        self.send_header('Content-Length', len(data))
        self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
        self.end_headers()
        self.wfile.write(data)

    def api_manager_reset_password(self, uid):
        sess = self.require_manager()
        if not sess:
            return
        with get_db() as db:
            row = db.execute('SELECT id, full_name FROM users WHERE id=?', (uid,)).fetchone()
            if not row:
                self.send_json({'error': 'Khong tim thay tai khoan'}, 404)
                return
            new_pw = generate_temp_password()
            db.execute('UPDATE users SET password=? WHERE id=?',
                       (hash_password(new_pw), uid))
            db.commit()
        self.send_json({'ok': True, 'new_password': new_pw, 'full_name': row['full_name']})

    def api_manager_cleanup_data(self):
        sess = self.require_manager()
        if not sess:
            return
        body = self.read_json()
        before_month = body.get('before_month', '')
        if not before_month:
            self.send_json({'error': 'Thieu truong before_month'}, 400)
            return
        with get_db() as db:
            rows = db.execute(
                "SELECT image_path FROM overtime_requests WHERE strftime('%Y-%m', request_date) < ?",
                (before_month,)
            ).fetchall()
            for r in rows:
                if r['image_path']:
                    try:
                        fpath = os.path.join(DATA_DIR, r['image_path'].lstrip('/'))
                        if os.path.exists(fpath):
                            os.remove(fpath)
                    except Exception:
                        pass
            db.execute(
                "DELETE FROM overtime_requests WHERE strftime('%Y-%m', request_date) < ?",
                (before_month,)
            )
            db.execute(
                "DELETE FROM notifications WHERE ot_id NOT IN (SELECT id FROM overtime_requests)"
            )
            db.execute('VACUUM')
            db.commit()
        self.send_json({'ok': True})

    def api_manager_reset_all(self):
        sess = self.require_manager()
        if not sess:
            return
        body = self.read_json()
        if body.get('confirm') != 'XAC_NHAN_XOA_HET':
            self.send_json({'error': 'Xac nhan khong dung'}, 400)
            return
        for fname in os.listdir(UPLOADS_DIR):
            try:
                os.remove(os.path.join(UPLOADS_DIR, fname))
            except Exception:
                pass
        with get_db() as db:
            db.execute('DELETE FROM overtime_requests')
            db.execute('DELETE FROM notifications')
            db.execute('DELETE FROM users WHERE role != ?', ('manager',))
            db.execute('VACUUM')
            db.commit()
        self.send_json({'ok': True})

    def api_manager_storage(self):
        sess = self.require_manager()
        if not sess:
            return
        try:
            db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
            uploads_size = sum(
                os.path.getsize(os.path.join(UPLOADS_DIR, f))
                for f in os.listdir(UPLOADS_DIR)
                if os.path.isfile(os.path.join(UPLOADS_DIR, f))
            ) if os.path.exists(UPLOADS_DIR) else 0
            total_bytes = db_size + uploads_size
            used_mb  = round(total_bytes / (1024*1024), 1)
            limit_mb = 500
            percent  = round(used_mb / limit_mb * 100, 1)
            self.send_json({
                'used_mb': used_mb,
                'limit_mb': limit_mb,
                'percent': percent,
                'is_warning': used_mb >= 410,
                'is_danger':  used_mb >= 480,
            })
        except Exception as e:
            self.send_json({'error': str(e)}, 500)

    # --- BACKUP (cham_cong.db -> R2) ---

    def api_manager_backups_list(self):
        sess = self.require_manager()
        if not sess:
            return
        self.send_json({'enabled': R2_ENABLED, 'backups': list_r2_backups()})

    def api_manager_backup_run(self):
        sess = self.require_manager()
        if not sess:
            return
        result = backup_db_to_r2()
        self.send_json(result, 200 if result.get('ok') else 500)

    def api_manager_backup_download(self, qs):
        sess = self.require_manager()
        if not sess:
            return
        key = (qs.get('key') or [''])[0]
        if not key or not key.startswith('backups/'):
            self.send_json({'error': 'Key khong hop le'}, 400); return
        try:
            client = get_r2_client()
            obj = client.get_object(Bucket=R2_BUCKET, Key=key)
            data = obj['Body'].read()
        except Exception as e:
            self.send_json({'error': f'Khong tai duoc backup: {e}'}, 500); return
        fname = key.split('/')[-1]
        self.send_response(200)
        self.send_header('Content-Type', 'application/octet-stream')
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Content-Disposition',
                          f"attachment; filename*=UTF-8''{urllib.parse.quote(fname)}")
        self.send_header('Cache-Control', 'no-cache')
        self.end_headers()
        self.wfile.write(data)

    # --- DOCUMENTS ---

    def _get_doc_perms(self, sess):
        """Lay quyen tai lieu cua user (manager luon co tat ca quyen)"""
        if sess['role'] == 'manager':
            return {'can_upload_docs':1,'can_view_docs':1,'can_download_docs':1,'can_edit_docs':1,'can_delete_docs':1}
        with get_db() as db:
            row = db.execute(
                'SELECT can_upload_docs,can_view_docs,can_download_docs,can_edit_docs,can_delete_docs FROM users WHERE id=?',
                (sess['userId'],)
            ).fetchone()
        if not row:
            return {'can_upload_docs':0,'can_view_docs':0,'can_download_docs':0,'can_edit_docs':0,'can_delete_docs':0}
        return dict(row)

    def can_upload_docs(self, sess):
        return self._get_doc_perms(sess)['can_upload_docs'] == 1

    def can_view_docs(self, sess):
        return self._get_doc_perms(sess)['can_view_docs'] == 1

    def can_download_docs(self, sess):
        return self._get_doc_perms(sess)['can_download_docs'] == 1

    def can_edit_docs(self, sess):
        return self._get_doc_perms(sess)['can_edit_docs'] == 1

    def can_delete_docs(self, sess):
        return self._get_doc_perms(sess)['can_delete_docs'] == 1

    def api_docs_my_permissions(self):
        """Tra ve quyen tai lieu cua user dang dang nhap"""
        sess = self.require_auth()
        if not sess: return
        self.send_json(self._get_doc_perms(sess))

    def api_doc_categories_list(self):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            rows = db.execute(
                'SELECT c.*, u.full_name as creator FROM doc_categories c LEFT JOIN users u ON c.created_by=u.id ORDER BY c.name'
            ).fetchall()
        self.send_json(rows_to_list(rows))

    def api_doc_categories_create(self):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        name      = body.get('name', '').strip()
        desc      = body.get('description', '').strip()
        color     = body.get('color', '#1B2A4A').strip()
        parent_id = body.get('parent_id') or None
        if parent_id: parent_id = int(parent_id)
        if not name:
            self.send_json({'error': 'Ten danh muc khong duoc de trong'}, 400); return
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO doc_categories (name, description, color, parent_id, created_by) VALUES (?,?,?,?,?)',
                (name, desc, color, parent_id, sess['userId'])
            )
            db.commit()
        self.send_json({'ok': True, 'id': cur.lastrowid})

    def api_doc_categories_update(self, cat_id):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        name      = body.get('name', '').strip()
        desc      = body.get('description', '').strip()
        color     = body.get('color', '#1B2A4A').strip()
        parent_id = body.get('parent_id') or None
        if parent_id: parent_id = int(parent_id)
        if parent_id == cat_id: parent_id = None  # prevent self-reference
        if not name:
            self.send_json({'error': 'Ten danh muc khong duoc de trong'}, 400); return
        with get_db() as db:
            db.execute('UPDATE doc_categories SET name=?, description=?, color=?, parent_id=? WHERE id=?',
                       (name, desc, color, parent_id, cat_id))
            db.commit()
        self.send_json({'ok': True})

    def api_doc_categories_delete(self, cat_id, qs=None):
        sess = self.require_manager()
        if not sess: return
        force = (qs or {}).get('force', [''])[0] == '1'
        with get_db() as db:
            child_cnt = db.execute(
                'SELECT COUNT(*) FROM doc_categories WHERE parent_id=?', (cat_id,)).fetchone()[0]
            if child_cnt > 0:
                self.send_json({'error': f'Danh muc co {child_cnt} danh muc con, vui long xoa con truoc'}, 400); return
            doc_cnt = db.execute(
                'SELECT COUNT(*) FROM documents WHERE category_id=?', (cat_id,)).fetchone()[0]
            if doc_cnt > 0 and not force:
                # Return warning — frontend will confirm
                self.send_json({'warning': True, 'doc_count': doc_cnt}); return
            if doc_cnt > 0:
                # Cascade: delete files on disk/R2, then delete DB rows
                rows = db.execute(
                    'SELECT file_path FROM documents WHERE category_id=?', (cat_id,)).fetchall()
                for row in rows:
                    try:
                        fp = row[0]
                        if fp.startswith('r2:'):
                            r2_delete(fp[3:])
                        else:
                            fpath = os.path.join(DATA_DIR, fp.lstrip('/'))
                            if os.path.exists(fpath):
                                os.remove(fpath)
                    except Exception:
                        pass
                db.execute('DELETE FROM documents WHERE category_id=?', (cat_id,))
            db.execute('DELETE FROM doc_categories WHERE id=?', (cat_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_docs_list(self, qs):
        sess = self.require_auth()
        if not sess: return
        if not self.can_view_docs(sess):
            self.send_json({'error': 'Ban khong co quyen xem tai lieu'}, 403); return
        cat_id      = qs.get('category_id', [None])[0]
        tag         = qs.get('tag', [None])[0]
        q           = qs.get('q', [None])[0]
        doc_type_id = qs.get('doc_type_id', [None])[0]
        issuer_f    = qs.get('issuer', [None])[0]
        sql = """SELECT d.*, u.full_name as uploader_name,
                        c.name as category_name, c.color as category_color,
                        t.name as doc_type_name
                 FROM documents d
                 LEFT JOIN users u ON d.uploaded_by = u.id
                 LEFT JOIN doc_categories c ON d.category_id = c.id
                 LEFT JOIN doc_types t ON d.doc_type_id = t.id
                 WHERE 1=1"""
        params = []
        if cat_id:
            # Lấy toàn bộ cây con (N cấp) bằng CTE đệ quy
            sql += (
                ' AND d.category_id IN ('
                '  WITH RECURSIVE sub(id) AS ('
                '    SELECT id FROM doc_categories WHERE id=?'
                '    UNION ALL'
                '    SELECT c.id FROM doc_categories c JOIN sub s ON c.parent_id=s.id'
                '  ) SELECT id FROM sub)'
            )
            params.append(int(cat_id))
        if doc_type_id:
            sql += ' AND d.doc_type_id=?'
            params.append(int(doc_type_id))
        if issuer_f:
            sql += ' AND d.issuer LIKE ?'
            params.append(f'%{issuer_f}%')
        if tag:
            sql += ' AND (d.tags LIKE ? OR d.tags LIKE ? OR d.tags LIKE ? OR d.tags = ?)'
            params += [f'{tag},%', f'%,{tag},%', f'%,{tag}', tag]
        if q:
            sql += ' AND (d.title LIKE ? OR d.description LIKE ? OR d.doc_number LIKE ? OR d.issuer LIKE ?)'
            params += [f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%']
        sql += ' ORDER BY d.issued_date DESC NULLS LAST, d.created_at DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json(rows_to_list(rows))

    def api_docs_extract_meta(self):
        """Extract doc_number, issued_date, issuer from an uploaded file via Claude."""
        sess = self.require_auth()
        if not sess: return
        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in ct:
            self.send_json({'error': 'multipart required'}, 400); return
        fields, files = self.read_multipart()
        if 'file' not in files:
            self.send_json({}, 200); return
        fitem = files['file']
        ext   = os.path.splitext(fitem.filename)[1].lower()
        file_bytes = fitem.file.read()
        if not os.environ.get('ANTHROPIC_API_KEY', ''):
            self.send_json({'_error': 'ANTHROPIC_API_KEY chua duoc cau hinh'}); return
        meta = call_claude_for_metadata(file_bytes, ext)
        if meta is None:
            self.send_json({'_error': 'Khong the trich xuat tu dong, vui long nhap tay'}); return
        self.send_json(meta)

    def api_docs_upload(self):
        sess = self.require_auth()
        if not sess: return
        if not self.can_upload_docs(sess):
            self.send_json({'error': 'Ban khong co quyen upload tai lieu'}, 403); return
        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' in ct:
            fields, files = self.read_multipart()
        else:
            self.send_json({'error': 'Yeu cau multipart/form-data'}, 400); return
        title       = fields.get('title', '').strip()
        desc        = fields.get('description', '').strip()
        cat_id      = fields.get('category_id', '').strip() or None
        tags        = fields.get('tags', '').strip()
        doc_number  = fields.get('doc_number', '').strip()
        issued_date = fields.get('issued_date', '').strip() or None
        issuer      = fields.get('issuer', '').strip()
        doc_type_id = fields.get('doc_type_id', '').strip() or None
        if not title:
            self.send_json({'error': 'Tieu de khong duoc de trong'}, 400); return
        if 'file' not in files:
            self.send_json({'error': 'Chua chon file'}, 400); return
        fitem = files['file']
        orig_name = fitem.filename
        ext = os.path.splitext(orig_name)[1].lower()
        allowed = {'.pdf': 'pdf', '.docx': 'docx', '.xlsx': 'xlsx',
                   '.jpg': 'image', '.jpeg': 'image', '.png': 'image'}
        if ext not in allowed:
            self.send_json({'error': 'Dinh dang khong ho tro (.pdf .docx .xlsx .jpg .png)'}, 400); return
        file_bytes = fitem.file.read()
        file_size  = len(file_bytes)
        ts = int(datetime.now(VN_TZ).timestamp() * 1000)
        safe_name = f'doc_{ts}_{sess["userId"]}{ext}'
        ct_map = {'.pdf':'application/pdf', '.docx':'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                  '.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  '.jpg':'image/jpeg', '.jpeg':'image/jpeg', '.png':'image/png'}
        content_type = ct_map.get(ext, 'application/octet-stream')
        if R2_ENABLED:
            ok = r2_upload(file_bytes, safe_name, content_type)
            if not ok:
                self.send_json({'error': 'Loi upload len cloud storage'}, 500); return
            file_path = f'r2:{safe_name}'
        else:
            fpath = os.path.join(UPLOADS_DIR, safe_name)
            with open(fpath, 'wb') as f:
                f.write(file_bytes)
            file_path = f'/uploads/{safe_name}'
        with get_db() as db:
            cur = db.execute(
                """INSERT INTO documents (title, description, file_path, file_name, file_type,
                   file_size, category_id, tags, uploaded_by,
                   doc_number, issued_date, issuer, doc_type_id)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (title, desc, file_path, orig_name,
                 allowed[ext], file_size, cat_id, tags, sess['userId'],
                 doc_number, issued_date, issuer, doc_type_id)
            )
            db.commit()
            new_id = cur.lastrowid
        self.send_json({'ok': True, 'id': new_id})

    def api_docs_update(self, doc_id):
        sess = self.require_auth()
        if not sess: return
        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' in ct:
            fields, files = self.read_multipart()
        else:
            fields = self.read_json()
            files = {}
        with get_db() as db:
            row = db.execute('SELECT * FROM documents WHERE id=?', (doc_id,)).fetchone()
            if not row:
                self.send_json({'error': 'Khong tim thay tai lieu'}, 404); return
            if not self.can_edit_docs(sess):
                self.send_json({'error': 'Ban khong co quyen chinh sua tai lieu'}, 403); return
            upd, params = [], []
            # Text fields
            for key, col in [('title','title'), ('description','description'),
                              ('doc_number','doc_number'), ('issuer','issuer'),
                              ('tags','tags'), ('notes','notes')]:
                if key in fields:
                    val = fields[key].strip() if isinstance(fields[key], str) else fields[key]
                    if key == 'title' and not val:
                        self.send_json({'error': 'Tieu de khong duoc de trong'}, 400); return
                    upd.append(f'{col}=?'); params.append(val)
            # Date field (nullable)
            if 'issued_date' in fields:
                val = (fields['issued_date'] or '').strip() or None
                upd.append('issued_date=?'); params.append(val)
            # FK fields (nullable int)
            for key, col in [('doc_type_id','doc_type_id'), ('category_id','category_id')]:
                if key in fields:
                    raw = fields[key]
                    try:
                        val = int(raw) if raw else None
                    except (ValueError, TypeError):
                        val = None
                    upd.append(f'{col}=?'); params.append(val)
            # File replacement
            if 'file' in files:
                fitem = files['file']
                orig_name = fitem.filename
                ext = os.path.splitext(orig_name)[1].lower()
                allowed = {'.pdf': 'pdf', '.docx': 'docx', '.xlsx': 'xlsx',
                           '.jpg': 'image', '.jpeg': 'image', '.png': 'image'}
                if ext not in allowed:
                    self.send_json({'error': 'Dinh dang khong ho tro (.pdf .docx .xlsx .jpg .png)'}, 400); return
                file_bytes = fitem.file.read()
                # No size limit — storage tracked separately
                ts = int(datetime.now(VN_TZ).timestamp() * 1000)
                safe_name = f'doc_{ts}_{sess["userId"]}{ext}'
                ct_map = {'.pdf':'application/pdf',
                          '.docx':'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                          '.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          '.jpg':'image/jpeg', '.jpeg':'image/jpeg', '.png':'image/png'}
                content_type = ct_map.get(ext, 'application/octet-stream')
                # Delete old file
                try:
                    old_fp = row['file_path']
                    if old_fp.startswith('r2:'):
                        r2_delete(old_fp[3:])
                    else:
                        fpath = os.path.join(DATA_DIR, old_fp.lstrip('/'))
                        if os.path.exists(fpath):
                            os.remove(fpath)
                except Exception:
                    pass
                # Save new file
                if R2_ENABLED:
                    ok = r2_upload(file_bytes, safe_name, content_type)
                    if not ok:
                        self.send_json({'error': 'Loi upload len cloud storage'}, 500); return
                    new_fp = f'r2:{safe_name}'
                else:
                    fpath = os.path.join(UPLOADS_DIR, safe_name)
                    with open(fpath, 'wb') as f:
                        f.write(file_bytes)
                    new_fp = f'/uploads/{safe_name}'
                upd.extend(['file_path=?', 'file_name=?', 'file_type=?', 'file_size=?'])
                params.extend([new_fp, orig_name, allowed[ext], len(file_bytes)])
            if not upd:
                self.send_json({'ok': True}); return
            params.append(doc_id)
            db.execute(f"UPDATE documents SET {', '.join(upd)} WHERE id=?", params)
            db.commit()
        self.send_json({'ok': True})

    def api_docs_delete(self, doc_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT * FROM documents WHERE id=?', (doc_id,)).fetchone()
            if not row:
                self.send_json({'error': 'Khong tim thay tai lieu'}, 404); return
            if not self.can_delete_docs(sess):
                self.send_json({'error': 'Ban khong co quyen xoa tai lieu'}, 403); return
            try:
                fp = row['file_path']
                if fp.startswith('r2:'):
                    r2_delete(fp[3:])
                else:
                    fpath = os.path.join(DATA_DIR, fp.lstrip('/'))
                    if os.path.exists(fpath):
                        os.remove(fpath)
            except Exception:
                pass
            db.execute('DELETE FROM documents WHERE id=?', (doc_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_docs_presigned_url(self, doc_id):
        sess = self.require_auth()
        if not sess: return
        if not self.can_download_docs(sess):
            self.send_json({'error': 'Ban khong co quyen tai ve tai lieu'}, 403); return
        with get_db() as db:
            row = db.execute('SELECT file_path, file_name FROM documents WHERE id=?', (doc_id,)).fetchone()
        if not row:
            self.send_json({'error': 'Khong tim thay tai lieu'}, 404); return
        fp = row['file_path']
        if fp.startswith('r2:'):
            url = r2_presigned_url(fp[3:], expires=3600)
            if not url:
                self.send_json({'error': 'Khong the tao URL'}, 500); return
        else:
            url = fp
        self.send_json({'url': url, 'file_name': row['file_name']})

    def api_docs_download(self, doc_id):
        """Proxy download: server lay file tu R2/local roi stream ve client voi Content-Disposition: attachment"""
        sess = self.require_auth()
        if not sess: return
        if not self.can_download_docs(sess):
            self.send_json({'error': 'Ban khong co quyen tai ve tai lieu'}, 403); return
        with get_db() as db:
            row = db.execute('SELECT file_path, file_name, file_type FROM documents WHERE id=?', (doc_id,)).fetchone()
        if not row:
            self.send_json({'error': 'Khong tim thay tai lieu'}, 404); return
        fp        = row['file_path']
        file_name = row['file_name'] or f'document_{doc_id}'
        ext       = os.path.splitext(file_name)[1].lower()
        ct_map    = {'.pdf':'application/pdf',
                     '.docx':'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                     '.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     '.jpg':'image/jpeg', '.jpeg':'image/jpeg', '.png':'image/png'}
        content_type = ct_map.get(ext, 'application/octet-stream')
        try:
            if fp.startswith('r2:'):
                client = get_r2_client()
                obj = client.get_object(Bucket=R2_BUCKET, Key=fp[3:])
                data = obj['Body'].read()
            else:
                fpath = os.path.join(DATA_DIR, fp.lstrip('/'))
                with open(fpath, 'rb') as f:
                    data = f.read()
        except Exception as e:
            self.send_json({'error': f'Khong the doc file: {e}'}, 500); return
        # Encode filename for Content-Disposition (RFC 5987)
        self.send_response(200)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Content-Disposition',
                         f"attachment; filename*=UTF-8''{urllib.parse.quote(file_name)}")
        self.send_header('Cache-Control', 'no-cache')
        self.end_headers()
        self.wfile.write(data)

    # --- DOC STORAGE ---

    def api_docs_storage(self):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute(
                'SELECT COALESCE(SUM(file_size),0) as total_bytes, COUNT(*) as file_count FROM documents'
            ).fetchone()
        total_bytes = row['total_bytes']
        file_count  = row['file_count']
        limit_bytes = 1 * 1024 * 1024 * 1024 * 1024  # 1 TB
        self.send_json({
            'used_bytes':  total_bytes,
            'limit_bytes': limit_bytes,
            'file_count':  file_count,
            'percent':     round(total_bytes / limit_bytes * 100, 4)
        })

    # --- DOC TYPES ---

    def api_doc_types_list(self):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            rows = db.execute('SELECT * FROM doc_types ORDER BY name').fetchall()
        self.send_json(rows_to_list(rows))

    def api_doc_types_create(self):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        name = body.get('name', '').strip()
        if not name:
            self.send_json({'error': 'Ten loai van ban khong duoc de trong'}, 400); return
        with get_db() as db:
            db.execute('INSERT INTO doc_types (name, description, created_by) VALUES (?,?,?)',
                       (name, body.get('description', '').strip(), sess['userId']))
            db.commit()
        self.send_json({'ok': True})

    def api_doc_types_update(self, type_id):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        name = body.get('name', '').strip()
        if not name:
            self.send_json({'error': 'Ten loai van ban khong duoc de trong'}, 400); return
        with get_db() as db:
            db.execute('UPDATE doc_types SET name=?, description=? WHERE id=?',
                       (name, body.get('description', '').strip(), type_id))
            db.commit()
        self.send_json({'ok': True})

    def api_doc_types_delete(self, type_id):
        sess = self.require_manager()
        if not sess: return
        with get_db() as db:
            cnt = db.execute('SELECT COUNT(*) FROM documents WHERE doc_type_id=?', (type_id,)).fetchone()[0]
            if cnt > 0:
                self.send_json({'error': f'Loai van ban co {cnt} tai lieu, khong the xoa'}, 400); return
            db.execute('DELETE FROM doc_types WHERE id=?', (type_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_manager_doc_permissions(self):
        sess = self.require_manager()
        if not sess: return
        with get_db() as db:
            rows = db.execute(
                """SELECT id, username, full_name, department,
                          can_upload_docs, can_view_docs, can_download_docs, can_edit_docs, can_delete_docs
                   FROM users WHERE role='employee' ORDER BY full_name"""
            ).fetchall()
        self.send_json(rows_to_list(rows))

    def api_manager_doc_permission_set(self, uid):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            db.execute(
                """UPDATE users SET
                   can_upload_docs=?, can_view_docs=?, can_download_docs=?, can_edit_docs=?, can_delete_docs=?
                   WHERE id=?""",
                (1 if body.get('can_upload_docs')   else 0,
                 1 if body.get('can_view_docs')     else 0,
                 1 if body.get('can_download_docs') else 0,
                 1 if body.get('can_edit_docs')     else 0,
                 1 if body.get('can_delete_docs')   else 0,
                 uid)
            )
            db.commit()
        self.send_json({'ok': True})

    # ══════════════════════════════════════════════════════════════════
    # BỒI THƯỜNG, HỖ TRỢ — API
    # ══════════════════════════════════════════════════════════════════

    # ── BT v2: CHỦ THỂ & NHÂN KHẨU ──────────────────────────────────
    def api_bt_parties_list(self, qs):
        sess = self.require_auth()
        if not sess: return
        q = qs.get('q', [''])[0].strip()
        sql = (
            'SELECT p.*, COUNT(DISTINCT m.id) as so_nhan_khau '
            'FROM bt_parties p LEFT JOIN bt_household_members m ON m.chu_the_id=p.id '
        )
        params = []
        if q:
            like = f'%{q}%'
            sql += 'WHERE p.ho_ten LIKE ? OR p.so_cccd LIKE ? OR p.so_dien_thoai LIKE ? '
            params += [like, like, like]
        sql += 'GROUP BY p.id ORDER BY p.id DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_parties_create(self):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        ho_ten = body.get('ho_ten', '').strip()
        if not ho_ten:
            self.send_json({'error': 'Họ tên không được để trống'}, 400); return
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_parties (loai_chu_the, ho_ten, gioi_tinh, ngay_sinh, so_cccd, '
                'ngay_cap_cccd, noi_cap_cccd, dia_chi_thuong_tru, so_dien_thoai, ghi_chu, custom_fields) '
                'VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (body.get('loai_chu_the', 'Cá nhân'), ho_ten, body.get('gioi_tinh', ''),
                 body.get('ngay_sinh') or None, body.get('so_cccd', ''), body.get('ngay_cap_cccd') or None,
                 body.get('noi_cap_cccd', ''), body.get('dia_chi_thuong_tru', ''), body.get('so_dien_thoai', ''),
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False))
            )
            db.commit()
        self.send_json({'ok': True, 'id': cur.lastrowid})

    def api_bt_parties_update(self, pid):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        ho_ten = body.get('ho_ten', '').strip()
        if not ho_ten:
            self.send_json({'error': 'Họ tên không được để trống'}, 400); return
        with get_db() as db:
            db.execute(
                'UPDATE bt_parties SET loai_chu_the=?, ho_ten=?, gioi_tinh=?, ngay_sinh=?, so_cccd=?, '
                'ngay_cap_cccd=?, noi_cap_cccd=?, dia_chi_thuong_tru=?, so_dien_thoai=?, ghi_chu=?, custom_fields=? '
                'WHERE id=?',
                (body.get('loai_chu_the', 'Cá nhân'), ho_ten, body.get('gioi_tinh', ''),
                 body.get('ngay_sinh') or None, body.get('so_cccd', ''), body.get('ngay_cap_cccd') or None,
                 body.get('noi_cap_cccd', ''), body.get('dia_chi_thuong_tru', ''), body.get('so_dien_thoai', ''),
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False), pid)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_bt_parties_delete(self, pid):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            # SQLite không tự cascade (foreign_keys pragma tắt mặc định trong get_db()),
            # nên xoá thủ công các bảng con tham chiếu tới chủ thể này.
            db.execute('DELETE FROM bt_household_members WHERE chu_the_id=?', (pid,))
            db.execute('DELETE FROM bt_parcel_owners WHERE chu_the_id=?', (pid,))
            db.execute('DELETE FROM bt_parties WHERE id=?', (pid,))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_members_list(self, party_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            rows = db.execute(
                'SELECT * FROM bt_household_members WHERE chu_the_id=? ORDER BY id', (party_id,)
            ).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_members_create(self, party_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        ho_ten = body.get('ho_ten', '').strip()
        if not ho_ten:
            self.send_json({'error': 'Họ tên không được để trống'}, 400); return
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_household_members (chu_the_id, ho_ten, quan_he_voi_chu_ho, ngay_sinh, so_cccd, ghi_chu, custom_fields) '
                'VALUES (?,?,?,?,?,?,?)',
                (party_id, ho_ten, body.get('quan_he_voi_chu_ho', ''), body.get('ngay_sinh') or None,
                 body.get('so_cccd', ''), body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False))
            )
            db.commit()
        self.send_json({'ok': True, 'id': cur.lastrowid})

    def api_bt_members_update(self, mid):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        ho_ten = body.get('ho_ten', '').strip()
        if not ho_ten:
            self.send_json({'error': 'Họ tên không được để trống'}, 400); return
        with get_db() as db:
            db.execute(
                'UPDATE bt_household_members SET ho_ten=?, quan_he_voi_chu_ho=?, ngay_sinh=?, so_cccd=?, ghi_chu=?, custom_fields=? '
                'WHERE id=?',
                (ho_ten, body.get('quan_he_voi_chu_ho', ''), body.get('ngay_sinh') or None,
                 body.get('so_cccd', ''), body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False), mid)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_bt_members_delete(self, mid):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            db.execute('DELETE FROM bt_household_members WHERE id=?', (mid,))
            db.commit()
        self.send_json({'ok': True})

    # ── BT v2: THỬA ĐẤT (gốc + theo dự án + đồng sở hữu) ────────────
    def api_bt_parcel_master_search(self, qs):
        sess = self.require_auth()
        if not sess: return
        q = qs.get('q', [''])[0].strip()
        sql = 'SELECT * FROM bt_parcel_master WHERE 1=1 '
        params = []
        if q:
            like = f'%{q}%'
            sql += 'AND (so_to_hien_hanh LIKE ? OR so_thua_hien_hanh LIKE ? OR dia_chi_vi_tri LIKE ?) '
            params += [like, like, like]
        sql += 'ORDER BY id DESC LIMIT 20'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_parcels_list(self, pid, qs):
        sess = self.require_auth()
        if not sess: return
        q = qs.get('q', [''])[0].strip()
        sql = (
            'SELECT pl.*, pm.so_to_hien_hanh, pm.so_thua_hien_hanh, '
            "(SELECT GROUP_CONCAT(pt.ho_ten, ', ') FROM bt_parcel_owners po "
            ' JOIN bt_parties pt ON pt.id=po.chu_the_id WHERE po.parcel_id=pl.id) as chu_so_huu '
            'FROM bt_parcels pl LEFT JOIN bt_parcel_master pm ON pm.id=pl.parcel_master_id '
            'WHERE pl.project_id=? '
        )
        params = [pid]
        if q:
            sql += 'AND (pl.so_to LIKE ? OR pl.so_thua LIKE ?) '
            like = f'%{q}%'
            params += [like, like]
        sql += 'ORDER BY pl.id DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_parcel_detail(self, parcel_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT * FROM bt_parcels WHERE id=?', (parcel_id,)).fetchone()
            if not row:
                self.send_json({'error': 'Không tìm thấy'}, 404); return
            owners = db.execute(
                'SELECT po.id as link_id, po.chu_the_id, po.vai_tro, po.ty_le_so_huu, pt.ho_ten '
                'FROM bt_parcel_owners po JOIN bt_parties pt ON pt.id=po.chu_the_id WHERE po.parcel_id=?',
                (parcel_id,)
            ).fetchall()
        d = dict(row)
        d['owners'] = [dict(o) for o in owners]
        self.send_json(d)

    def _save_parcel_owners(self, db, parcel_id, owners):
        db.execute('DELETE FROM bt_parcel_owners WHERE parcel_id=?', (parcel_id,))
        for o in (owners or []):
            chu_the_id = o.get('chu_the_id')
            if not chu_the_id:
                continue
            db.execute(
                'INSERT INTO bt_parcel_owners (parcel_id, chu_the_id, vai_tro, ty_le_so_huu) VALUES (?,?,?,?)',
                (parcel_id, chu_the_id, o.get('vai_tro', 'Đại diện đứng tên'), o.get('ty_le_so_huu'))
            )

    def api_bt_parcels_create(self, pid):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            master_id = body.get('parcel_master_id')
            if not master_id:
                cur = db.execute(
                    'INSERT INTO bt_parcel_master (so_to_hien_hanh, so_thua_hien_hanh, dia_chi_vi_tri) VALUES (?,?,?)',
                    (body.get('so_to', ''), body.get('so_thua', ''), body.get('dia_chi_vi_tri', ''))
                )
                master_id = cur.lastrowid
            cur = db.execute(
                'INSERT INTO bt_parcels (project_id, parcel_master_id, so_to, so_thua, loai_dat, '
                'tong_dien_tich, dien_tich_thu_hoi, dien_tich_con_lai, nguon_goc_su_dung, so_gcn, ngay_cap_gcn, '
                'ghi_chu, custom_fields) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
                (pid, master_id, body.get('so_to', ''), body.get('so_thua', ''), body.get('loai_dat', ''),
                 body.get('tong_dien_tich', 0), body.get('dien_tich_thu_hoi', 0), body.get('dien_tich_con_lai', 0),
                 body.get('nguon_goc_su_dung', ''), body.get('so_gcn', ''), body.get('ngay_cap_gcn') or None,
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False))
            )
            parcel_id = cur.lastrowid
            self._save_parcel_owners(db, parcel_id, body.get('owners'))
            db.commit()
        self.send_json({'ok': True, 'id': parcel_id, 'parcel_master_id': master_id})

    def api_bt_parcels_update(self, parcel_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            db.execute(
                'UPDATE bt_parcels SET so_to=?, so_thua=?, loai_dat=?, tong_dien_tich=?, dien_tich_thu_hoi=?, '
                'dien_tich_con_lai=?, nguon_goc_su_dung=?, so_gcn=?, ngay_cap_gcn=?, ghi_chu=?, custom_fields=? '
                'WHERE id=?',
                (body.get('so_to', ''), body.get('so_thua', ''), body.get('loai_dat', ''),
                 body.get('tong_dien_tich', 0), body.get('dien_tich_thu_hoi', 0), body.get('dien_tich_con_lai', 0),
                 body.get('nguon_goc_su_dung', ''), body.get('so_gcn', ''), body.get('ngay_cap_gcn') or None,
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False), parcel_id)
            )
            self._save_parcel_owners(db, parcel_id, body.get('owners'))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_parcels_delete(self, parcel_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            # Xoá thủ công bảng con vì foreign_keys pragma đang tắt (xem ghi chú get_db()).
            db.execute('DELETE FROM bt_parcel_owners WHERE parcel_id=?', (parcel_id,))
            db.execute('DELETE FROM bt_asset_parcels WHERE parcel_id=?', (parcel_id,))
            db.execute('DELETE FROM bt_parcel_decisions WHERE parcel_id=?', (parcel_id,))
            # Gỡ liên kết ở thửa trên Bản đồ GPMB (nếu có) để tránh tham chiếu treo,
            # xoá trực tiếp ở tab Thửa đất chỉ nên dùng cho thửa cũ/không gắn Bản đồ GPMB.
            db.execute('UPDATE bt_map_parcels SET parcel_id=NULL WHERE parcel_id=?', (parcel_id,))
            db.execute('DELETE FROM bt_parcels WHERE id=?', (parcel_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_project_gpmb_status(self, pid):
        """Kiểm tra dự án đã có Bản đồ GPMB với ít nhất 1 thửa hay chưa — dùng để khoá/mở khoá các tab khác."""
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            cnt = db.execute(
                "SELECT COUNT(*) FROM bt_map_parcels mp JOIN bt_maps m ON m.id=mp.map_id "
                "WHERE m.project_id=? AND m.loai_ban_do='Bản đồ GPMB' AND mp.parcel_id IS NOT NULL",
                (pid,)
            ).fetchone()[0]
        self.send_json({'has_gpmb': cnt > 0, 'parcel_count': cnt})

    # ── BT v2: BẢN ĐỒ (theo dự án hoặc độc lập, đính kèm nhiều thửa gốc + nhiều file) ──

    def api_bt_maps_list(self, qs):
        sess = self.require_auth()
        if not sess: return
        project_id = qs.get('project_id', [''])[0].strip()
        q = qs.get('q', [''])[0].strip()
        sql = ('SELECT m.*, COUNT(DISTINCT mp.id) as so_thua, COUNT(DISTINCT mf.id) as so_file '
               'FROM bt_maps m '
               'LEFT JOIN bt_map_parcels mp ON mp.map_id=m.id '
               'LEFT JOIN bt_map_files mf ON mf.map_id=m.id '
               'WHERE 1=1 ')
        params = []
        if project_id:
            sql += 'AND m.project_id=? '
            params.append(project_id)
        if q:
            like = f'%{q}%'
            sql += 'AND (m.ten_ban_do LIKE ? OR m.loai_ban_do LIKE ? OR m.don_vi_lap LIKE ? OR m.nhom_ban_do LIKE ?) '
            params += [like, like, like, like]
        sql += 'GROUP BY m.id ORDER BY m.id DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_map_detail(self, map_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT * FROM bt_maps WHERE id=?', (map_id,)).fetchone()
            if not row:
                self.send_json({'error': 'Không tìm thấy'}, 404); return
            parcels = db.execute(
                'SELECT mp.*, pm.so_to_hien_hanh, pm.so_thua_hien_hanh '
                'FROM bt_map_parcels mp LEFT JOIN bt_parcel_master pm ON pm.id=mp.parcel_master_id '
                'WHERE mp.map_id=? ORDER BY mp.id', (map_id,)
            ).fetchall()
            files = db.execute('SELECT * FROM bt_map_files WHERE map_id=? ORDER BY id', (map_id,)).fetchall()
        d = dict(row)
        d['parcels'] = [dict(p) for p in parcels]
        d['files'] = [dict(f) for f in files]
        self.send_json(d)

    def api_bt_maps_create(self):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        mocgioi_manh_ids = json.dumps(body.get('mocgioi_manh_ids') or [], ensure_ascii=False)
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_maps (project_id, loai_ban_do, ten_ban_do, nhom_ban_do, ngay_lap, don_vi_lap, ghi_chu, '
                'custom_fields, mocgioi_manh_ids, mocgioi_trong_la) VALUES (?,?,?,?,?,?,?,?,?,?)',
                (body.get('project_id') or None, body.get('loai_ban_do', ''), body.get('ten_ban_do', ''),
                 body.get('nhom_ban_do', ''), body.get('ngay_lap') or None, body.get('don_vi_lap', ''),
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False),
                 mocgioi_manh_ids, body.get('mocgioi_trong_la', ''))
            )
            db.commit()
            new_id = cur.lastrowid
        self.send_json({'ok': True, 'id': new_id})

    def api_bt_maps_update(self, map_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        mocgioi_manh_ids = json.dumps(body.get('mocgioi_manh_ids') or [], ensure_ascii=False)
        with get_db() as db:
            db.execute(
                'UPDATE bt_maps SET project_id=?, loai_ban_do=?, ten_ban_do=?, nhom_ban_do=?, ngay_lap=?, don_vi_lap=?, ghi_chu=?, '
                'custom_fields=?, mocgioi_manh_ids=?, mocgioi_trong_la=? WHERE id=?',
                (body.get('project_id') or None, body.get('loai_ban_do', ''), body.get('ten_ban_do', ''),
                 body.get('nhom_ban_do', ''), body.get('ngay_lap') or None, body.get('don_vi_lap', ''),
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False),
                 mocgioi_manh_ids, body.get('mocgioi_trong_la', ''), map_id)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_bt_maps_delete(self, map_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT loai_ban_do, mocgioi_manh_ids FROM bt_maps WHERE id=?', (map_id,)).fetchone()
            affected_manh_ids = []
            if row and row['loai_ban_do'] == 'Bản đồ mốc giới GPMB':
                # Bản đồ sắp xóa là Mốc giới GPMB — ghi lại các Mảnh trích đo GPMB nó từng liên kết
                # TRƯỚC KHI xóa, để sau khi xóa xong có thể rà soát lại diện tích thu hồi của các
                # thửa từng bị nó ảnh hưởng (theo yêu cầu của Minh: xóa Mốc giới GPMB phải tính lại).
                try:
                    affected_manh_ids = json.loads(row['mocgioi_manh_ids'] or '[]')
                except Exception:
                    affected_manh_ids = []

            # Xoá thủ công bảng con + file vật lý vì foreign_keys pragma đang tắt (xem ghi chú get_db()).
            frows = db.execute('SELECT file_path FROM bt_map_files WHERE map_id=?', (map_id,)).fetchall()
            for f in frows:
                try:
                    fp = f['file_path']
                    if fp.startswith('r2:'):
                        r2_delete(fp[3:])
                    else:
                        fpath = os.path.join(DATA_DIR, fp.lstrip('/'))
                        if os.path.exists(fpath):
                            os.remove(fpath)
                except Exception:
                    pass
            db.execute('DELETE FROM bt_map_files WHERE map_id=?', (map_id,))
            db.execute('DELETE FROM bt_map_parcels WHERE map_id=?', (map_id,))
            db.execute('DELETE FROM bt_maps WHERE id=?', (map_id,))
            db.commit()

            mocgioi_diffs = []
            if affected_manh_ids:
                try:
                    mocgioi_diffs = self._compute_mocgioi_recompute_for_manh_ids(db, affected_manh_ids)
                except RuntimeError:
                    # Thiếu shapely — việc xóa vẫn đã thành công, chỉ là không rà soát tự động được.
                    # Không chặn/lỗi hóa toàn bộ thao tác xóa chỉ vì bước rà soát phụ này thất bại.
                    mocgioi_diffs = []
        self.send_json({'ok': True, 'mocgioi_diffs': mocgioi_diffs, 'mocgioi_manh_ids': affected_manh_ids})

    def _create_parcel_from_map_entry(self, db, project_id, parcel_master_id, so_to, so_thua, tong_dt, thu_hoi_dt):
        """Bản đồ GPMB là nguồn gốc của Thửa đất: mỗi thửa thêm trên Bản đồ GPMB tự sinh 1 bản ghi bt_parcels."""
        con_lai = max((tong_dt or 0) - (thu_hoi_dt or 0), 0)
        cur = db.execute(
            'INSERT INTO bt_parcels (project_id, parcel_master_id, so_to, so_thua, tong_dien_tich, '
            'dien_tich_thu_hoi, dien_tich_con_lai) VALUES (?,?,?,?,?,?,?)',
            (project_id, parcel_master_id, so_to, so_thua, tong_dt or 0, thu_hoi_dt or 0, con_lai)
        )
        return cur.lastrowid

    def _sync_parcel_from_map_entry(self, db, parcel_id, parcel_master_id, so_to, so_thua, tong_dt, thu_hoi_dt):
        """Cập nhật lại các trường không gian của Thửa đất khi Bản đồ GPMB thay đổi.
        Chỉ đồng bộ số tờ/số thửa/diện tích — các trường do người dùng nhập ở tab Thửa đất
        (loại đất, GCN, chủ sử dụng...) được giữ nguyên."""
        con_lai = max((tong_dt or 0) - (thu_hoi_dt or 0), 0)
        db.execute(
            'UPDATE bt_parcels SET parcel_master_id=?, so_to=?, so_thua=?, tong_dien_tich=?, '
            'dien_tich_thu_hoi=?, dien_tich_con_lai=? WHERE id=?',
            (parcel_master_id, so_to, so_thua, tong_dt or 0, thu_hoi_dt or 0, con_lai, parcel_id)
        )

    def _compute_mocgioi_overlap(self, db, mocgioi_map_id):
        """Tính giao hình học giữa polygon(s) của 1 Bản đồ mốc giới GPMB và từng thửa trên các Mảnh
        trích đo GPMB đã khai liên kết (mocgioi_manh_ids), trả về danh sách thửa có diện tích thu hồi
        TÍNH MỚI khác với giá trị đang lưu (chênh > 1m² — bỏ qua sai số làm tròn vặt), kèm số cũ/mới.
        KHÔNG tự ghi đè — chỉ trả về để phía gọi (recompute-check) hiển thị cho người dùng xác nhận,
        hoặc (apply-recompute) áp dụng sau khi đã xác nhận. Dùng shapely — ném RuntimeError rõ ràng
        nếu thư viện chưa cài được trên môi trường đang chạy (không có sẵn trong dev sandbox lúc viết
        code này, chỉ cài qua requirements.txt lúc Railway build — xem ghi chú trong tài liệu thiết kế)."""
        try:
            from shapely.geometry import Polygon
            from shapely.ops import unary_union
        except ImportError:
            raise RuntimeError('Thiếu thư viện shapely trên server — không tính được diện tích giao. Liên hệ quản trị để cài đặt (requirements.txt đã khai, có thể cần deploy lại).')

        mgm = db.execute(
            'SELECT project_id, mocgioi_manh_ids, mocgioi_trong_la FROM bt_maps WHERE id=?',
            (mocgioi_map_id,)
        ).fetchone()
        if not mgm:
            raise ValueError('Không tìm thấy Bản đồ mốc giới GPMB')
        try:
            manh_ids = json.loads(mgm['mocgioi_manh_ids'] or '[]')
        except Exception:
            manh_ids = []
        trong_la = mgm['mocgioi_trong_la'] or ''
        if not manh_ids or trong_la not in ('thu_hoi', 'con_lai'):
            return []  # chưa khai đủ liên kết + hướng tính — chưa có gì để tính

        # Gom toàn bộ polygon của Bản đồ mốc giới GPMB thành 1 hình (có thể nhiều đoạn/nhiều xã)
        mg_rows = db.execute('SELECT toa_do FROM bt_map_parcels WHERE map_id=?', (mocgioi_map_id,)).fetchall()
        mg_polys = []
        for r in mg_rows:
            ring = self._ring_from_toa_do(r['toa_do'])
            if not ring:
                continue
            try:
                p = Polygon(ring)
                if p.is_valid and p.area > 0:
                    mg_polys.append(p)
            except Exception:
                continue
        if not mg_polys:
            return []
        mg_union = unary_union(mg_polys)

        diffs = []
        for manh_id in manh_ids:
            manh = db.execute('SELECT id, ten_ban_do FROM bt_maps WHERE id=?', (manh_id,)).fetchone()
            manh_ten = manh['ten_ban_do'] if manh else f'#{manh_id}'
            thua_rows = db.execute(
                'SELECT id, parcel_id, parcel_master_id, so_to_tren_ban_do, so_thua_tren_ban_do, '
                'dien_tich_tren_ban_do, dien_tich_thu_hoi_tren_ban_do, toa_do FROM bt_map_parcels WHERE map_id=?',
                (manh_id,)
            ).fetchall()
            for tr in thua_rows:
                ring = self._ring_from_toa_do(tr['toa_do'])
                if not ring:
                    continue
                try:
                    thua_poly = Polygon(ring)
                    if not thua_poly.is_valid or thua_poly.area <= 0:
                        continue
                    inter_area = thua_poly.intersection(mg_union).area
                except Exception:
                    continue
                tong_dt = tr['dien_tich_tren_ban_do'] or thua_poly.area
                if trong_la == 'thu_hoi':
                    new_thu_hoi = round(inter_area, 2)
                else:
                    new_thu_hoi = round(max(tong_dt - inter_area, 0), 2)
                cur_thu_hoi = round(tr['dien_tich_thu_hoi_tren_ban_do'] or 0, 2)
                if abs(new_thu_hoi - cur_thu_hoi) > 1.0:
                    diffs.append({
                        'link_id': tr['id'], 'manh_id': manh_id, 'manh_ten': manh_ten,
                        'so_to': tr['so_to_tren_ban_do'], 'so_thua': tr['so_thua_tren_ban_do'],
                        'dien_tich_thu_hoi_cu': cur_thu_hoi, 'dien_tich_thu_hoi_moi': new_thu_hoi,
                        'parcel_id': tr['parcel_id'], 'parcel_master_id': tr['parcel_master_id'],
                        'tong_dien_tich': tong_dt,
                    })
        return diffs

    def _compute_mocgioi_recompute_for_manh_ids(self, db, manh_ids):
        """Dùng khi 1 Bản đồ mốc giới GPMB VỪA BỊ XÓA — tính lại diện tích thu hồi cho các thửa trên
        những Mảnh trích đo GPMB từng bị nó ảnh hưởng. Không nhận map_id (bản đồ đã không còn tồn tại
        để tra cứu), mà nhận thẳng danh sách manh_ids đã ghi lại từ trước lúc xóa.
        Với mỗi mảnh: nếu còn Bản đồ mốc giới GPMB KHÁC vẫn liên kết tới nó, tính lại theo (các) bản đồ
        còn lại đó (tái sử dụng _compute_mocgioi_overlap, không viết lại công thức). Nếu KHÔNG còn bản
        đồ mốc giới nào bao phủ mảnh đó nữa, đề xuất đưa diện tích thu hồi (vốn được tính từ bản đồ vừa
        xóa) về 0 — vì không còn căn cứ hình học nào để giữ số cũ. Luôn trả về danh sách để người dùng
        xác nhận qua popup rà soát, không tự ghi CSDL (giống _compute_mocgioi_overlap)."""
        manh_ids = list(dict.fromkeys(manh_ids or []))  # loại trùng, giữ thứ tự
        if not manh_ids:
            return []

        all_diffs = []
        covered_manh = set()
        remaining_mocgioi = db.execute(
            "SELECT id, mocgioi_manh_ids FROM bt_maps WHERE loai_ban_do='Bản đồ mốc giới GPMB'"
        ).fetchall()
        for mg in remaining_mocgioi:
            try:
                ids = json.loads(mg['mocgioi_manh_ids'] or '[]')
            except Exception:
                ids = []
            relevant = [i for i in ids if i in manh_ids]
            if not relevant:
                continue
            covered_manh.update(relevant)
            diffs = self._compute_mocgioi_overlap(db, mg['id'])  # có thể ném RuntimeError nếu thiếu shapely
            all_diffs.extend([d for d in diffs if d['manh_id'] in manh_ids])

        # Mảnh không còn Bản đồ mốc giới GPMB nào bao phủ nữa — đề xuất về 0 nếu đang lưu > 0.
        uncovered = [m for m in manh_ids if m not in covered_manh]
        for manh_id in uncovered:
            manh = db.execute('SELECT id, ten_ban_do FROM bt_maps WHERE id=?', (manh_id,)).fetchone()
            manh_ten = manh['ten_ban_do'] if manh else f'#{manh_id}'
            rows = db.execute(
                'SELECT id, parcel_id, parcel_master_id, so_to_tren_ban_do, so_thua_tren_ban_do, '
                'dien_tich_tren_ban_do, dien_tich_thu_hoi_tren_ban_do FROM bt_map_parcels WHERE map_id=?',
                (manh_id,)
            ).fetchall()
            for r in rows:
                cur_thu_hoi = round(r['dien_tich_thu_hoi_tren_ban_do'] or 0, 2)
                if cur_thu_hoi > 1.0:
                    all_diffs.append({
                        'link_id': r['id'], 'manh_id': manh_id, 'manh_ten': manh_ten,
                        'so_to': r['so_to_tren_ban_do'], 'so_thua': r['so_thua_tren_ban_do'],
                        'dien_tich_thu_hoi_cu': cur_thu_hoi, 'dien_tich_thu_hoi_moi': 0.0,
                        'parcel_id': r['parcel_id'], 'parcel_master_id': r['parcel_master_id'],
                        'tong_dien_tich': r['dien_tich_tren_ban_do'] or 0,
                    })
        return all_diffs

    def api_bt_mocgioi_apply_recompute_for_manh(self):
        """Áp dụng rà soát diện tích thu hồi sau khi 1 Bản đồ mốc giới GPMB đã bị xóa — nhận
        {manh_ids: [...]} (client gửi lại đúng danh sách mảnh server đã trả về lúc xóa), TỰ TÍNH LẠI
        từ đầu (không tin số cũ) rồi áp dụng, cùng nguyên tắc an toàn như apply-recompute theo map_id."""
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        manh_ids = body.get('manh_ids') or []
        with get_db() as db:
            try:
                diffs = self._compute_mocgioi_recompute_for_manh_ids(db, manh_ids)
            except RuntimeError as e:
                self.send_json({'error': str(e)}, 500); return
            for d in diffs:
                db.execute(
                    'UPDATE bt_map_parcels SET dien_tich_thu_hoi_tren_ban_do=? WHERE id=?',
                    (d['dien_tich_thu_hoi_moi'], d['link_id'])
                )
                if d['parcel_id']:
                    self._sync_parcel_from_map_entry(
                        db, d['parcel_id'], d['parcel_master_id'], d['so_to'], d['so_thua'],
                        d['tong_dien_tich'], d['dien_tich_thu_hoi_moi']
                    )
            db.commit()
        self.send_json({'ok': True, 'updated': len(diffs)})

    def api_bt_mocgioi_recompute_check(self, map_id):
        """Tính thử (không ghi CSDL) chênh lệch diện tích thu hồi cho các thửa bị ảnh hưởng bởi 1
        Bản đồ mốc giới GPMB — dùng để hỏi xác nhận trước khi thực sự cập nhật (apply-recompute)."""
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            try:
                diffs = self._compute_mocgioi_overlap(db, map_id)
            except ValueError as e:
                self.send_json({'error': str(e)}, 404); return
            except RuntimeError as e:
                self.send_json({'error': str(e)}, 500); return
        self.send_json({'diffs': diffs})

    def api_bt_mocgioi_apply_recompute(self, map_id):
        """Áp dụng cập nhật diện tích thu hồi đã được người dùng xác nhận sau recompute-check.
        Tự tính lại từ đầu (không tin số phía client gửi lên) để tránh dữ liệu cũ/đua lệnh —
        diện tích thu hồi là số liệu ảnh hưởng tiền bồi thường nên ưu tiên an toàn hơn tiện lợi."""
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            try:
                diffs = self._compute_mocgioi_overlap(db, map_id)
            except ValueError as e:
                self.send_json({'error': str(e)}, 404); return
            except RuntimeError as e:
                self.send_json({'error': str(e)}, 500); return
            for d in diffs:
                db.execute(
                    'UPDATE bt_map_parcels SET dien_tich_thu_hoi_tren_ban_do=? WHERE id=?',
                    (d['dien_tich_thu_hoi_moi'], d['link_id'])
                )
                if d['parcel_id']:
                    self._sync_parcel_from_map_entry(
                        db, d['parcel_id'], d['parcel_master_id'], d['so_to'], d['so_thua'],
                        d['tong_dien_tich'], d['dien_tich_thu_hoi_moi']
                    )
            db.commit()
        self.send_json({'ok': True, 'updated': len(diffs)})

    def api_bt_map_parcels_add(self, map_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        so_to = body.get('so_to_tren_ban_do', '')
        so_thua = body.get('so_thua_tren_ban_do', '')
        dien_tich = body.get('dien_tich_tren_ban_do') or 0
        thu_hoi = body.get('dien_tich_thu_hoi_tren_ban_do') or 0
        master_id = body.get('parcel_master_id') or None
        toa_do = self._normalize_toa_do(body.get('toa_do'))
        ring = self._ring_from_toa_do(toa_do)
        if ring:
            dien_tich = round(self._polygon_area(ring), 2)
        with get_db() as db:
            m = db.execute('SELECT project_id, loai_ban_do FROM bt_maps WHERE id=?', (map_id,)).fetchone()
            parcel_id = None
            if m and m['loai_ban_do'] == 'Bản đồ GPMB' and m['project_id']:
                parcel_id = self._create_parcel_from_map_entry(
                    db, m['project_id'], master_id, so_to, so_thua, dien_tich, thu_hoi
                )
            cur = db.execute(
                'INSERT INTO bt_map_parcels (map_id, parcel_master_id, so_to_tren_ban_do, so_thua_tren_ban_do, '
                'dien_tich_tren_ban_do, dien_tich_thu_hoi_tren_ban_do, toa_do, parcel_id) VALUES (?,?,?,?,?,?,?,?)',
                (map_id, master_id, so_to, so_thua, dien_tich, thu_hoi, toa_do, parcel_id)
            )
            db.commit()
            new_id = cur.lastrowid
        self.send_json({'ok': True, 'id': new_id, 'parcel_id': parcel_id})

    def api_bt_map_parcels_update(self, link_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        so_to = body.get('so_to_tren_ban_do', '')
        so_thua = body.get('so_thua_tren_ban_do', '')
        dien_tich = body.get('dien_tich_tren_ban_do') or 0
        thu_hoi = body.get('dien_tich_thu_hoi_tren_ban_do') or 0
        master_id = body.get('parcel_master_id') or None
        toa_do = self._normalize_toa_do(body.get('toa_do'))
        ring = self._ring_from_toa_do(toa_do)
        if ring:
            dien_tich = round(self._polygon_area(ring), 2)
        with get_db() as db:
            row = db.execute('SELECT map_id, parcel_id FROM bt_map_parcels WHERE id=?', (link_id,)).fetchone()
            if not row:
                self.send_json({'error': 'Không tìm thấy'}, 404); return
            db.execute(
                'UPDATE bt_map_parcels SET parcel_master_id=?, so_to_tren_ban_do=?, so_thua_tren_ban_do=?, '
                'dien_tich_tren_ban_do=?, dien_tich_thu_hoi_tren_ban_do=?, toa_do=? WHERE id=?',
                (master_id, so_to, so_thua, dien_tich, thu_hoi, toa_do, link_id)
            )
            parcel_id = row['parcel_id']
            if parcel_id:
                self._sync_parcel_from_map_entry(db, parcel_id, master_id, so_to, so_thua, dien_tich, thu_hoi)
            else:
                # Thửa này chưa từng được sinh (vd. dữ liệu cũ) — sinh mới nếu bản đồ là GPMB thuộc 1 dự án.
                m = db.execute('SELECT project_id, loai_ban_do FROM bt_maps WHERE id=?', (row['map_id'],)).fetchone()
                if m and m['loai_ban_do'] == 'Bản đồ GPMB' and m['project_id']:
                    parcel_id = self._create_parcel_from_map_entry(
                        db, m['project_id'], master_id, so_to, so_thua, dien_tich, thu_hoi
                    )
                    db.execute('UPDATE bt_map_parcels SET parcel_id=? WHERE id=?', (parcel_id, link_id))
            db.commit()
        self.send_json({'ok': True, 'parcel_id': parcel_id})

    def api_bt_map_parcels_delete(self, link_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT parcel_id FROM bt_map_parcels WHERE id=?', (link_id,)).fetchone()
            db.execute('DELETE FROM bt_map_parcels WHERE id=?', (link_id,))
            if row and row['parcel_id']:
                # Xoá thửa khỏi Bản đồ GPMB nghĩa là xoá luôn Thửa đất tương ứng — xoá thủ công bảng con
                # vì foreign_keys pragma đang tắt (xem ghi chú get_db()).
                pid = row['parcel_id']
                db.execute('DELETE FROM bt_parcel_owners WHERE parcel_id=?', (pid,))
                db.execute('DELETE FROM bt_asset_parcels WHERE parcel_id=?', (pid,))
                db.execute('DELETE FROM bt_parcel_decisions WHERE parcel_id=?', (pid,))
                db.execute('DELETE FROM bt_parcels WHERE id=?', (pid,))
            db.commit()
        self.send_json({'ok': True})

    def _ring_from_toa_do(self, toa_do_json):
        """Lấy vành ngoài (outer ring) [[x,y],...] từ chuỗi GeoJSON đã chuẩn hóa lưu trong toa_do.
        Trả về None nếu không đọc được hoặc chưa đủ 3 đỉnh."""
        if not toa_do_json:
            return None
        try:
            geo = json.loads(toa_do_json)
            if geo.get('type') == 'Polygon':
                ring = geo['coordinates'][0]
            elif geo.get('type') == 'MultiPolygon':
                ring = geo['coordinates'][0][0]
            else:
                return None
            ring = [p for p in ring if isinstance(p, (list, tuple)) and len(p) >= 2]
            if len(ring) < 3:
                return None
            return ring
        except Exception:
            return None

    def _polygon_area(self, ring):
        """Diện tích đa giác theo công thức Shoelace (Gauss). ring: [[x,y],...], không cần tự khép kín trước."""
        pts = list(ring)
        if pts[0][0] != pts[-1][0] or pts[0][1] != pts[-1][1]:
            pts = pts + [pts[0]]
        s = 0.0
        for i in range(len(pts) - 1):
            x1, y1 = pts[i][0], pts[i][1]
            x2, y2 = pts[i + 1][0], pts[i + 1][1]
            s += x1 * y2 - x2 * y1
        return abs(s) / 2.0

    def _polygon_centroid(self, ring):
        """Tâm hình học (centroid) của đa giác — dùng để đặt nhãn thửa đúng vị trí tâm thực,
        khác với tâm hình chữ nhật bao quanh (bbox center) vốn có thể lệch ra ngoài với hình lõm."""
        pts = list(ring)
        if pts[0][0] != pts[-1][0] or pts[0][1] != pts[-1][1]:
            pts = pts + [pts[0]]
        a = cx = cy = 0.0
        for i in range(len(pts) - 1):
            x1, y1 = pts[i][0], pts[i][1]
            x2, y2 = pts[i + 1][0], pts[i + 1][1]
            cross = x1 * y2 - x2 * y1
            a += cross
            cx += (x1 + x2) * cross
            cy += (y1 + y2) * cross
        a = a / 2.0
        if abs(a) < 1e-9:
            xs = [p[0] for p in pts[:-1]]
            ys = [p[1] for p in pts[:-1]]
            return sum(xs) / len(xs), sum(ys) / len(ys)
        return cx / (6 * a), cy / (6 * a)

    def _nearest_neighbor_reorder(self, pts):
        """Sắp xếp lại danh sách đỉnh theo heuristic 'láng giềng gần nhất': bắt đầu từ đỉnh đầu tiên
        trong file, luôn nối tới đỉnh CHƯA dùng gần nhất về khoảng cách. Dùng cho trường hợp file dữ
        liệu đo đạc bị xáo trộn thứ tự (VD ranh giới GPMB xuất từ nhiều đợt đo, không theo đúng đường
        biên) — không đảm bảo tuyệt đối đúng với hình dạng lõm/phức tạp, chỉ là suy đoán hợp lý dựa
        trên việc các mốc liền kề trên thực địa thường ở gần nhau. Chỉ áp dụng khi người dùng chủ động
        bật tùy chọn — KHÔNG áp dụng mặc định vì dữ liệu vốn đã đúng thứ tự (trường hợp thông thường)
        không được đụng tới."""
        remaining = list(range(len(pts)))
        order = [remaining.pop(0)]
        while remaining:
            cx, cy = pts[order[-1]]
            best_i = min(range(len(remaining)),
                         key=lambda i: (pts[remaining[i]][0] - cx) ** 2 + (pts[remaining[i]][1] - cy) ** 2)
            order.append(remaining.pop(best_i))
        return [pts[i] for i in order]

    # ---- Quy đổi WGS84 (lon,lat, dùng cho import KML/KMZ) -> VN-2000 (X=Đông,Y=Bắc, dùng lưu CSDL) ----
    # Khớp CHÍNH XÁC với phép chiếu ngược dùng ở phía trình duyệt (vn2000ToLonLat/proj4 trong
    # boi-thuong.html, xem +towgs84=-191.90441429,-39.30318279,-111.45032835,-0.00928836,0.01975479,
    # -0.00427372,0.252906278) — viết thủ công bằng Python thuần (không dùng pyproj) vì môi trường
    # sandbox lúc phát triển không có mạng để cài, và pyproj cần biên dịch C-extension nên rủi ro
    # deploy Railway không chắc có sẵn wheel phù hợp; các hằng số hình học/dịch chuyển gốc tọa độ dưới
    # đây được kiểm chứng bằng test round-trip độc lập (VN-2000 -> WGS84 -> VN-2000, sai số < 1mm).
    _GEO_A = 6378137.0
    _GEO_F = 1.0 / 298.257223563
    _GEO_E2 = _GEO_F * (2 - _GEO_F)
    _TOWGS84_DX = -191.90441429
    _TOWGS84_DY = -39.30318279
    _TOWGS84_DZ = -111.45032835
    _TOWGS84_RX_ARCSEC = -0.00928836
    _TOWGS84_RY_ARCSEC = 0.01975479
    _TOWGS84_RZ_ARCSEC = -0.00427372
    _TOWGS84_S_PPM = 0.252906278

    def _geodetic_to_geocentric(self, lat, lon, h=0.0):
        import math
        a, e2 = self._GEO_A, self._GEO_E2
        sin_lat = math.sin(lat)
        n = a / math.sqrt(1 - e2 * sin_lat * sin_lat)
        x = (n + h) * math.cos(lat) * math.cos(lon)
        y = (n + h) * math.cos(lat) * math.sin(lon)
        z = (n * (1 - e2) + h) * sin_lat
        return x, y, z

    def _geocentric_to_geodetic(self, x, y, z):
        import math
        a, e2 = self._GEO_A, self._GEO_E2
        lon = math.atan2(y, x)
        p = math.sqrt(x * x + y * y)
        lat = math.atan2(z, p * (1 - e2))
        for _ in range(6):
            sin_lat = math.sin(lat)
            n = a / math.sqrt(1 - e2 * sin_lat * sin_lat)
            h = p / math.cos(lat) - n
            lat = math.atan2(z, p * (1 - e2 * n / (n + h)))
        return lat, lon

    def _helmert_wgs84_to_vn2000_src(self, x, y, z):
        """Đảo ngược 7-tham số Helmert (Bursa-Wolf, quy ước position-vector giống proj4 +towgs84):
        WGS84 (geocentric) -> hệ gốc VN-2000 (geocentric, cùng hình dạng ellipsoid WGS84). Dùng ma
        trận chuyển vị (R^T) làm nghịch đảo gần đúng của ma trận xoay góc nhỏ — sai số bậc 2 theo góc
        xoay (~1e-8 rad), thực tế không đáng kể (đã kiểm chứng round-trip dưới mm)."""
        import math
        rx = math.radians(self._TOWGS84_RX_ARCSEC / 3600.0)
        ry = math.radians(self._TOWGS84_RY_ARCSEC / 3600.0)
        rz = math.radians(self._TOWGS84_RZ_ARCSEC / 3600.0)
        s = self._TOWGS84_S_PPM * 1e-6
        dx = x - self._TOWGS84_DX
        dy = y - self._TOWGS84_DY
        dz = z - self._TOWGS84_DZ
        inv = 1.0 / (1 + s)
        xs = inv * (dx + rz * dy - ry * dz)
        ys = inv * (-rz * dx + dy + rx * dz)
        zs = inv * (ry * dx - rx * dy + dz)
        return xs, ys, zs

    def _forward_tmerc(self, lat, lon, lon0, k0=0.9999, x0=500000.0, y0=0.0):
        """Phép chiếu Transverse Mercator thuận (công thức Snyder) — lat/lon/lon0 tính bằng radian.
        Trả về (Đông, Bắc). Đây là hàm nghịch đảo của phép quy đổi Bắc/Đông->kinh/vĩ độ đang chạy
        phía trình duyệt qua proj4 (cùng tham số k0=0.9999, x0=500000, y0=0, ellipsoid WGS84)."""
        import math
        a, e2 = self._GEO_A, self._GEO_E2
        ep2 = e2 / (1 - e2)
        n = a / math.sqrt(1 - e2 * math.sin(lat) ** 2)
        t = math.tan(lat) ** 2
        c = ep2 * math.cos(lat) ** 2
        a_ = (lon - lon0) * math.cos(lat)
        m = a * (
            (1 - e2 / 4 - 3 * e2 ** 2 / 64 - 5 * e2 ** 3 / 256) * lat
            - (3 * e2 / 8 + 3 * e2 ** 2 / 32 + 45 * e2 ** 3 / 1024) * math.sin(2 * lat)
            + (15 * e2 ** 2 / 256 + 45 * e2 ** 3 / 1024) * math.sin(4 * lat)
            - (35 * e2 ** 3 / 3072) * math.sin(6 * lat)
        )
        x = x0 + k0 * n * (
            a_ + (1 - t + c) * a_ ** 3 / 6
            + (5 - 18 * t + t ** 2 + 72 * c - 58 * ep2) * a_ ** 5 / 120
        )
        y = y0 + k0 * (
            m + n * math.tan(lat) * (
                a_ ** 2 / 2 + (5 - t + 9 * c + 4 * c ** 2) * a_ ** 4 / 24
                + (61 - 58 * t + t ** 2 + 600 * c - 330 * ep2) * a_ ** 6 / 720
            )
        )
        return x, y

    def _wgs84_lonlat_to_vn2000(self, lon_deg, lat_deg, kinh_tuyen_truc_deg):
        """Điểm WGS84 (kinh độ, vĩ độ — hệ tọa độ mọi file KML/KMZ luôn dùng) -> VN-2000 (X=Đông,Y=Bắc),
        khớp đúng quy ước đang lưu trong CSDL của app. Dùng khi import ranh giới/thửa từ KML/KMZ."""
        import math
        lat_w = math.radians(lat_deg)
        lon_w = math.radians(lon_deg)
        xw, yw, zw = self._geodetic_to_geocentric(lat_w, lon_w, 0.0)
        xg, yg, zg = self._helmert_wgs84_to_vn2000_src(xw, yw, zw)
        lat_src, lon_src = self._geocentric_to_geodetic(xg, yg, zg)
        lon0 = math.radians(kinh_tuyen_truc_deg)
        return self._forward_tmerc(lat_src, lon_src, lon0)

    def _kml_local_tag(self, tag):
        return tag.split('}')[-1] if '}' in tag else tag

    def _kml_strip_diacritics(self, s):
        """Bỏ dấu tiếng Việt để so khớp tên Placemark không phụ thuộc dấu thanh/dấu mũ
        (VD 'Tờ'->'to', 'Thửa'->'thua') — chữ 'đ' xử lý riêng vì unicodedata không tách nó
        thành 'd' + dấu kết hợp như các nguyên âm có dấu."""
        s = s.replace('đ', 'd').replace('Đ', 'D')
        nfkd = unicodedata.normalize('NFD', s)
        return ''.join(c for c in nfkd if unicodedata.category(c) != 'Mn')

    def _kml_guess_so_to_so_thua(self, name, fallback_idx):
        """Đoán Tờ/Thửa từ tên Placemark KML theo vài mẫu thường gặp (không phân biệt dấu tiếng Việt).
        Không đoán được thì trả về nhóm riêng biệt (so_to rỗng, so_thua = tên gốc hoặc số thứ tự) để
        không gộp nhầm 2 hình khác nhau cùng không đoán được tên thành 1 nhóm trùng lặp."""
        name = (name or '').strip()
        if name:
            norm = self._kml_strip_diacritics(name).lower()
            m = re.search(r'\bto\s*[:.\-]?\s*(\d+)[^\d]+thua\s*[:.\-]?\s*(\d+)', norm)
            if m:
                return m.group(1), m.group(2)
            m = re.search(r'^\s*(\d+)\s*[.\-]\s*(\d+)\s*$', name)
            if m:
                return m.group(2), m.group(1)  # quy ước nhãn "Thửa.Tờ" đã dùng trong app
        return '', name if name else f'KML-{fallback_idx}'

    def _extract_kml_groups(self, file_bytes, filename):
        """Đọc file .kml hoặc .kmz (zip chứa .kml), trả về list các nhóm đỉnh dạng
        {'so_to', 'so_thua', 'pts_wgs84': [[lon,lat], ...]} — 1 Placemark (hoặc 1 hình con trong
        MultiGeometry) = 1 nhóm. Giữ NGUYÊN thứ tự đỉnh trong file — không cần tùy chọn auto-reorder
        như luồng Excel, vì phần mềm vẽ bản đồ (Google Earth, QGIS...) luôn xuất đỉnh theo đúng thứ
        tự đường vẽ trên thực địa/màn hình. Ưu tiên đọc Polygon (outerBoundaryIs, bỏ qua lỗ hổng bên
        trong nếu có); nếu Placemark không có Polygon thì thử LineString (nhiều người vẽ ranh giới
        bằng công cụ 'Đường' trong Google Earth thay vì 'Đa giác')."""
        local = self._kml_local_tag
        ext = (filename or '').lower()
        if ext.endswith('.kmz'):
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
                kml_name = next((n for n in zf.namelist() if n.lower().endswith('.kml')), None)
                if not kml_name:
                    raise ValueError('File .kmz không chứa tệp .kml nào bên trong')
                kml_bytes = zf.read(kml_name)
        else:
            kml_bytes = file_bytes

        root = ET.fromstring(kml_bytes)
        placemarks = [el for el in root.iter() if local(el.tag) == 'Placemark']
        if not placemarks:
            raise ValueError('Không tìm thấy Placemark nào trong file KML')

        groups = []
        idx = 0
        for pm in placemarks:
            name_el = next((el for el in pm if local(el.tag) == 'name'), None)
            name = name_el.text.strip() if name_el is not None and name_el.text else ''

            rings = []
            for poly in [el for el in pm.iter() if local(el.tag) == 'Polygon']:
                outer = next((el for el in poly.iter() if local(el.tag) == 'outerBoundaryIs'), None)
                if outer is None:
                    continue
                ring_el = next((el for el in outer.iter() if local(el.tag) == 'LinearRing'), None)
                if ring_el is None:
                    continue
                coord_el = next((el for el in ring_el.iter() if local(el.tag) == 'coordinates'), None)
                if coord_el is not None and coord_el.text:
                    rings.append(coord_el.text)

            if not rings:
                for line in [el for el in pm.iter() if local(el.tag) == 'LineString']:
                    coord_el = next((el for el in line.iter() if local(el.tag) == 'coordinates'), None)
                    if coord_el is not None and coord_el.text:
                        rings.append(coord_el.text)

            for ring_i, coord_text in enumerate(rings):
                pts = []
                for tok in coord_text.split():
                    parts = tok.strip().split(',')
                    if len(parts) < 2:
                        continue
                    try:
                        pts.append([float(parts[0]), float(parts[1])])
                    except ValueError:
                        continue
                if len(pts) < 3:
                    continue
                idx += 1
                so_to, so_thua = self._kml_guess_so_to_so_thua(name, idx)
                if len(rings) > 1:
                    so_thua = f"{so_thua}_{ring_i + 1}"
                groups.append({'so_to': so_to, 'so_thua': so_thua, 'pts_wgs84': pts})
        return groups

    def _normalize_toa_do(self, raw):
        """Chuẩn hóa ô nhập tọa độ thửa về GeoJSON, chấp nhận 2 kiểu nhập — dùng chung cho thêm/sửa 1 thửa
        (form tay) và import Excel hàng loạt, để không ai phải tự tay viết đúng cú pháp GeoJSON:
        1) Định dạng đơn giản 'X1,Y1; X2,Y2; X3,Y3; ...' — liệt kê các đỉnh theo thứ tự, tự động khép kín vòng.
        2) GeoJSON đầy đủ (bắt đầu bằng '{') — dành cho ai đã có sẵn dữ liệu chuẩn, giữ nguyên nếu hợp lệ."""
        raw = str(raw).strip() if raw is not None else ''
        if not raw:
            return None
        if raw.startswith('{') or raw.startswith('['):
            try:
                json.loads(raw)
                return raw
            except Exception:
                return None
        try:
            pts = []
            for pair in raw.split(';'):
                pair = pair.strip()
                if not pair:
                    continue
                x_str, y_str = pair.split(',')
                pts.append([float(x_str.strip()), float(y_str.strip())])
            if len(pts) < 3:
                return None
            if pts[0] != pts[-1]:
                pts.append(pts[0])
            return json.dumps({'type': 'Polygon', 'coordinates': [pts]}, ensure_ascii=False)
        except Exception:
            return None

    def api_bt_map_parcels_import(self, map_id):
        """Import hàng loạt thửa đất vào 1 mảnh trích đo Bản đồ, nhận 1 trong 3 định dạng file:
        - .xlsx: MỖI DÒNG LÀ 1 ĐỈNH THỬA (cột STT, Tờ, Thửa, X, Y — tọa độ VN-2000, X=Đông/Y=Bắc).
          Các dòng liên tiếp có cùng Tờ+Thửa là các đỉnh của cùng 1 thửa, nối theo thứ tự trong file.
          Field tùy chọn auto_reorder=1: sắp xếp lại thứ tự đỉnh trong mỗi nhóm theo láng giềng gần
          nhất trước khi khép kín — dùng khi nghi ngờ thứ tự đỉnh trong file bị xáo trộn. Mặc định TẮT.
        - .kml / .kmz: đọc mỗi Placemark (Polygon, hoặc LineString nếu không có Polygon) thành 1 nhóm
          đỉnh — tọa độ trong KML luôn ở hệ WGS84 (kinh độ, vĩ độ), tự động quy đổi sang VN-2000 theo
          Kinh tuyến trục của dự án (bắt buộc phải khai trước ở Sửa dự án). Không cần auto_reorder vì
          phần mềm vẽ bản đồ luôn xuất đỉnh đúng thứ tự đường vẽ. Tên Placemark được dò để đoán Tờ/Thửa
          (VD "Tờ 3 Thửa 45"); nếu không đoán được, mỗi Placemark thành 1 nhóm riêng dùng chính tên đó
          làm "số thửa" tạm (không sinh Thửa đất tự động trong trường hợp này nếu là Bản đồ GPMB, vì
          không có số tờ/số thửa thật — vẫn lưu được hình để xem, chỉ không đồng bộ sang bt_parcels).
        Ở cả 3 định dạng: đối chiếu trùng theo Tờ+Thửa trên cùng mảnh, yêu cầu xác nhận ghi đè
        (field confirm_overwrite=1) nếu có trùng — không âm thầm ghi đè dữ liệu đã có. Diện tích luôn
        TÍNH TỰ ĐỘNG bằng công thức Shoelace từ tọa độ các đỉnh sau khi khép kín."""
        sess = self.require_auth()
        if not sess: return
        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in ct:
            self.send_json({'error': 'multipart required'}, 400); return
        fields, files = self.read_multipart()
        if 'file' not in files:
            self.send_json({'error': 'Thiếu file'}, 400); return
        confirm_overwrite = fields.get('confirm_overwrite', '') == '1'
        auto_reorder = fields.get('auto_reorder', '') == '1'
        fitem = files['file']
        file_bytes = fitem.file.read()
        filename = fitem.filename or ''
        ext = os.path.splitext(filename)[1].lower()

        with get_db() as db:
            m = db.execute('SELECT project_id, loai_ban_do FROM bt_maps WHERE id=?', (map_id,)).fetchone()
            if not m:
                self.send_json({'error': 'Không tìm thấy Bản đồ'}, 404); return

            if ext in ('.kml', '.kmz'):
                ktt = None
                if m['project_id']:
                    proj = db.execute('SELECT kinh_tuyen_truc FROM bt_projects WHERE id=?', (m['project_id'],)).fetchone()
                    if proj and proj['kinh_tuyen_truc'] not in (None, ''):
                        ktt = proj['kinh_tuyen_truc']
                if ktt is None:
                    self.send_json({'error': 'Dự án chưa khai Kinh tuyến trục — cần thiết để quy đổi tọa độ '
                                              'KML/KMZ (hệ WGS84) sang VN-2000. Vào "Sửa dự án" để nhập trước '
                                              'khi import (tra theo tỉnh dự án tọa lạc, Phụ lục 2 Thông tư '
                                              '25/2014/TT-BTNMT, hoặc hỏi đơn vị đo đạc).'}, 400)
                    return
                try:
                    kml_groups = self._extract_kml_groups(file_bytes, filename)
                except Exception as e:
                    self.send_json({'error': f'Không đọc được file: {e}'}, 400); return

                parsed_rows = []
                skipped_short = []
                for g in kml_groups:
                    pts_wgs84 = g['pts_wgs84']
                    if len(pts_wgs84) > 1 and pts_wgs84[0] == pts_wgs84[-1]:
                        pts_wgs84 = pts_wgs84[:-1]
                    if len(pts_wgs84) < 3:
                        label = f"Tờ {g['so_to']} - Thửa {g['so_thua']}" if g['so_to'] else g['so_thua']
                        skipped_short.append(label)
                        continue
                    ring = [list(self._wgs84_lonlat_to_vn2000(lon, lat, ktt)) for lon, lat in pts_wgs84]
                    if ring[0][0] != ring[-1][0] or ring[0][1] != ring[-1][1]:
                        ring.append(ring[0])
                    toa_do = json.dumps({'type': 'Polygon', 'coordinates': [ring]}, ensure_ascii=False)
                    dien_tich = round(self._polygon_area(ring), 2)
                    parsed_rows.append({
                        'so_to': g['so_to'], 'so_thua': g['so_thua'],
                        'dien_tich': dien_tich, 'thu_hoi': 0, 'toa_do': toa_do,
                    })
                if not parsed_rows:
                    msg = 'Không có hình nào hợp lệ trong file KML/KMZ (mỗi Placemark cần ít nhất 3 đỉnh).'
                    if skipped_short:
                        msg += ' Bỏ qua: ' + ', '.join(skipped_short)
                    self.send_json({'error': msg}, 400); return
            else:
                if file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                    self.send_json({'error': 'File này đang ở định dạng Excel cũ (.xls). Hệ thống chỉ đọc được '
                                              '.xlsx — mở file bằng Excel rồi "Save As" sang định dạng .xlsx, '
                                              'sau đó import lại.'}, 400)
                    return
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                    ws = wb.active
                    raw_rows = list(ws.iter_rows(values_only=True))
                except Exception as e:
                    self.send_json({'error': f'Không đọc được file: {e}'}, 400); return

                if not raw_rows:
                    self.send_json({'error': 'File rỗng'}, 400); return
                header = [str(c).strip().lower() if c else '' for c in raw_rows[0]]

                def find(keywords):
                    for kw in keywords:
                        for i, v in enumerate(header):
                            if kw in v:
                                return i
                    return None

                COL = {
                    'stt':     find(['stt', 'số thứ tự']),
                    'so_to':   find(['tờ', 'to']),
                    'so_thua': find(['thửa', 'thua']),
                    'x':       find(['x']),
                    'y':       find(['y']),
                }
                if COL['so_to'] is None or COL['so_thua'] is None or COL['x'] is None or COL['y'] is None:
                    self.send_json({'error': 'Không tìm thấy đủ cột Tờ / Thửa / X / Y trong file. Vui lòng dùng đúng file mẫu.'}, 400)
                    return

                def cell(row, key):
                    idx = COL.get(key)
                    if idx is None or idx >= len(row):
                        return None
                    return row[idx]

                # Mỗi dòng là 1 đỉnh — gom các dòng LIÊN TIẾP có cùng (Tờ,Thửa) thành 1 thửa (1 nhóm đỉnh).
                groups = []
                cur = None
                for row in raw_rows[1:]:
                    if not any(row):
                        continue
                    so_to = cell(row, 'so_to')
                    so_thua = cell(row, 'so_thua')
                    so_to = str(so_to).strip() if so_to is not None else ''
                    so_thua = str(so_thua).strip() if so_thua is not None else ''
                    x_raw, y_raw = cell(row, 'x'), cell(row, 'y')
                    if not so_to and not so_thua:
                        continue
                    try:
                        x = float(x_raw)
                        y = float(y_raw)
                    except Exception:
                        continue
                    key = (so_to, so_thua)
                    if cur is None or cur['key'] != key:
                        cur = {'key': key, 'so_to': so_to, 'so_thua': so_thua, 'pts': []}
                        groups.append(cur)
                    cur['pts'].append([x, y])

                parsed_rows = []
                skipped_short = []
                for g in groups:
                    pts = g['pts']
                    if len(pts) < 3:
                        skipped_short.append(f"Tờ {g['so_to']} - Thửa {g['so_thua']}")
                        continue
                    if auto_reorder:
                        # Bỏ đỉnh khép kín trùng lặp (nếu có) trước khi sắp xếp lại, tránh thuật toán bị
                        # nhiễu bởi 1 điểm trùng khoảng cách 0 — sẽ tự khép kín lại đúng cách ngay bên dưới.
                        work = list(pts)
                        if len(work) > 1 and work[0][0] == work[-1][0] and work[0][1] == work[-1][1]:
                            work = work[:-1]
                        pts = self._nearest_neighbor_reorder(work)
                    ring = list(pts)
                    if ring[0][0] != ring[-1][0] or ring[0][1] != ring[-1][1]:
                        ring.append(ring[0])
                    toa_do = json.dumps({'type': 'Polygon', 'coordinates': [ring]}, ensure_ascii=False)
                    dien_tich = round(self._polygon_area(ring), 2)
                    parsed_rows.append({
                        'so_to': g['so_to'], 'so_thua': g['so_thua'],
                        'dien_tich': dien_tich, 'thu_hoi': 0, 'toa_do': toa_do,
                    })

                if not parsed_rows:
                    msg = 'Không có thửa hợp lệ trong file (mỗi thửa cần ít nhất 3 đỉnh với X, Y hợp lệ).'
                    if skipped_short:
                        msg += ' Bỏ qua vì thiếu đỉnh: ' + ', '.join(skipped_short)
                    self.send_json({'error': msg}, 400); return

            existing = db.execute(
                'SELECT id, so_to_tren_ban_do, so_thua_tren_ban_do, parcel_id FROM bt_map_parcels WHERE map_id=?',
                (map_id,)
            ).fetchall()
            existing_map = {}
            for e in existing:
                key = (str(e['so_to_tren_ban_do'] or '').strip().lower(), str(e['so_thua_tren_ban_do'] or '').strip().lower())
                existing_map[key] = e

            conflicts = []
            for r in parsed_rows:
                key = (r['so_to'].lower(), r['so_thua'].lower())
                if key in existing_map:
                    conflicts.append(f"Tờ {r['so_to']} - Thửa {r['so_thua']}")

            if conflicts and not confirm_overwrite:
                self.send_json({'conflict': True, 'conflicts': conflicts, 'total': len(parsed_rows)}, 409)
                return

            created, updated = 0, 0
            for r in parsed_rows:
                key = (r['so_to'].lower(), r['so_thua'].lower())
                existing_row = existing_map.get(key)
                if existing_row:
                    link_id = existing_row['id']
                    pid = existing_row['parcel_id']
                    db.execute(
                        'UPDATE bt_map_parcels SET so_to_tren_ban_do=?, so_thua_tren_ban_do=?, dien_tich_tren_ban_do=?, '
                        'dien_tich_thu_hoi_tren_ban_do=?, toa_do=? WHERE id=?',
                        (r['so_to'], r['so_thua'], r['dien_tich'], r['thu_hoi'], r['toa_do'], link_id)
                    )
                    if pid:
                        self._sync_parcel_from_map_entry(db, pid, None, r['so_to'], r['so_thua'], r['dien_tich'], r['thu_hoi'])
                    elif m['loai_ban_do'] == 'Bản đồ GPMB' and m['project_id'] and r['so_to']:
                        # Chỉ tự sinh Thửa đất khi có số tờ thật — Placemark KML không đoán được Tờ/Thửa
                        # (so_to rỗng) chỉ là hình tham khảo, không đủ dữ liệu để coi là 1 thửa đất thật.
                        pid = self._create_parcel_from_map_entry(
                            db, m['project_id'], None, r['so_to'], r['so_thua'], r['dien_tich'], r['thu_hoi']
                        )
                        db.execute('UPDATE bt_map_parcels SET parcel_id=? WHERE id=?', (pid, link_id))
                    updated += 1
                else:
                    pid = None
                    if m['loai_ban_do'] == 'Bản đồ GPMB' and m['project_id'] and r['so_to']:
                        pid = self._create_parcel_from_map_entry(
                            db, m['project_id'], None, r['so_to'], r['so_thua'], r['dien_tich'], r['thu_hoi']
                        )
                    db.execute(
                        'INSERT INTO bt_map_parcels (map_id, parcel_master_id, so_to_tren_ban_do, so_thua_tren_ban_do, '
                        'dien_tich_tren_ban_do, dien_tich_thu_hoi_tren_ban_do, toa_do, parcel_id) VALUES (?,?,?,?,?,?,?,?)',
                        (map_id, None, r['so_to'], r['so_thua'], r['dien_tich'], r['thu_hoi'], r['toa_do'], pid)
                    )
                    created += 1
            db.commit()
        self.send_json({'ok': True, 'created': created, 'updated': updated})

    def api_bt_gpmb_template(self):
        """Trả về file Excel mẫu để import hàng loạt thửa vào Bản đồ GPMB."""
        sess = self.require_auth()
        if not sess: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.send_json({'error': 'openpyxl chưa cài'}, 500); return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dữ liệu'
        headers = ['STT', 'Tờ', 'Thửa', 'X', 'Y']
        hdr_fill = PatternFill('solid', start_color='1B2A4A')
        hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        # Ví dụ 2 thửa liền kề, mỗi dòng là 1 đỉnh — đỉnh cuối lặp lại đỉnh đầu để khép kín (không bắt buộc,
        # hệ thống tự khép kín nếu thiếu), minh họa đúng cấu trúc dữ liệu đo đạc thực tế.
        sample = [
            (1, '1', '4', 590943.981, 2313150.975),
            (2, '1', '4', 591014.629, 2313128.736),
            (3, '1', '4', 591009.393, 2313109.651),
            (4, '1', '4', 590938.240, 2313132.048),
            (5, '1', '4', 590943.981, 2313150.975),
            (6, '1', '5', 590938.240, 2313132.048),
            (7, '1', '5', 591009.393, 2313109.651),
            (8, '1', '5', 591003.867, 2313089.506),
            (9, '1', '5', 590932.179, 2313112.071),
            (10, '1', '5', 590938.240, 2313132.048),
        ]
        for ri, row in enumerate(sample, 2):
            for ci, v in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=v)
        for ci, w in enumerate([8, 8, 8, 16, 16], 1):
            ws.column_dimensions[chr(64+ci)].width = w

        ws2 = wb.create_sheet('Hướng dẫn')
        guide = [
            ('Cột', 'Ý nghĩa'),
            ('STT', 'Số thứ tự đỉnh, chỉ để tham chiếu — hệ thống đọc theo đúng thứ tự các dòng trong file, không theo số này.'),
            ('Tờ / Thửa', 'Bắt buộc — theo hồ sơ địa chính.'),
            ('X, Y', 'Tọa độ VN-2000 của đỉnh thửa (mét), lấy đúng theo hồ sơ đo đạc/bản vẽ CAD. '
             'Thường X là tọa độ Đông (Easting), Y là tọa độ Bắc (Northing) — theo đúng thứ tự cột trong hồ sơ của anh, '
             'không cần đảo chỗ.'),
            ('Mỗi dòng là 1 đỉnh thửa',
             'KHÔNG phải mỗi dòng 1 thửa. Các dòng liên tiếp có cùng Tờ+Thửa là các đỉnh của cùng 1 thửa, '
             'nối theo đúng thứ tự trong file (dòng dưới nối dòng trên) để tạo thành các cạnh. '
             'Đỉnh cuối cùng của 1 thửa tự động nối lại đỉnh đầu tiên để khép kín thành đa giác — '
             'có thể lặp lại đỉnh đầu ở dòng cuối (như ví dụ) hoặc không, hệ thống đều tự xử lý đúng.'),
            ('Ví dụ trong file mẫu', 'Thửa 4 - Tờ 1 gồm 4 đỉnh (STT 1-4), dòng STT 5 lặp lại đỉnh đầu để khép kín. '
             'Thửa 5 - Tờ 1 bắt đầu từ dòng STT 6 (Tờ/Thửa đổi sang 1/5), gồm 4 đỉnh khép kín ở dòng STT 10.'),
            ('Diện tích thửa', 'KHÔNG cần nhập — hệ thống tự tính bằng công thức Shoelace từ tọa độ các đỉnh sau khi khép kín, '
             'đơn vị m². Nhãn thửa trên bản đồ sẽ hiện dạng "Thửa.Tờ.Diện tíchm2" tại đúng tâm hình học của thửa.'),
            ('Diện tích thu hồi', 'Không có trong file import (không tính được từ hình học) — nhập/sửa riêng sau khi import, '
             'bằng nút ✏️ sửa từng thửa.'),
            ('Nếu thửa đã tồn tại', 'Hệ thống nhận diện trùng theo Tờ + Thửa trong cùng mảnh trích đo Bản đồ GPMB, '
             'và sẽ hỏi xác nhận trước khi ghi đè — không tự động ghi đè.'),
        ]
        for ri, (a, b) in enumerate(guide, 1):
            ws2.cell(row=ri, column=1, value=a).font = Font(bold=(ri == 1))
            ws2.cell(row=ri, column=2, value=b).alignment = Alignment(wrap_text=True, vertical='top')
        ws2.column_dimensions['A'].width = 26
        ws2.column_dimensions['B'].width = 90

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', 'attachment; filename="mau-import-ban-do-gpmb.xlsx"')
        self.send_header('Content-Length', str(len(buf.getvalue())))
        self.end_headers()
        self.wfile.write(buf.getvalue())

    def api_bt_map_files_upload(self, map_id):
        sess = self.require_auth()
        if not sess: return
        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in ct:
            self.send_json({'error': 'multipart required'}, 400); return
        fields, files = self.read_multipart()
        if 'file' not in files:
            self.send_json({'error': 'Chưa chọn file'}, 400); return
        fitem = files['file']
        orig_name = fitem.filename
        ext = os.path.splitext(orig_name)[1].lower()
        allowed = {'.pdf': 'application/pdf',
                   '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                   '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                   '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png',
                   '.dwg': 'application/acad', '.dxf': 'application/dxf'}
        if ext not in allowed:
            self.send_json({'error': 'Định dạng không hỗ trợ (.pdf .docx .xlsx .jpg .png .dwg .dxf)'}, 400); return
        file_bytes = fitem.file.read()
        ts = int(datetime.now(VN_TZ).timestamp() * 1000)
        safe_name = f'map_{ts}_{sess["userId"]}{ext}'
        content_type = allowed[ext]
        if R2_ENABLED:
            ok = r2_upload(file_bytes, safe_name, content_type)
            if not ok:
                self.send_json({'error': 'Lỗi upload lên cloud storage'}, 500); return
            file_path = f'r2:{safe_name}'
        else:
            fpath = os.path.join(UPLOADS_DIR, safe_name)
            with open(fpath, 'wb') as f:
                f.write(file_bytes)
            file_path = f'/uploads/{safe_name}'
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_map_files (map_id, file_path, file_name) VALUES (?,?,?)',
                (map_id, file_path, orig_name)
            )
            db.commit()
            new_id = cur.lastrowid
        self.send_json({'ok': True, 'id': new_id})

    def api_bt_map_files_delete(self, file_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT file_path FROM bt_map_files WHERE id=?', (file_id,)).fetchone()
            if row:
                try:
                    fp = row['file_path']
                    if fp.startswith('r2:'):
                        r2_delete(fp[3:])
                    else:
                        fpath = os.path.join(DATA_DIR, fp.lstrip('/'))
                        if os.path.exists(fpath):
                            os.remove(fpath)
                except Exception:
                    pass
            db.execute('DELETE FROM bt_map_files WHERE id=?', (file_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_map_files_download(self, file_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT file_path, file_name FROM bt_map_files WHERE id=?', (file_id,)).fetchone()
        if not row:
            self.send_json({'error': 'Không tìm thấy'}, 404); return
        fp = row['file_path']
        file_name = row['file_name'] or f'map_file_{file_id}'
        ext = os.path.splitext(file_name)[1].lower()
        ct_map = {'.pdf':'application/pdf',
                  '.docx':'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                  '.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  '.jpg':'image/jpeg', '.jpeg':'image/jpeg', '.png':'image/png',
                  '.dwg':'application/acad', '.dxf':'application/dxf'}
        content_type = ct_map.get(ext, 'application/octet-stream')
        try:
            if fp.startswith('r2:'):
                client = get_r2_client()
                obj = client.get_object(Bucket=R2_BUCKET, Key=fp[3:])
                data = obj['Body'].read()
            else:
                fpath = os.path.join(DATA_DIR, fp.lstrip('/'))
                with open(fpath, 'rb') as f:
                    data = f.read()
        except Exception as e:
            self.send_json({'error': f'Không thể đọc file: {e}'}, 500); return
        self.send_response(200)
        self.send_header('Content-Type', content_type)
        self.send_header('Content-Length', str(len(data)))
        self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{urllib.parse.quote(file_name)}")
        self.send_header('Cache-Control', 'no-cache')
        self.end_headers()
        self.wfile.write(data)

    # ── BT v2: TÀI SẢN TRÊN ĐẤT (kiểm đếm khách quan, gắn nhiều thửa) ──

    def api_bt_assets_list(self, project_id, qs):
        sess = self.require_auth()
        if not sess: return
        q = qs.get('q', [''])[0].strip()
        sql = (
            'SELECT a.*, pt.ho_ten as chu_tai_san_ten, '
            "(SELECT GROUP_CONCAT(pl.so_to || '.' || pl.so_thua, ', ') "
            ' FROM bt_asset_parcels ap2 JOIN bt_parcels pl ON pl.id=ap2.parcel_id '
            ' WHERE ap2.asset_id=a.id) as thua_dat '
            'FROM bt_assets a '
            'JOIN bt_asset_parcels ap ON ap.asset_id=a.id '
            'JOIN bt_parcels p ON p.id=ap.parcel_id '
            'LEFT JOIN bt_parties pt ON pt.id=a.chu_tai_san_id '
            'WHERE p.project_id=? '
        )
        params = [project_id]
        if q:
            like = f'%{q}%'
            sql += 'AND (a.loai_tai_san_cu_the LIKE ? OR a.loai_tai_san_nhom LIKE ?) '
            params += [like, like]
        sql += 'GROUP BY a.id ORDER BY a.id DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_asset_detail(self, asset_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT * FROM bt_assets WHERE id=?', (asset_id,)).fetchone()
            if not row:
                self.send_json({'error': 'Không tìm thấy'}, 404); return
            parcels = db.execute(
                'SELECT ap.id as link_id, pl.id as parcel_id, pl.so_to, pl.so_thua '
                'FROM bt_asset_parcels ap JOIN bt_parcels pl ON pl.id=ap.parcel_id '
                'WHERE ap.asset_id=? ORDER BY ap.id', (asset_id,)
            ).fetchall()
        d = dict(row)
        d['parcels'] = [dict(p) for p in parcels]
        self.send_json(d)

    def _save_asset_parcels(self, db, asset_id, parcel_ids):
        db.execute('DELETE FROM bt_asset_parcels WHERE asset_id=?', (asset_id,))
        for pid in (parcel_ids or []):
            if not pid:
                continue
            db.execute('INSERT INTO bt_asset_parcels (asset_id, parcel_id) VALUES (?,?)', (asset_id, pid))

    def api_bt_assets_create(self, project_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_assets (loai_tai_san_nhom, loai_tai_san_cu_the, chu_tai_san_id, don_vi_tinh, '
                'so_luong_khoi_luong, thoi_diem_hinh_thanh, tinh_trang_phap_ly, ngay_kiem_dem, nguoi_kiem_dem, '
                'ghi_chu, custom_fields) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (body.get('loai_tai_san_nhom', ''), body.get('loai_tai_san_cu_the', ''),
                 body.get('chu_tai_san_id') or None, body.get('don_vi_tinh', ''),
                 body.get('so_luong_khoi_luong') or 0, body.get('thoi_diem_hinh_thanh') or None,
                 body.get('tinh_trang_phap_ly', 'Đúng quy định'), body.get('ngay_kiem_dem') or None,
                 body.get('nguoi_kiem_dem', ''), body.get('ghi_chu', ''),
                 json.dumps(body.get('custom_fields', {}), ensure_ascii=False))
            )
            asset_id = cur.lastrowid
            self._save_asset_parcels(db, asset_id, body.get('parcel_ids'))
            db.commit()
        self.send_json({'ok': True, 'id': asset_id})

    def api_bt_assets_update(self, asset_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            db.execute(
                'UPDATE bt_assets SET loai_tai_san_nhom=?, loai_tai_san_cu_the=?, chu_tai_san_id=?, don_vi_tinh=?, '
                'so_luong_khoi_luong=?, thoi_diem_hinh_thanh=?, tinh_trang_phap_ly=?, ngay_kiem_dem=?, '
                'nguoi_kiem_dem=?, ghi_chu=?, custom_fields=? WHERE id=?',
                (body.get('loai_tai_san_nhom', ''), body.get('loai_tai_san_cu_the', ''),
                 body.get('chu_tai_san_id') or None, body.get('don_vi_tinh', ''),
                 body.get('so_luong_khoi_luong') or 0, body.get('thoi_diem_hinh_thanh') or None,
                 body.get('tinh_trang_phap_ly', 'Đúng quy định'), body.get('ngay_kiem_dem') or None,
                 body.get('nguoi_kiem_dem', ''), body.get('ghi_chu', ''),
                 json.dumps(body.get('custom_fields', {}), ensure_ascii=False), asset_id)
            )
            self._save_asset_parcels(db, asset_id, body.get('parcel_ids'))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_assets_delete(self, asset_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            # Xoá thủ công bảng con vì foreign_keys pragma đang tắt (xem ghi chú get_db()).
            db.execute('DELETE FROM bt_asset_parcels WHERE asset_id=?', (asset_id,))
            db.execute('DELETE FROM bt_assets WHERE id=?', (asset_id,))
            db.commit()
        self.send_json({'ok': True})

    # ── BT v2: HỒ SƠ HỘ + QUYẾT ĐỊNH THEO THỬA ──────────────────────

    def api_bt_dossiers_list(self, project_id, qs):
        sess = self.require_auth()
        if not sess: return
        q = qs.get('q', [''])[0].strip()
        sql = (
            'SELECT d.*, pt.ho_ten as chu_ho_ten, '
            'COUNT(DISTINCT pd.id) as so_quyet_dinh, '
            'COALESCE(SUM(pd.tien_bt_dat),0) as tong_tien_bt_dat, '
            'COALESCE(SUM(pd.tien_bt_tai_san),0) as tong_tien_bt_tai_san '
            'FROM bt_dossiers d '
            'LEFT JOIN bt_parties pt ON pt.id=d.chu_the_id '
            'LEFT JOIN bt_parcel_decisions pd ON pd.ho_so_ho_id=d.id '
            'WHERE d.project_id=? '
        )
        params = [project_id]
        if q:
            sql += 'AND pt.ho_ten LIKE ? '
            params.append(f'%{q}%')
        sql += 'GROUP BY d.id ORDER BY d.id DESC'
        with get_db() as db:
            rows = db.execute(sql, params).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_dossier_detail(self, dossier_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute('SELECT * FROM bt_dossiers WHERE id=?', (dossier_id,)).fetchone()
            if not row:
                self.send_json({'error': 'Không tìm thấy'}, 404); return
            decisions = db.execute(
                'SELECT pd.*, pl.so_to, pl.so_thua '
                'FROM bt_parcel_decisions pd JOIN bt_parcels pl ON pl.id=pd.parcel_id '
                'WHERE pd.ho_so_ho_id=? ORDER BY pd.id', (dossier_id,)
            ).fetchall()
            persons = db.execute(
                'SELECT dp.*, COALESCE(pt.ho_ten, hm.ho_ten) as ho_ten '
                'FROM bt_dossier_persons dp '
                'LEFT JOIN bt_parties pt ON pt.id=dp.chu_the_id '
                'LEFT JOIN bt_household_members hm ON hm.id=dp.nhan_khau_id '
                'WHERE dp.ho_so_ho_id=? ORDER BY dp.id', (dossier_id,)
            ).fetchall()
        d = dict(row)
        d['decisions'] = [dict(x) for x in decisions]
        d['persons'] = [dict(x) for x in persons]
        self.send_json(d)

    def _save_dossier_decisions(self, db, dossier_id, decisions):
        db.execute('DELETE FROM bt_parcel_decisions WHERE ho_so_ho_id=?', (dossier_id,))
        for dec in (decisions or []):
            parcel_id = dec.get('parcel_id')
            if not parcel_id:
                continue
            db.execute(
                'INSERT INTO bt_parcel_decisions (ho_so_ho_id, parcel_id, so_quyet_dinh_pd, ngay_quyet_dinh_pd, '
                'tien_bt_dat, tien_bt_tai_san, trang_thai, ghi_chu) VALUES (?,?,?,?,?,?,?,?)',
                (dossier_id, parcel_id, dec.get('so_quyet_dinh_pd', ''), dec.get('ngay_quyet_dinh_pd') or None,
                 dec.get('tien_bt_dat') or 0, dec.get('tien_bt_tai_san') or 0,
                 dec.get('trang_thai', 'Chưa thực hiện'), dec.get('ghi_chu', ''))
            )

    def _save_dossier_persons(self, db, dossier_id, persons):
        db.execute('DELETE FROM bt_dossier_persons WHERE ho_so_ho_id=?', (dossier_id,))
        for p in (persons or []):
            chu_the_id = p.get('chu_the_id') or None
            nhan_khau_id = p.get('nhan_khau_id') or None
            if not chu_the_id and not nhan_khau_id:
                continue
            db.execute(
                'INSERT INTO bt_dossier_persons (ho_so_ho_id, chu_the_id, nhan_khau_id, la_doi_tuong_sxnn_truc_tiep, ghi_chu) '
                'VALUES (?,?,?,?,?)',
                (dossier_id, chu_the_id, nhan_khau_id, 1 if p.get('la_doi_tuong_sxnn_truc_tiep') else 0, p.get('ghi_chu', ''))
            )

    def api_bt_dossiers_create(self, project_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        chu_the_id = body.get('chu_the_id')
        if not chu_the_id:
            self.send_json({'error': 'Vui lòng chọn chủ hộ'}, 400); return
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_dossiers (project_id, chu_the_id, tien_thuong_tien_do, tien_ho_tro_on_dinh_doi_song, '
                'tien_ho_tro_dao_tao_chuyen_doi_nghe, tien_ho_tro_tai_dinh_cu, tien_ho_tro_khac, so_tien_tam_ung, '
                'so_tien_da_chi_tra, ngay_chi_tra, co_don_y_kien, noi_dung_y_kien, ghi_chu, custom_fields) '
                'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                (project_id, chu_the_id, body.get('tien_thuong_tien_do') or 0, body.get('tien_ho_tro_on_dinh_doi_song') or 0,
                 body.get('tien_ho_tro_dao_tao_chuyen_doi_nghe') or 0, body.get('tien_ho_tro_tai_dinh_cu') or 0,
                 body.get('tien_ho_tro_khac') or 0, body.get('so_tien_tam_ung') or 0, body.get('so_tien_da_chi_tra') or 0,
                 body.get('ngay_chi_tra') or None, 1 if body.get('co_don_y_kien') else 0, body.get('noi_dung_y_kien', ''),
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False))
            )
            dossier_id = cur.lastrowid
            self._save_dossier_decisions(db, dossier_id, body.get('decisions'))
            self._save_dossier_persons(db, dossier_id, body.get('persons'))
            db.commit()
        self.send_json({'ok': True, 'id': dossier_id})

    def api_bt_dossiers_update(self, dossier_id):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        chu_the_id = body.get('chu_the_id')
        if not chu_the_id:
            self.send_json({'error': 'Vui lòng chọn chủ hộ'}, 400); return
        with get_db() as db:
            db.execute(
                'UPDATE bt_dossiers SET chu_the_id=?, tien_thuong_tien_do=?, tien_ho_tro_on_dinh_doi_song=?, '
                'tien_ho_tro_dao_tao_chuyen_doi_nghe=?, tien_ho_tro_tai_dinh_cu=?, tien_ho_tro_khac=?, '
                'so_tien_tam_ung=?, so_tien_da_chi_tra=?, ngay_chi_tra=?, co_don_y_kien=?, noi_dung_y_kien=?, '
                'ghi_chu=?, custom_fields=?, updated_at=datetime(\'now\',\'localtime\') WHERE id=?',
                (chu_the_id, body.get('tien_thuong_tien_do') or 0, body.get('tien_ho_tro_on_dinh_doi_song') or 0,
                 body.get('tien_ho_tro_dao_tao_chuyen_doi_nghe') or 0, body.get('tien_ho_tro_tai_dinh_cu') or 0,
                 body.get('tien_ho_tro_khac') or 0, body.get('so_tien_tam_ung') or 0, body.get('so_tien_da_chi_tra') or 0,
                 body.get('ngay_chi_tra') or None, 1 if body.get('co_don_y_kien') else 0, body.get('noi_dung_y_kien', ''),
                 body.get('ghi_chu', ''), json.dumps(body.get('custom_fields', {}), ensure_ascii=False), dossier_id)
            )
            self._save_dossier_decisions(db, dossier_id, body.get('decisions'))
            self._save_dossier_persons(db, dossier_id, body.get('persons'))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_dossiers_delete(self, dossier_id):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            # Xoá thủ công bảng con vì foreign_keys pragma đang tắt (xem ghi chú get_db()).
            db.execute('DELETE FROM bt_parcel_decisions WHERE ho_so_ho_id=?', (dossier_id,))
            db.execute('DELETE FROM bt_dossier_persons WHERE ho_so_ho_id=?', (dossier_id,))
            db.execute('DELETE FROM bt_dossiers WHERE id=?', (dossier_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_projects_list(self):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            rows = db.execute(
                'SELECT p.*, COUNT(DISTINCT r.id) as so_ho_so FROM bt_projects p '
                'LEFT JOIN bt_parcels pl ON pl.project_id=p.id '
                'LEFT JOIN bt_records r ON r.parcel_id=pl.id '
                'GROUP BY p.id ORDER BY p.id DESC'
            ).fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_projects_create(self):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        name = body.get('name', '').strip()
        if not name:
            self.send_json({'error': 'Tên dự án không được để trống'}, 400); return
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_projects (name, mo_ta, dia_diem, chu_dau_tu, dien_tich_du_an, ranh_gioi_du_an, '
                'ngay_thong_bao_thu_hoi, ngay_bat_dau, ngay_ket_thuc_du_kien, custom_fields, kinh_tuyen_truc) '
                'VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (name, body.get('mo_ta',''), body.get('dia_diem',''), body.get('chu_dau_tu',''),
                 body.get('dien_tich_du_an') or 0, body.get('ranh_gioi_du_an') or None,
                 body.get('ngay_thong_bao_thu_hoi') or None, body.get('ngay_bat_dau') or None,
                 body.get('ngay_ket_thuc_du_kien') or None,
                 json.dumps(body.get('custom_fields', {}), ensure_ascii=False),
                 body.get('kinh_tuyen_truc') if body.get('kinh_tuyen_truc') not in ('', None) else None)
            )
            db.commit()
        self.send_json({'ok': True, 'id': cur.lastrowid})

    def api_bt_projects_update(self, pid):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            db.execute(
                'UPDATE bt_projects SET name=?, mo_ta=?, dia_diem=?, chu_dau_tu=?, dien_tich_du_an=?, '
                'ranh_gioi_du_an=?, ngay_thong_bao_thu_hoi=?, ngay_bat_dau=?, ngay_ket_thuc_du_kien=?, '
                'custom_fields=?, kinh_tuyen_truc=? WHERE id=?',
                (body.get('name',''), body.get('mo_ta',''), body.get('dia_diem',''), body.get('chu_dau_tu',''),
                 body.get('dien_tich_du_an') or 0, body.get('ranh_gioi_du_an') or None,
                 body.get('ngay_thong_bao_thu_hoi') or None, body.get('ngay_bat_dau') or None,
                 body.get('ngay_ket_thuc_du_kien') or None,
                 json.dumps(body.get('custom_fields', {}), ensure_ascii=False),
                 body.get('kinh_tuyen_truc') if body.get('kinh_tuyen_truc') not in ('', None) else None, pid)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_bt_projects_delete(self, pid):
        sess = self.require_manager()
        if not sess: return
        with get_db() as db:
            # Xoá thủ công toàn bộ dữ liệu con vì foreign_keys pragma đang tắt (xem ghi chú get_db()).
            parcel_ids = [r[0] for r in db.execute('SELECT id FROM bt_parcels WHERE project_id=?', (pid,)).fetchall()]
            if parcel_ids:
                qmarks = ','.join('?' * len(parcel_ids))
                db.execute(f'DELETE FROM bt_asset_parcels WHERE parcel_id IN ({qmarks})', parcel_ids)
                db.execute(f'DELETE FROM bt_parcel_owners WHERE parcel_id IN ({qmarks})', parcel_ids)
                db.execute(f'DELETE FROM bt_records WHERE parcel_id IN ({qmarks})', parcel_ids)
                db.execute(f'DELETE FROM bt_parcel_decisions WHERE parcel_id IN ({qmarks})', parcel_ids)
            db.execute('DELETE FROM bt_parcels WHERE project_id=?', (pid,))

            map_ids = [r[0] for r in db.execute('SELECT id FROM bt_maps WHERE project_id=?', (pid,)).fetchall()]
            if map_ids:
                qmarks = ','.join('?' * len(map_ids))
                frows = db.execute(f'SELECT file_path FROM bt_map_files WHERE map_id IN ({qmarks})', map_ids).fetchall()
                for f in frows:
                    try:
                        fp = f['file_path']
                        if fp.startswith('r2:'):
                            r2_delete(fp[3:])
                        else:
                            fpath = os.path.join(DATA_DIR, fp.lstrip('/'))
                            if os.path.exists(fpath):
                                os.remove(fpath)
                    except Exception:
                        pass
                db.execute(f'DELETE FROM bt_map_files WHERE map_id IN ({qmarks})', map_ids)
                db.execute(f'DELETE FROM bt_map_parcels WHERE map_id IN ({qmarks})', map_ids)
            db.execute('DELETE FROM bt_maps WHERE project_id=?', (pid,))

            dossier_ids = [r[0] for r in db.execute('SELECT id FROM bt_dossiers WHERE project_id=?', (pid,)).fetchall()]
            if dossier_ids:
                qmarks = ','.join('?' * len(dossier_ids))
                db.execute(f'DELETE FROM bt_parcel_decisions WHERE ho_so_ho_id IN ({qmarks})', dossier_ids)
                db.execute(f'DELETE FROM bt_dossier_persons WHERE ho_so_ho_id IN ({qmarks})', dossier_ids)
            db.execute('DELETE FROM bt_dossiers WHERE project_id=?', (pid,))

            db.execute('DELETE FROM bt_projects WHERE id=?', (pid,))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_project_stats(self, pid):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            # Tổng hồ sơ
            total = db.execute(
                'SELECT COUNT(*) FROM bt_parcels WHERE project_id=?', (pid,)
            ).fetchone()[0]
            # Theo trạng thái
            by_status = db.execute(
                'SELECT r.trang_thai, COUNT(*) as cnt, '
                'SUM(pl.dien_tich_thu_hoi) as tong_dt, '
                'SUM(r.tong_tien_btht) as tong_tien '
                'FROM bt_records r JOIN bt_parcels pl ON r.parcel_id=pl.id '
                'WHERE pl.project_id=? GROUP BY r.trang_thai', (pid,)
            ).fetchall()
            # Tổng tài chính
            finance = db.execute(
                'SELECT SUM(r.tien_bt_dat) as tien_dat, '
                'SUM(r.tien_bt_cay_trong) as tien_cay, '
                'SUM(r.tien_csht) as tien_csht, '
                'SUM(r.tien_ho_tro) as tien_ho_tro, '
                'SUM(r.tong_tien_btht) as tong_tien, '
                'SUM(r.so_tien_da_chi_tra) as da_chi_tra, '
                'SUM(pl.dien_tich_thu_hoi) as tong_dt_thu_hoi '
                'FROM bt_records r JOIN bt_parcels pl ON r.parcel_id=pl.id '
                'WHERE pl.project_id=?', (pid,)
            ).fetchone()
        self.send_json({
            'total': total,
            'by_status': [dict(r) for r in by_status],
            'finance': dict(finance) if finance else {}
        })

    def api_bt_records_list(self, pid, qs):
        sess = self.require_auth()
        if not sess: return
        search   = (qs.get('q',[''])[0] or '').strip()
        trang_thai = qs.get('trang_thai',[''])[0]
        loai_dat = qs.get('loai_dat',[''])[0]
        page     = max(1, int(qs.get('page',['1'])[0]))
        per_page = int(qs.get('per_page',['50'])[0])
        offset   = (page - 1) * per_page

        where = ['pl.project_id=?']
        params = [pid]
        if search:
            where.append('(o.ho_ten LIKE ? OR pl.so_to||"."||pl.so_thua LIKE ?)')
            params += [f'%{search}%', f'%{search}%']
        if trang_thai:
            where.append('r.trang_thai=?')
            params.append(trang_thai)
        if loai_dat:
            where.append('pl.loai_dat=?')
            params.append(loai_dat)

        w = ' AND '.join(where)
        base_sql = (
            'FROM bt_parcels pl '
            'LEFT JOIN bt_owners o ON o.id=pl.owner_id '
            'LEFT JOIN bt_records r ON r.parcel_id=pl.id '
            f'WHERE {w}'
        )
        with get_db() as db:
            total = db.execute(f'SELECT COUNT(*) {base_sql}', params).fetchone()[0]
            rows  = db.execute(
                f'SELECT pl.*, o.ho_ten, o.dia_chi_thuong_tru, o.so_cccd, o.so_dien_thoai, '
                f'r.id as record_id, r.trang_thai, r.tien_bt_dat, r.tien_bt_cay_trong, '
                f'r.tien_csht, r.tien_ho_tro, r.tong_tien_nyck, r.so_tien_tam_ung, '
                f'r.tong_tien_btht, r.so_tien_da_chi_tra, r.ngay_chi_tra, '
                f'r.so_quyet_dinh_pd, r.so_quyet_dinh_thd, r.ngay_quyet_dinh_pd, '
                f'r.dot_niem_yet_ck, r.co_don_y_kien, r.noi_dung_y_kien, r.ghi_chu, '
                f'r.updated_at {base_sql} '
                f'ORDER BY pl.so_to*1, pl.so_thua*1 '
                f'LIMIT ? OFFSET ?',
                params + [per_page, offset]
            ).fetchall()
        self.send_json({'total': total, 'page': page, 'per_page': per_page,
                        'records': [dict(r) for r in rows]})

    def api_bt_records_create(self, pid):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            # Upsert owner
            owner_id = body.get('owner_id')
            if not owner_id:
                ho_ten = body.get('ho_ten', '').strip()
                if ho_ten:
                    cur = db.execute(
                        'INSERT INTO bt_owners (ho_ten, dia_chi_thuong_tru, so_cccd, so_dien_thoai) '
                        'VALUES (?,?,?,?)',
                        (ho_ten, body.get('dia_chi_thuong_tru',''),
                         body.get('so_cccd',''), body.get('so_dien_thoai',''))
                    )
                    owner_id = cur.lastrowid
            # Tạo parcel
            cur = db.execute(
                'INSERT INTO bt_parcels (project_id,owner_id,so_to,so_thua,loai_dat,'
                'tong_dien_tich,dien_tich_thu_hoi,dien_tich_con_lai,dia_diem_thu_hoi,'
                'so_gcn,ngay_cap_gcn,ghi_chu) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                (pid, owner_id,
                 body.get('so_to',''), body.get('so_thua',''), body.get('loai_dat',''),
                 body.get('tong_dien_tich',0), body.get('dien_tich_thu_hoi',0),
                 body.get('dien_tich_con_lai',0), body.get('dia_diem_thu_hoi',''),
                 body.get('so_gcn',''), body.get('ngay_cap_gcn'), body.get('ghi_chu_thua',''))
            )
            parcel_id = cur.lastrowid
            # Tạo record bồi thường
            db.execute(
                'INSERT INTO bt_records (parcel_id,trang_thai,dot_niem_yet_ck,'
                'so_quyet_dinh_pd,ngay_quyet_dinh_pd,so_tien_pd,'
                'so_quyet_dinh_thd,ngay_quyet_dinh_thd,'
                'tien_bt_dat,tien_bt_cay_trong,tien_csht,tien_ho_tro,'
                'tong_tien_nyck,so_tien_tam_ung,tong_tien_btht,'
                'so_tien_da_chi_tra,ngay_chi_tra,co_don_y_kien,noi_dung_y_kien,ghi_chu) '
                'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                (parcel_id, body.get('trang_thai','Chưa thực hiện'),
                 body.get('dot_niem_yet_ck',''),
                 body.get('so_quyet_dinh_pd',''), body.get('ngay_quyet_dinh_pd'),
                 body.get('so_tien_pd',0),
                 body.get('so_quyet_dinh_thd',''), body.get('ngay_quyet_dinh_thd'),
                 body.get('tien_bt_dat',0), body.get('tien_bt_cay_trong',0),
                 body.get('tien_csht',0), body.get('tien_ho_tro',0),
                 body.get('tong_tien_nyck',0), body.get('so_tien_tam_ung',0),
                 body.get('tong_tien_btht',0),
                 body.get('so_tien_da_chi_tra',0), body.get('ngay_chi_tra'),
                 1 if body.get('co_don_y_kien') else 0,
                 body.get('noi_dung_y_kien',''), body.get('ghi_chu',''))
            )
            db.commit()
        self.send_json({'ok': True, 'parcel_id': parcel_id})

    def api_bt_records_update(self, rid):
        sess = self.require_auth()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            # Cập nhật owner
            row = db.execute(
                'SELECT pl.owner_id, pl.id as pid FROM bt_records r '
                'JOIN bt_parcels pl ON r.parcel_id=pl.id WHERE r.id=?', (rid,)
            ).fetchone()
            if not row:
                self.send_json({'error': 'Không tìm thấy'}, 404); return
            owner_id = row['owner_id']
            if owner_id:
                db.execute(
                    'UPDATE bt_owners SET ho_ten=?, dia_chi_thuong_tru=?, '
                    'so_cccd=?, so_dien_thoai=? WHERE id=?',
                    (body.get('ho_ten',''), body.get('dia_chi_thuong_tru',''),
                     body.get('so_cccd',''), body.get('so_dien_thoai',''), owner_id)
                )
            # Cập nhật parcel
            db.execute(
                'UPDATE bt_parcels SET so_to=?, so_thua=?, loai_dat=?, '
                'tong_dien_tich=?, dien_tich_thu_hoi=?, dien_tich_con_lai=?, '
                'dia_diem_thu_hoi=?, so_gcn=?, ngay_cap_gcn=?, ghi_chu=? '
                'WHERE id=?',
                (body.get('so_to',''), body.get('so_thua',''), body.get('loai_dat',''),
                 body.get('tong_dien_tich',0), body.get('dien_tich_thu_hoi',0),
                 body.get('dien_tich_con_lai',0), body.get('dia_diem_thu_hoi',''),
                 body.get('so_gcn',''), body.get('ngay_cap_gcn'), body.get('ghi_chu_thua',''),
                 row['pid'])
            )
            # Cập nhật record
            db.execute(
                'UPDATE bt_records SET trang_thai=?, dot_niem_yet_ck=?, '
                'so_quyet_dinh_pd=?, ngay_quyet_dinh_pd=?, so_tien_pd=?, '
                'so_quyet_dinh_thd=?, ngay_quyet_dinh_thd=?, '
                'tien_bt_dat=?, tien_bt_cay_trong=?, tien_csht=?, tien_ho_tro=?, '
                'tong_tien_nyck=?, so_tien_tam_ung=?, tong_tien_btht=?, '
                'so_tien_da_chi_tra=?, ngay_chi_tra=?, '
                'co_don_y_kien=?, noi_dung_y_kien=?, ghi_chu=?, '
                'updated_at=datetime("now","localtime") WHERE id=?',
                (body.get('trang_thai','Chưa thực hiện'),
                 body.get('dot_niem_yet_ck',''),
                 body.get('so_quyet_dinh_pd',''), body.get('ngay_quyet_dinh_pd'),
                 body.get('so_tien_pd',0),
                 body.get('so_quyet_dinh_thd',''), body.get('ngay_quyet_dinh_thd'),
                 body.get('tien_bt_dat',0), body.get('tien_bt_cay_trong',0),
                 body.get('tien_csht',0), body.get('tien_ho_tro',0),
                 body.get('tong_tien_nyck',0), body.get('so_tien_tam_ung',0),
                 body.get('tong_tien_btht',0),
                 body.get('so_tien_da_chi_tra',0), body.get('ngay_chi_tra'),
                 1 if body.get('co_don_y_kien') else 0,
                 body.get('noi_dung_y_kien',''), body.get('ghi_chu',''), rid)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_bt_records_delete(self, rid):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            row = db.execute(
                'SELECT pl.id FROM bt_records r JOIN bt_parcels pl ON r.parcel_id=pl.id WHERE r.id=?',
                (rid,)
            ).fetchone()
            # Xoá thủ công bảng con vì foreign_keys pragma đang tắt (xem ghi chú get_db()).
            db.execute('DELETE FROM bt_records WHERE id=?', (rid,))
            if row:
                parcel_id = row['id']
                db.execute('DELETE FROM bt_parcel_owners WHERE parcel_id=?', (parcel_id,))
                db.execute('DELETE FROM bt_asset_parcels WHERE parcel_id=?', (parcel_id,))
                db.execute('DELETE FROM bt_parcel_decisions WHERE parcel_id=?', (parcel_id,))
                db.execute('DELETE FROM bt_parcels WHERE id=?', (parcel_id,))
            db.commit()
        self.send_json({'ok': True})

    def api_bt_import_excel(self, pid):
        """Import hàng loạt từ file Excel theo cấu trúc chuẩn."""
        sess = self.require_auth()
        if not sess: return
        ct = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in ct:
            self.send_json({'error': 'multipart required'}, 400); return
        fields, files = self.read_multipart()
        if 'file' not in files:
            self.send_json({'error': 'Thiếu file'}, 400); return
        fitem = files['file']
        file_bytes = fitem.file.read()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            self.send_json({'error': f'Không đọc được file: {e}'}, 400); return

        # Tìm dòng header (có "Họ tên" hoặc "Số tờ")
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows):
            row_str = [str(c).lower() if c else '' for c in row]
            if any('họ tên' in c or 'ho ten' in c for c in row_str):
                header_row = row
                header_idx = i
                break

        if header_row is None:
            # Dùng mapping cố định cho file Book1.xlsx
            COL = {'ho_ten':3,'dia_chi':4,'so_to':5,'so_thua':6,
                   'tong_dt':7,'dt_thu_hoi':8,'dt_con_lai':12,'loai_dat':13,
                   'dia_diem':14,'dot_nyck':38,'tien_dat':53,'tien_cay':56,
                   'tien_csht':57,'tien_ho_tro':58,'tong_nyck':60,
                   'tam_ung':61,'tong_btht':68,'da_chi_tra':75,'ghi_chu':25}
            data_start = 4
        else:
            # Tự dò cột
            h = [str(c).strip().lower() if c else '' for c in header_row]
            def find(keywords):
                for kw in keywords:
                    for i,v in enumerate(h):
                        if kw in v: return i
                return None
            COL = {
                'ho_ten':    find(['họ tên','ho ten']),
                'dia_chi':   find(['địa chỉ thường','dia chi']),
                'so_to':     find(['số tờ','so to']),
                'so_thua':   find(['số thửa','so thua']),
                'tong_dt':   find(['tổng diện tích','tong dien tich']),
                'dt_thu_hoi':find(['s thu hồi','thu hoi']),
                'dt_con_lai':find(['s còn','con lai']),
                'loai_dat':  find(['loại đất','loai dat']),
                'dia_diem':  find(['địa chỉ thu hồi']),
                'tien_dat':  find(['về đất','tien dat']),
                'tien_cay':  find(['cây trồng','tien cay']),
                'tong_btht': find(['tổng tiền btht','tong tien']),
                'da_chi_tra':find(['tiền chi trả','chi tra']),
                'ghi_chu':   find(['note','ghi chú']),
            }
            data_start = header_idx + 1

        def safe(val, default=None):
            if val is None or str(val).strip() in ('', 'None', 'nan'): return default
            return val

        def safe_float(val):
            try: return float(str(val).replace(',','.'))
            except: return 0

        STATUS_MAP = [
            ('chi tra', 'Đã nhận tiền bồi thường, hỗ trợ'),
            ('quyet dinh thu hoi', 'Đã ban hành QĐ thu hồi đất'),
            ('phe duyet', 'Đã lập PA bồi thường, hỗ trợ'),
            ('niem yet', 'Đã kiểm kê / Niêm yết công khai'),
            ('hop dan', 'Đã họp dân'),
        ]

        def infer_status(row_data):
            """Suy luận trạng thái từ milestone cuối cùng có dữ liệu."""
            chi_tra = safe(row_data.get('da_chi_tra'))
            if chi_tra and safe_float(chi_tra) > 0:
                return 'Đã nhận tiền bồi thường, hỗ trợ'
            tong = safe(row_data.get('tong_btht'))
            if tong and safe_float(tong) > 0:
                return 'Đã lập PA bồi thường, hỗ trợ'
            nyck = safe(row_data.get('dot_nyck', row_data.get('tong_nyck')))
            if nyck:
                return 'Đã kiểm kê / Niêm yết công khai'
            return 'Chưa thực hiện'

        imported = 0
        skipped  = 0

        def g(row, key):
            idx = COL.get(key)
            if idx is None or idx >= len(row): return None
            return row[idx]

        with get_db() as db:
            for row in rows[data_start:]:
                if not any(row): continue
                ho_ten  = safe(g(row,'ho_ten'), '')
                so_to   = safe(g(row,'so_to'), '')
                so_thua = safe(g(row,'so_thua'), '')
                if not ho_ten and not so_to and not so_thua:
                    skipped += 1; continue
                # Owner
                owner_id = None
                if str(ho_ten).strip():
                    cur = db.execute(
                        'INSERT INTO bt_owners (ho_ten, dia_chi_thuong_tru) VALUES (?,?)',
                        (str(ho_ten).strip(), str(safe(g(row,'dia_chi'),'') or '').strip())
                    )
                    owner_id = cur.lastrowid
                # Parcel
                tong_dt    = safe_float(g(row,'tong_dt'))
                dt_thu_hoi = safe_float(g(row,'dt_thu_hoi'))
                dt_con_lai = safe_float(g(row,'dt_con_lai'))
                cur = db.execute(
                    'INSERT INTO bt_parcels (project_id,owner_id,so_to,so_thua,loai_dat,'
                    'tong_dien_tich,dien_tich_thu_hoi,dien_tich_con_lai,dia_diem_thu_hoi,ghi_chu) '
                    'VALUES (?,?,?,?,?,?,?,?,?,?)',
                    (pid, owner_id, str(so_to), str(so_thua),
                     str(safe(g(row,'loai_dat'),'') or ''),
                     tong_dt, dt_thu_hoi, dt_con_lai,
                     str(safe(g(row,'dia_diem'),'') or ''),
                     str(safe(g(row,'ghi_chu'),'') or ''))
                )
                parcel_id = cur.lastrowid
                # Record
                tong_btht    = safe_float(g(row,'tong_btht'))
                da_chi_tra   = safe_float(g(row,'da_chi_tra'))
                dot_nyck_val = safe(g(row,'dot_nyck'), '')
                trang_thai   = infer_status({
                    'da_chi_tra': da_chi_tra, 'tong_btht': tong_btht,
                    'dot_nyck': dot_nyck_val
                })
                db.execute(
                    'INSERT INTO bt_records (parcel_id,trang_thai,dot_niem_yet_ck,'
                    'tien_bt_dat,tien_bt_cay_trong,tien_csht,tien_ho_tro,'
                    'tong_tien_nyck,so_tien_tam_ung,tong_tien_btht,so_tien_da_chi_tra) '
                    'VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                    (parcel_id, trang_thai, str(dot_nyck_val or ''),
                     safe_float(g(row,'tien_dat')),
                     safe_float(g(row,'tien_cay')),
                     safe_float(g(row,'tien_csht') if COL.get('tien_csht') else 0),
                     safe_float(g(row,'tien_ho_tro') if COL.get('tien_ho_tro') else 0),
                     safe_float(g(row,'tong_nyck') if COL.get('tong_nyck') else 0),
                     safe_float(g(row,'tam_ung') if COL.get('tam_ung') else 0),
                     tong_btht, da_chi_tra)
                )
                imported += 1
            db.commit()
        self.send_json({'ok': True, 'imported': imported, 'skipped': skipped})

    def api_bt_export(self, pid):
        """Xuất Excel danh sách hồ sơ của 1 dự án."""
        sess = self.require_auth()
        if not sess: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.send_json({'error': 'openpyxl chưa cài'}, 500); return
        with get_db() as db:
            proj = db.execute('SELECT name FROM bt_projects WHERE id=?', (pid,)).fetchone()
            rows = db.execute(
                'SELECT pl.so_to, pl.so_thua, pl.loai_dat, pl.tong_dien_tich, '
                'pl.dien_tich_thu_hoi, pl.dia_diem_thu_hoi, '
                'o.ho_ten, o.dia_chi_thuong_tru, o.so_cccd, '
                'r.trang_thai, r.tien_bt_dat, r.tien_bt_cay_trong, r.tien_csht, '
                'r.tien_ho_tro, r.tong_tien_btht, r.so_tien_da_chi_tra, '
                'r.so_quyet_dinh_pd, r.so_quyet_dinh_thd, r.ngay_chi_tra, r.ghi_chu '
                'FROM bt_parcels pl '
                'LEFT JOIN bt_owners o ON o.id=pl.owner_id '
                'LEFT JOIN bt_records r ON r.parcel_id=pl.id '
                'WHERE pl.project_id=? ORDER BY pl.so_to*1, pl.so_thua*1', (pid,)
            ).fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Danh sách hồ sơ'
        headers = ['Số Tờ','Số Thửa','Loại đất','Tổng DT (m²)','DT thu hồi (m²)',
                   'Địa điểm','Họ tên CSD','Địa chỉ','CCCD','Trạng thái',
                   'Tiền BT đất','Tiền BT cây trồng','Tiền CSHT','Tiền hỗ trợ',
                   'Tổng tiền BTHT','Đã chi trả','Số QĐ PD','Số QĐ THĐ','Ngày chi trả','Ghi chú']
        hdr_fill = PatternFill('solid', start_color='1B2A4A')
        hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        proj_name = proj['name'] if proj else 'du-an'
        fname = f"boi-thuong-{proj_name}.xlsx".encode('ascii', errors='replace').decode()
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
        self.send_header('Content-Length', str(len(buf.getvalue())))
        self.end_headers()
        self.wfile.write(buf.getvalue())

    def api_bt_status_config_list(self):
        sess = self.require_auth()
        if not sess: return
        with get_db() as db:
            rows = db.execute('SELECT * FROM bt_status_config ORDER BY thu_tu').fetchall()
        self.send_json([dict(r) for r in rows])

    def api_bt_status_config_create(self):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            cur = db.execute(
                'INSERT INTO bt_status_config (ten, mau, thu_tu) VALUES (?,?,?)',
                (body.get('ten',''), body.get('mau','#95a5a6'), body.get('thu_tu',0))
            )
            db.commit()
        self.send_json({'ok': True, 'id': cur.lastrowid})

    def api_bt_status_config_update(self, sid):
        sess = self.require_manager()
        if not sess: return
        body = self.read_json()
        with get_db() as db:
            db.execute(
                'UPDATE bt_status_config SET ten=?, mau=?, thu_tu=? WHERE id=?',
                (body.get('ten',''), body.get('mau','#95a5a6'), body.get('thu_tu',0), sid)
            )
            db.commit()
        self.send_json({'ok': True})

    def api_bt_status_config_delete(self, sid):
        sess = self.require_manager()
        if not sess: return
        with get_db() as db:
            db.execute('DELETE FROM bt_status_config WHERE id=?', (sid,))
            db.commit()
        self.send_json({'ok': True})

def run():
    os.chdir(PUBLIC_DIR)
    init_db()
    if R2_ENABLED:
        threading.Thread(target=backup_scheduler_loop, daemon=True).start()
        print('[Backup] Da bat auto-backup cham_cong.db -> R2 (moi 24h)')
    else:
        print('[Backup] R2 chua bat nen KHONG co auto-backup. Xem cac bien R2_* tren Railway.')
    handler = Handler
    with socketserver.TCPServer(('', PORT), handler) as httpd:
        httpd.allow_reuse_address = True
        print(f'Server running on port {PORT}')
        print(f'DATA_DIR: {DATA_DIR}')
        print(f'R2 enabled: {R2_ENABLED}')
        httpd.serve_forever()

if __name__ == '__main__':
    run()
